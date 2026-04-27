from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import numpy_financial as npf
from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from Dashboard_LaaS_vs_NYTG.tier_dashboard_data import (  # noqa: E402
    build_tier_tables,
    tier_traceability_dict,
)
from Dashboard_LaaS_vs_NYTG.evidence_cn import cards_for_selected_tier  # noqa: E402
from Dashboard_LaaS_vs_NYTG.product_profiles import (  # noqa: E402
    ProductKey,
    capex_scale_vs_reference,
    forces_grid_electricity_zero,
    get_product_profile,
    routine_om_scale,
)


NAVY = "#1F3864"
MUTED = "#6B7280"
ACCENT = "#2563EB"
GREEN = "#16A34A"
RED = "#DC2626"

# Reduce metric font size so large numbers fit.
st.markdown(
    """
<style>
  div[data-testid="metric-container"] * { line-height: 1.05; }
  div[data-testid="stMetricValue"] { font-size: 1.35rem; }
  div[data-testid="stMetricLabel"] { font-size: 0.9rem; }
</style>
""",
    unsafe_allow_html=True,
)


def _money(x: float | None) -> str:
    if x is None or x != x:
        return "-"
    return f"{x:,.0f}"


def _pct(x: float | None) -> str:
    if x is None or x != x:
        return "-"
    return f"{x:.2%}"


def _safe_float(x: object, default: float = 0.0) -> float:
    try:
        v = float(x)  # type: ignore[arg-type]
        return v if v == v else default
    except Exception:
        return default


def _shift_electricity_payment(
    *,
    owner_spend_y: list[float],
    owner_save_y: list[float],
    elec_after_y: list[float],
    provider_pays_elec: bool,
) -> tuple[list[float], list[float]]:
    """
    Workbook default series assume owner pays electricity.
    If provider pays electricity (what-if), shift electricity outflow from owner to provider:
      - owner_spend decreases by elec_after
      - owner_net_savings increases by elec_after
    """
    if not provider_pays_elec:
        return owner_spend_y, owner_save_y
    spend = [float(owner_spend_y[i]) - float(elec_after_y[i]) for i in range(min(len(owner_spend_y), len(elec_after_y)))]
    save = [float(owner_save_y[i]) + float(elec_after_y[i]) for i in range(min(len(owner_save_y), len(elec_after_y)))]
    return spend, save


def _irr_yearly(flows_y0_to_y10: list[float]) -> float | None:
    try:
        r = float(npf.irr([float(x) for x in flows_y0_to_y10]))
        return r if r == r else None
    except Exception:
        return None


def _npv_yearly(*, disc_annual: float, flows_y0_to_yN: list[float]) -> float | None:
    """NPV using annual discounting (same convention as `numpy_financial.npv`)."""
    try:
        d = float(disc_annual)
        if d != d or d < 0:
            return None
        return float(npf.npv(d, [float(x) for x in flows_y0_to_yN]))
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def _load(new_models_dir: str, cache_bust: str) -> tuple[pd.DataFrame, pd.DataFrame, list[dict]]:
    df_sum, df_long, tiers = build_tier_tables(new_models_dir=Path(new_models_dir))
    tiers_dict = [tier_traceability_dict(t) for t in tiers]
    return df_sum, df_long, tiers_dict


def main() -> None:
    st.set_page_config(page_title="EMC → LaaS | 单方案查看", layout="wide", initial_sidebar_state="expanded")
    st.markdown(
        f"<h2 style='color:{NAVY};margin-bottom:0.2rem;'>能源托管/EMC → LaaS：单方案“工作簿查看器”（中文）</h2>"
        f"<p style='color:{MUTED};margin-top:0;'>选择一个方案后，本页直接读取该工作簿的缓存数值（与WPS打开看到一致），并以更直观的方式展示："
        f"<code>01_Dashboard</code> 核心KPI、<code>05_Annual_Model</code> 年度现金流/节省、以及 OPEX/CAPEX 结构与“为什么更好”的证据说明。</p>",
        unsafe_allow_html=True,
    )
    with st.expander("先看这个：谁付钱 / 钱从哪来（模型口径）", expanded=True):
        st.markdown(
            "- **业主（客户）支付**：年度服务费/订阅费（EMC 与 LaaS），以及 **电费**（默认业主承担）。\n"
            "- **服务商（Huapu/HPwinner）支付**：项目 **CAPEX（Y0一次性投入）**，以及 **运维OPEX**（人工/维修、平台、备件）。\n"
            "- **电费承担开关（工作簿）**：`02_Inputs!D25`（EMC）、`02_Inputs!D35`（LaaS）（`0=业主承担电费；1=服务商承担电费`）。\n"
            "- **本页 what-if**：侧边栏可分别切换 EMC / LaaS 的电费承担（不写回工作簿；工作簿默认仍以缓存值为准）。\n"
            "- **理解方式**：服务商的OPEX/CAPEX是“成本端现金流出”；资金来源主要来自业主付的服务费（收入端），两者差额形成利润/回收期。"
        )

    default_dir = str(_ROOT / "Dashboard_LaaS_vs_NYTG" / "new_models")
    with st.sidebar:
        st.subheader("数据源（来自 new_models）")
        new_models_dir = st.text_input("new_models folder", value=default_dir)
        view_mode = st.radio("展示模式", options=["只看关键结论（1分钟版）", "展开明细（投委会版）"], index=0)
        show_all_years_table = st.checkbox("显示年度表格（像Excel）", value=(view_mode != "只看关键结论（1分钟版）"))

    df_sum, df_long, tiers_dict = _load(new_models_dir, cache_bust="v12_authoritative_opex_inputs")

    if df_sum.empty:
        st.warning("No workbooks found. Check the folder path.")
        return

    # Single-tier selector
    tier_list = df_sum["tier"].tolist()
    sel = st.selectbox("选择方案（切换工作簿）", options=tier_list, index=0)

    dsel = df_sum[df_sum["tier"] == sel].iloc[0].to_dict()
    df_t = df_long[df_long["tier"] == sel].copy()

    # Product selector (what-if override). Default to workbook stored product_key.
    workbook_pk = str(dsel.get("product_key") or "AI_lightning_grid")
    product_options: list[ProductKey] = ["AI_lightning_grid", "AI_battery_integrated_grid", "AI_plus_solar_offgrid"]
    with st.sidebar:
        st.subheader("产品类型（可做what-if）")
        pk_sel: ProductKey = st.selectbox(
            "产品类型(product_key)",
            options=product_options,
            index=product_options.index(workbook_pk) if workbook_pk in product_options else 0,
        )
        product_overrides_active = bool(pk_sel != workbook_pk)
        if product_overrides_active:
            st.caption("假设覆盖（未写回工作簿）：图表/表格按当前产品类型重算；工作簿默认值仍可审计。")
        st.caption("切换产品会影响的关键单元格（工作簿内）")
        st.code(
            "02_Inputs!D48  产品类型(product_key)\n"
            "02_Inputs!D28  LaaS CAPEX/盏\n"
            "02_Inputs!D31  LaaS 节电率\n"
            "02_Inputs!D32:D34  LaaS 非电费运维（人工/平台/备件）",
            language="text",
        )
        st.caption("提示：如果你在WPS/Excel里改了 D48，要想让本页读取到变化，需要保存文件（且若未自动重算，则先触发重算）。")

        # Small explainer table so audience understands the implications.
        st.caption("产品差异说明（模型口径）")
        rows = []
        for pk in product_options:
            prof = get_product_profile(pk)
            rows.append(
                {
                    "product_key": pk,
                    "节电率下限(推导)": prof.implied_grid_saving_rate(),
                    "CAPEX缩放(相对AI_lightning_grid)": capex_scale_vs_reference(pk),
                    "非电费运维缩放(相对AI_lightning_grid)": routine_om_scale(pk),
                    "电费规则": "离网：电费≈0" if forces_grid_electricity_zero(pk) else "并网：按节电率计算",
                }
            )
        st.dataframe(pd.DataFrame(rows), use_container_width=True, height=160)

        st.subheader("电费承担（可做what-if）")
        emc_elec_bear = st.radio(
            "EMC 谁承担电费（用于本页重算展示）",
            options=["业主承担电费（默认）", "服务商承担电费（what-if）"],
            index=0,
            key="emc_elec_bearer",
        )
        laas_elec_bear = st.radio(
            "LaaS 谁承担电费（用于本页重算展示）",
            options=["业主承担电费（默认）", "服务商承担电费（what-if）"],
            index=0,
            key="laas_elec_bearer",
        )
        emc_provider_pays_elec = emc_elec_bear.startswith("服务商")
        laas_provider_pays_elec = laas_elec_bear.startswith("服务商")
        if emc_provider_pays_elec or laas_provider_pays_elec:
            st.caption("假设覆盖（未写回工作簿）：把对应方案的“节电后电费”从业主支出转移到服务商成本端。")

    # Apply what-if overlay to selected-tier dict (only affects derived views; does not touch cached workbook values)
    dsel_view = dict(dsel)
    dsel_view["product_key_active"] = pk_sel
    # When the selected product matches the workbook, keep LaaS drivers identical to the workbook cache
    # so “对齐Excel/Notebook” views (incl. what-if IRR) don’t accidentally double-apply product physics.
    if product_overrides_active:
        # Scale LaaS per-lamp CAPEX using product profile (relative to AI_lightning_grid ref).
        dsel_view["capex_laas_per_lamp"] = float(dsel.get("capex_laas_per_lamp") or 0.0) * float(capex_scale_vs_reference(pk_sel))
        # Update LaaS saving rate baseline from product physics.
        try:
            dsel_view["laas_saving_rate"] = max(
                float(dsel.get("laas_saving_rate") or 0.0), float(get_product_profile(pk_sel).implied_grid_saving_rate())
            )
        except Exception:
            pass
        # Scale non-electric OPEX components by product routine factor.
        om_scale = float(routine_om_scale(pk_sel))
        # IMPORTANT: only scale LaaS OPEX by product; EMC OPEX stays as workbook EMC baseline.
        for k in ("laas_opex_om_per_lamp", "laas_opex_platform", "laas_opex_spares"):
            if dsel.get(k) is not None:
                dsel_view[k] = float(dsel.get(k) or 0.0) * om_scale

    # Authoritative OPEX inputs (avoid any stale/missing cached extraction):
    # Use workbook values directly so EMC vs LaaS non-electric OPEX cannot accidentally collapse to 0.
    try:
        wb_inp = load_workbook(str(dsel["file_path"]), data_only=True, read_only=True)
        ws_inp = wb_inp["02_Inputs"]
        dsel_view["opex_om_per_lamp"] = float(ws_inp["D22"].value or dsel_view.get("opex_om_per_lamp") or 0.0)
        dsel_view["opex_platform"] = float(ws_inp["D23"].value or dsel_view.get("opex_platform") or 0.0)
        dsel_view["opex_spares"] = float(ws_inp["D24"].value or dsel_view.get("opex_spares") or 0.0)
        dsel_view["laas_opex_om_per_lamp"] = float(ws_inp["D32"].value or dsel_view.get("laas_opex_om_per_lamp") or 0.0)
        dsel_view["laas_opex_platform"] = float(ws_inp["D33"].value or dsel_view.get("laas_opex_platform") or 0.0)
        dsel_view["laas_opex_spares"] = float(ws_inp["D34"].value or dsel_view.get("laas_opex_spares") or 0.0)
        wb_inp.close()
    except Exception:
        pass

    def _emc_opex_parts() -> tuple[float, float, float]:
        return (
            float(dsel_view.get("opex_om_per_lamp") or 0.0),
            float(dsel_view.get("opex_platform") or 0.0),
            float(dsel_view.get("opex_spares") or 0.0),
        )

    def _laas_opex_parts() -> tuple[float, float, float]:
        # Fall back to EMC components if older workbooks didn't cache D32:D34.
        return (
            float(dsel_view.get("laas_opex_om_per_lamp") or dsel_view.get("opex_om_per_lamp") or 0.0),
            float(dsel_view.get("laas_opex_platform") or dsel_view.get("opex_platform") or 0.0),
            float(dsel_view.get("laas_opex_spares") or dsel_view.get("opex_spares") or 0.0),
        )

    # Electricity baseline and after-retrofit series (Y1..Y10) for both models.
    lamps = _safe_float(dsel_view.get("lamps"))
    # `baseline_electricity_y1` is extracted from `03_Baseline!D12`, i.e. **pre-retrofit** annual electricity cost.
    # Template wiring (`04_Mode_Params` row9) applies savings as: baseline_pre * (1 - saving_rate).
    # IMPORTANT: do not start from post-retrofit electricity here — it double-applies savings and breaks what-if IRR.
    elec_pre_y1 = dsel.get("baseline_electricity_y1")
    if elec_pre_y1 is None or elec_pre_y1 != elec_pre_y1:
        price = _safe_float(dsel.get("electricity_price_per_kwh"))
        watts = _safe_float(dsel.get("watts_per_lamp"))
        hpd = _safe_float(dsel.get("hours_per_day"))
        dpy = _safe_float(dsel.get("days_per_year"))
        elec_pre_y1 = lamps * watts / 1000.0 * hpd * dpy * price if (lamps and watts and hpd and dpy and price) else 0.0
    elec_pre_y1 = float(elec_pre_y1 or 0.0)
    emc_save = _safe_float(dsel_view.get("emc_saving_rate"))
    laas_save = _safe_float(dsel_view.get("laas_saving_rate"))
    elec_emc_y = [elec_pre_y1 * (1.0 - emc_save)] * 10
    elec_laas_y = [elec_pre_y1 * (1.0 - laas_save)] * 10

    # Owner headline savings (KPI cards): use the same electricity what-if as charts/tables (`_shift_electricity_payment`).
    _sp_emc_m = [float(x or 0.0) for x in df_t["owner_spend_emc"].tolist()]
    _sv_emc_m = [float(x or 0.0) for x in df_t["owner_save_emc"].tolist()]
    _sp_laas_m = [float(x or 0.0) for x in df_t["owner_spend_laas"].tolist()]
    _sv_laas_m = [float(x or 0.0) for x in df_t["owner_save_laas"].tolist()]
    _, owner_save_emc_card = _shift_electricity_payment(
        owner_spend_y=_sp_emc_m,
        owner_save_y=_sv_emc_m,
        elec_after_y=elec_emc_y,
        provider_pays_elec=emc_provider_pays_elec,
    )
    _, owner_save_laas_card = _shift_electricity_payment(
        owner_spend_y=_sp_laas_m,
        owner_save_y=_sv_laas_m,
        elec_after_y=elec_laas_y,
        provider_pays_elec=laas_provider_pays_elec,
    )
    owner_save_emc_y1_card = float(owner_save_emc_card[0]) if owner_save_emc_card else None
    owner_save_laas_y1_card = float(owner_save_laas_card[0]) if owner_save_laas_card else None

    # Provider IRR what-if (annual, simplified):
    # - Electricity uses **pre-retrofit** baseline (`03_Baseline!D12`) times (1 - saving), matching template row logic.
    # - LaaS subscription revenue uses the extracted `05_Annual_Model!D40:M40` series (net of prepay/tail when cached).
    # - Other/第三方收入 uses `02_Inputs!D20/D30` as a flat Y1..Y10 add-on (template structure).
    # - Upfront cash is modeled as Y0 inflow (+upfront) alongside CAPEX (-capex), matching the annual table convention.
    try:
        wb_chk_inputs = load_workbook(str(dsel["file_path"]), data_only=True, read_only=True)
        ws_chk = wb_chk_inputs["02_Inputs"]
        trust_other_y1 = _safe_float(ws_chk["D20"].value, default=0.0)
        laa_other_y1 = _safe_float(ws_chk["D30"].value, default=0.0)
        upfront_y0 = _safe_float(ws_chk["D45"].value, default=_safe_float(dsel.get("upfront")))
        wb_chk_inputs.close()
    except Exception:
        trust_other_y1 = 0.0
        laa_other_y1 = 0.0
        upfront_y0 = _safe_float(dsel.get("upfront"))

    lamps_i = float(dsel_view.get("lamps") or 0.0)
    trust_capex_y0 = float(dsel.get("capex_emc_per_lamp") or 0.0) * lamps_i
    laa_capex_y0 = float(dsel_view.get("capex_laas_per_lamp") or 0.0) * lamps_i

    trust_om_y = float(_emc_opex_parts()[0]) * lamps_i
    trust_plat_y = float(_emc_opex_parts()[1])
    trust_spare_y = float(_emc_opex_parts()[2])
    laa_om_y = float(_laas_opex_parts()[0]) * lamps_i
    laa_plat_y = float(_laas_opex_parts()[1])
    laa_spare_y = float(_laas_opex_parts()[2])

    laas_fee_y = [float(x or 0.0) for x in df_t["laas_fee"].tolist()]
    emc_fee_y = float(dsel.get("emc_fee_y1") or 0.0)

    def _compute_provider_block(*, kind: str) -> dict[str, list[float]]:
        """
        Build a provider-side cashflow block (Y0..Y10) using workbook inputs and
        the annual fee schedule already extracted from 05_Annual_Model.
        This avoids relying on cached values in formula-heavy rows that may be blank.
        """
        lamps = float(dsel_view.get("lamps") or 0.0)
        capex_per_lamp = float(dsel_view.get("capex_laas_per_lamp") or 0.0) if kind == "laas" else float(dsel_view.get("capex_emc_per_lamp") or 0.0)
        capex_y0 = capex_per_lamp * lamps

        # Electricity block: always compute the post-saving electricity amount,
        # then decide whether it sits on provider cost side based on what-if toggle.
        saving = float(dsel_view.get("laas_saving_rate") or 0.0) if kind == "laas" else float(dsel_view.get("emc_saving_rate") or 0.0)
        elec_after = [0.0] + [elec_pre_y1 * (1.0 - saving)] * 10
        if kind == "laas" and forces_grid_electricity_zero(pk_sel):
            elec_after = [0.0] * 11
        pays = emc_provider_pays_elec if kind == "emc" else laas_provider_pays_elec
        elec_cost = [0.0] + ([elec_after[i] for i in range(1, 11)] if pays else [0.0] * 10)

        # OPEX components (non-electric): EMC uses D22:D24, LaaS uses D32:D34
        if kind == "laas":
            om_per_lamp, platform, spares = _laas_opex_parts()
        else:
            om_per_lamp, platform, spares = _emc_opex_parts()
        om = float(om_per_lamp) * lamps
        om_y = [0.0] + [om] * 10
        platform_y = [0.0] + [platform] * 10
        spares_y = [0.0] + [spares] * 10

        # Revenue: service fee
        if kind == "laas":
            fee = [0.0] + [float(x or 0.0) for x in df_t["laas_fee"].tolist()]
        else:
            emc_fee_local = float(dsel.get("emc_fee_y1") or 0.0)
            fee = [0.0] + [emc_fee_local] * 10

        other_income = [0.0] * 11
        asset_income = [0.0] * 11

        total_rev = [fee[i] + other_income[i] + asset_income[i] for i in range(11)]
        total_cost = [elec_cost[i] + om_y[i] + platform_y[i] + spares_y[i] for i in range(11)]
        net_cf = [0.0] * 11
        net_cf[0] = -capex_y0
        for i in range(1, 11):
            net_cf[i] = total_rev[i] - total_cost[i]
        cum: list[float] = []
        s = 0.0
        for i in range(11):
            s += net_cf[i]
            cum.append(s)

        return {
            "初始CAPEX支出（正数，扣减）": [capex_y0] + [0.0] * 10,
            "固定服务费收入": fee,
            "其他/第三方收入": other_income,
            "资产转让收入": asset_income,
            "节电后电费总额": elec_after,
            "服务商承担电费成本（正数，扣减）": elec_cost,
            "运维成本（正数，扣减）": om_y,
            # Provide both EMC/LaaS label variants so downstream charts can always find the right series.
            "平台/管理成本（正数，扣减）": platform_y,
            "平台/AI升级成本（正数，扣减）": platform_y,
            "备件/小改造储备（正数，扣减）": spares_y,
            "升级/备件储备（正数，扣减）": spares_y,
            "年度总收入": total_rev,
            "年度总现金成本": total_cost,
            "年度净现金流": net_cf,
            "累计现金流": cum,
        }

    prov_emc = _compute_provider_block(kind="emc")
    prov_laas = _compute_provider_block(kind="laas")
    _cum_emc = prov_emc.get("累计现金流") or []
    _cum_laas = prov_laas.get("累计现金流") or []
    provider_cum_net_cf_emc_y10 = float(_cum_emc[-1]) if _cum_emc else None
    provider_cum_net_cf_laas_y10 = float(_cum_laas[-1]) if _cum_laas else None

    def _provider_net_cf_y0_to_y10(*, kind: str) -> list[float]:
        saving = float(dsel_view.get("laas_saving_rate") or 0.0) if kind == "laas" else float(dsel_view.get("emc_saving_rate") or 0.0)
        elec_after_y = [float(elec_pre_y1) * (1.0 - saving)] * 10
        if kind == "laas" and forces_grid_electricity_zero(pk_sel):
            elec_after_y = [0.0] * 10
        pays = emc_provider_pays_elec if kind == "emc" else laas_provider_pays_elec
        elec_cost_y = elec_after_y if pays else [0.0] * 10

        capex = float(laa_capex_y0 if kind == "laas" else trust_capex_y0)
        om_y = float(laa_om_y if kind == "laas" else trust_om_y)
        plat_y = float(laa_plat_y if kind == "laas" else trust_plat_y)
        spare_y = float(laa_spare_y if kind == "laas" else trust_spare_y)
        other_y = float(laa_other_y1 if kind == "laas" else trust_other_y1)

        flows: list[float] = [float(upfront_y0) - float(capex)]
        for i in range(10):
            rev = (float(laas_fee_y[i]) if kind == "laas" else float(emc_fee_y)) + float(other_y)
            flows.append(rev - (float(elec_cost_y[i]) + om_y + plat_y + spare_y))
        return flows

    irr_emc_whatif = _irr_yearly(_provider_net_cf_y0_to_y10(kind="emc"))
    irr_laas_whatif = _irr_yearly(_provider_net_cf_y0_to_y10(kind="laas"))

    # 1-minute: KPI cards
    st.subheader("A) 关键结论（1分钟版）")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("服务商 NPV（EMC → LaaS）", f"{_money(dsel.get('emc_npv'))} → {_money(dsel.get('laas_npv'))}")
    if emc_provider_pays_elec or laas_provider_pays_elec:
        k2.metric("服务商 IRR（EMC → LaaS）", f"{_pct(irr_emc_whatif)} → {_pct(irr_laas_whatif)}")
        k2.caption(
            "注：此处按侧边栏“电费承担(what-if)”重算IRR：电费按 `03_Baseline!D12×(1-节电率)`；"
            "LaaS订阅收入按 `05_Annual_Model` 提取的年费序列；Y0含 `02_Inputs!D45` 首期款现金流（与年度表展示一致）。"
            "这与 `01_Dashboard`/`00_LaaS收益来源` 缓存IRR仍可能不同（Excel端可能是更细粒度/不同IRR定义）。"
        )
    else:
        k2.metric("服务商 IRR（EMC → LaaS）", f"{_pct(dsel.get('emc_irr'))} → {_pct(dsel.get('laas_irr'))}")
    k3.metric("业主净节省 Y1（EMC → LaaS）", f"{_money(owner_save_emc_y1_card)} → {_money(owner_save_laas_y1_card)}")
    k4.metric(
        "10年累计净现金流（服务商，至Y10｜EMC → LaaS）",
        f"{_money(provider_cum_net_cf_emc_y10)} → {_money(provider_cum_net_cf_laas_y10)}",
    )
    k4.caption(
        "口径：**服务商**视角；为 Y0..Y10 年度净现金流累加（与下方对比表中的「累计现金流」列、以及累计现金流图一致）。"
        "Y0 含 CAPEX 流出；电费成本是否计入服务商侧随侧边栏电费承担 what-if。"
    )

    st.caption(f"工作簿路径：`{dsel['file_path']}`（数据来自该工作簿缓存值，便于审计与对齐WPS）")

    st.divider()

    st.subheader("对比总表（像Excel，便于汇报）")
    # A wide table similar to the screenshot: year 0..10, EMC vs LaaS core cashflow building blocks.
    years = [0] + list(df_t["year"].tolist())
    trust_capex = [-(dsel.get("capex_emc_per_lamp") or 0.0) * float(dsel.get("lamps") or 0.0)] + [0.0] * 10
    laas_capex = [-(dsel_view.get("capex_laas_per_lamp") or 0.0) * float(dsel_view.get("lamps") or 0.0)] + [0.0] * 10
    # Owner total spend rows already mirror Annual_Model; use 0 for year0.
    trust_spend_y = [float(x or 0.0) for x in df_t["owner_spend_emc"].tolist()]
    laas_spend_y = [float(x or 0.0) for x in df_t["owner_spend_laas"].tolist()]
    trust_save_y = [float(x or 0.0) for x in df_t["owner_save_emc"].tolist()]
    laas_save_y = [float(x or 0.0) for x in df_t["owner_save_laas"].tolist()]
    trust_spend_y, trust_save_y = _shift_electricity_payment(
        owner_spend_y=trust_spend_y, owner_save_y=trust_save_y, elec_after_y=elec_emc_y, provider_pays_elec=emc_provider_pays_elec
    )
    laas_spend_y, laas_save_y = _shift_electricity_payment(
        owner_spend_y=laas_spend_y, owner_save_y=laas_save_y, elec_after_y=elec_laas_y, provider_pays_elec=laas_provider_pays_elec
    )
    trust_spend = [0.0] + trust_spend_y
    laas_spend = [0.0] + laas_spend_y
    trust_save = [0.0] + trust_save_y
    laas_save = [0.0] + laas_save_y
    laas_fee_sched = [0.0] + [float(x or 0.0) for x in df_t["laas_fee"].tolist()]
    # Use fee inputs as a proxy for EMC annual fee for display (workbook-only; actual annual fee schedule is flat in template).
    emc_fee = float(dsel.get("emc_fee_y1") or 0.0)
    trust_fee = [0.0] + [emc_fee] * 10
    laas_fee = [0.0] + laas_fee_sched[1:]

    # `prov_emc` / `prov_laas` already computed above for the 1-minute KPI card.

    table_dict = {
        "year": years,
        # Owner view (top part)
        "EMC_capex(owner_view)": trust_capex,
        "EMC_fee(owner_view)": trust_fee,
        "EMC_total_spend(owner)": trust_spend,
        "EMC_net_savings(owner)": trust_save,
        "LaaS_capex(owner_view)": laas_capex,
        "LaaS_fee(owner_view)": laas_fee,
        "LaaS_total_spend(owner)": laas_spend,
        "LaaS_net_savings(owner)": laas_save,
    }
    for lab in prov_emc.keys():
        table_dict[f"EMC_服务商_{lab}"] = prov_emc[lab]
        table_dict[f"LaaS_服务商_{lab}"] = prov_laas[lab]

    table = pd.DataFrame(table_dict)

    st.caption("建议：默认仅展示关键列（适合汇报）；需要完整明细再展开。")
    show_full = st.toggle("展开：显示全部列（明细/审计）", value=False)

    key_cols = [
        "year",
        # Owner headline
        "EMC_total_spend(owner)",
        "LaaS_total_spend(owner)",
        "EMC_net_savings(owner)",
        "LaaS_net_savings(owner)",
        # Provider headline
        "EMC_服务商_年度净现金流",
        "LaaS_服务商_年度净现金流",
        "EMC_服务商_累计现金流",
        "LaaS_服务商_累计现金流",
    ]
    key_cols = [c for c in key_cols if c in table.columns]
    t_show = table if show_full else table[key_cols].copy()

    # Add deltas in key view
    if not show_full:
        if {"EMC_total_spend(owner)", "LaaS_total_spend(owner)"} <= set(t_show.columns):
            t_show.insert(
                3,
                "Δ业主总支出(LaaS-EMC)",
                t_show["LaaS_total_spend(owner)"] - t_show["EMC_total_spend(owner)"],
            )
        if {"EMC_net_savings(owner)", "LaaS_net_savings(owner)"} <= set(t_show.columns):
            t_show.insert(
                6,
                "Δ业主净节省(LaaS-EMC)",
                t_show["LaaS_net_savings(owner)"] - t_show["EMC_net_savings(owner)"],
            )
        if {"EMC_服务商_年度净现金流", "LaaS_服务商_年度净现金流"} <= set(t_show.columns):
            t_show["Δ服务商净现金流(LaaS-EMC)"] = t_show["LaaS_服务商_年度净现金流"] - t_show["EMC_服务商_年度净现金流"]

    def _fmt_money(v: float) -> str:
        try:
            return f"{float(v):,.0f}"
        except Exception:
            return "-"

    def _style(df: pd.DataFrame) -> "pd.io.formats.style.Styler":
        sty = df.style
        # Base number formatting
        num_cols = [c for c in df.columns if c != "year"]
        sty = sty.format({c: _fmt_money for c in num_cols})

        # Color-code columns so it's obvious which model is which
        emc_cols = [c for c in df.columns if c.startswith("EMC_")]
        laas_cols = [c for c in df.columns if c.startswith("LaaS_")]
        delta_cols = [c for c in df.columns if c.startswith("Δ")]

        if emc_cols:
            sty = sty.set_properties(
                subset=emc_cols,
                **{
                    "background-color": "#F8FAFC",  # very light slate
                    "color": "#334155",  # slate-700
                },
            )
        if laas_cols:
            sty = sty.set_properties(
                subset=laas_cols,
                **{
                    "background-color": "#ECFDF5",  # very light green
                    "color": "#065F46",  # green-800
                    "font-weight": "600",
                },
            )
        # Add a clear divider before the first LaaS column
        if laas_cols:
            first_laas = laas_cols[0]
            sty = sty.set_properties(subset=[first_laas], **{"border-left": "3px solid #10B981"})
        # Add a clear divider before delta block if present
        if delta_cols:
            first_delta = delta_cols[0]
            sty = sty.set_properties(subset=[first_delta], **{"border-left": "3px solid #F59E0B"})

        # Highlight most important columns
        highlight_cols = [c for c in df.columns if ("净节省" in c or "累计现金流" in c or "Δ" in c)]
        if highlight_cols:
            sty = sty.set_properties(subset=highlight_cols, **{"background-color": "#FFF7ED"})  # light orange

        # Conditional colors for deltas / savings
        def _color_pos_neg(val: float) -> str:
            try:
                x = float(val)
            except Exception:
                return ""
            if x > 0:
                return "color: #16A34A; font-weight: 600;"
            if x < 0:
                return "color: #DC2626; font-weight: 600;"
            return "color: #0F172A;"

        if delta_cols:
            sty = sty.map(_color_pos_neg, subset=delta_cols)

        save_cols = [c for c in df.columns if c.endswith("net_savings(owner)") or "净节省" in c]
        if save_cols:
            sty = sty.map(_color_pos_neg, subset=save_cols)

        return sty

    st.dataframe(_style(t_show), use_container_width=True, height=420)

    # Cumulative cashflow chart (provider) to show payback timing clearly
    st.subheader("累计现金流（服务商视角）：LaaS vs EMC（看回本更直观）")
    st.caption("口径：这是**服务商**的累计现金流（包含CAPEX投入与运维OPEX成本），用于展示回本速度。")
    cum_emc = prov_emc.get("累计现金流", [0.0] * 11)
    cum_laas = prov_laas.get("累计现金流", [0.0] * 11)
    df_cum = pd.DataFrame({"year": years, "EMC": cum_emc, "LaaS": cum_laas})
    fig_cum = go.Figure()
    fig_cum.add_trace(go.Scatter(x=df_cum["year"], y=df_cum["EMC"], name="EMC 累计现金流", mode="lines+markers", line=dict(color="#94A3B8", width=3)))
    fig_cum.add_trace(go.Scatter(x=df_cum["year"], y=df_cum["LaaS"], name="LaaS 累计现金流", mode="lines+markers", line=dict(color=GREEN, width=3)))
    fig_cum.add_hline(y=0, line_dash="dash", line_color="#CBD5E1")
    fig_cum.update_layout(template="plotly_white", height=360, font_color=NAVY, xaxis_title="年（Y0..Y10）", yaxis_title="累计现金流（元）")
    st.plotly_chart(fig_cum, use_container_width=True)

    # OPEX factor story chart (Y1) - why LaaS is smaller
    st.subheader("为什么LaaS的OPEX更小：拆解对比（Y1，电费 + 非电费运维）")
    st.caption(
        "口径：这里展示的是成本拆解。你可以在左侧分别切换 **EMC / LaaS 的电费承担** 来观察："
        "当“服务商承担电费”时，电费从业主支出口径转移到服务商成本口径（本页 what-if，不写回工作簿）。"
    )
    lamps = float(dsel_view.get("lamps") or 0.0)
    emc_om_per_lamp, emc_platform, emc_spares = _emc_opex_parts()
    laas_om_per_lamp, laas_platform, laas_spares = _laas_opex_parts()

    # Authoritative check: read the selected workbook’s cached values directly (bypasses Streamlit cache).
    # This avoids any mismatch if the cached extracted dataframe is stale.
    try:
        wb_chk = load_workbook(str(dsel["file_path"]), data_only=True, read_only=True)
        ws_chk = wb_chk["02_Inputs"]
        emc_om_per_lamp = float(ws_chk["D22"].value or emc_om_per_lamp)
        emc_platform = float(ws_chk["D23"].value or emc_platform)
        emc_spares = float(ws_chk["D24"].value or emc_spares)
        laas_om_per_lamp = float(ws_chk["D32"].value or laas_om_per_lamp)
        laas_platform = float(ws_chk["D33"].value or laas_platform)
        laas_spares = float(ws_chk["D34"].value or laas_spares)
        wb_chk.close()
    except Exception:
        pass
    st.caption(
        "当前使用的非电费运维输入（来自工作簿缓存值）：\n"
        f"- EMC：人工/维修 {emc_om_per_lamp:.2f} 元/盏/年，平台 {emc_platform:,.0f} 元/年，备件 {emc_spares:,.0f} 元/年\n"
        f"- LaaS：人工/维修 {laas_om_per_lamp:.2f} 元/盏/年，平台 {laas_platform:,.0f} 元/年，备件 {laas_spares:,.0f} 元/年"
    )
    emc_om_y1 = float(emc_om_per_lamp) * lamps
    emc_platform_y1 = float(emc_platform)
    emc_spares_y1 = float(emc_spares)
    laas_om_y1 = float(laas_om_per_lamp) * lamps
    laas_platform_y1 = float(laas_platform)
    laas_spares_y1 = float(laas_spares)

    elec_emc_y1 = float(elec_emc_y[0])
    elec_laas_y1 = float(elec_laas_y[0])

    # Baseline: no savings (electricity stays baseline), non-electric OPEX shown as the same template components for storytelling consistency.
    df_opex_story = pd.DataFrame(
        [
            # Baseline(现状)的非电费运维口径暂用EMC输入做近似（如需可从 03_Baseline 再提取一套）。
            {"方案": "Baseline(现状)", "电费": elec_pre_y1, "人工/维修": emc_om_y1, "平台": emc_platform_y1, "备件": emc_spares_y1},
            {"方案": "EMC(托管)", "电费": elec_emc_y1, "人工/维修": emc_om_y1, "平台": emc_platform_y1, "备件": emc_spares_y1},
            {"方案": "LaaS", "电费": elec_laas_y1, "人工/维修": laas_om_y1, "平台": laas_platform_y1, "备件": laas_spares_y1},
        ]
    )
    df_m = df_opex_story.melt(id_vars=["方案"], var_name="成本项", value_name="金额")
    fig_opex = px.bar(
        df_m,
        x="方案",
        y="金额",
        color="成本项",
        barmode="stack",
        template="plotly_white",
        height=420,
        title="OPEX拆解对比（Y1）",
        color_discrete_map={"电费": "#60A5FA", "人工/维修": "#94A3B8", "平台": "#FBBF24", "备件": "#A78BFA"},
    )
    fig_opex.update_layout(font_color=NAVY, yaxis_title="元/年")
    st.plotly_chart(fig_opex, use_container_width=True)
    st.caption("注：电费来自“基准电费 × (1-节电率)”。非电费运维拆解来自模板输入（按灯人工/维修、平台费、备件）。如需“现状非电费运维”单独口径，我们可以从 `03_Baseline` 再补充提取。")

    # Explain the “why LaaS wins for the provider” gap: 10y **provider** net CF decomposition (Y1–Y10).
    st.subheader("为什么 LaaS 对服务商更“跳”：10 年经营净现金流差异拆解（服务商视角）")
    st.caption(
        "口径：**服务商**（与上方对比表里 `年度净现金流` 一致）。"
        "把 **Y1–Y10 年度净现金流之和** 的差异 **(LaaS − EMC)** 拆成三块——它们代数和应等于右侧“核对合计”：\n"
        "- **订阅/服务收入差**：`年度总收入`（LaaS − EMC）\n"
        "- **电费成本差（服务商侧）**：`服务商承担电费成本`（EMC − LaaS），正值表示 EMC 在服务商账上电费更重\n"
        "- **非电费运维差**：人工+平台+备件（EMC − LaaS），正值表示 EMC 更重\n"
        "（电费是否进服务商成本，随侧边栏 what-if；与业主支出拆解不是同一件事。）"
    )

    def _sum_y1_y10(xs: list[float | None]) -> float:
        return float(sum(float(xs[i] or 0.0) for i in range(1, 11)))

    def _prov_series(p: dict[str, list[float] | None], key: str, *, n: int = 11) -> list[float]:
        """
        Copy provider block series as plain floats.
        IMPORTANT: avoid `dict.get(...) or default_list` — some objects (e.g. ndarray) can be falsy
        in surprising ways; we only want the explicit-missing-key fallback.
        """
        v = p.get(key)
        if not isinstance(v, (list, tuple)) or len(v) < n:
            return [0.0] * n
        return [float(v[i] or 0.0) for i in range(n)]

    rev_e = _prov_series(prov_emc, "年度总收入")
    rev_l = _prov_series(prov_laas, "年度总收入")
    el_e = _prov_series(prov_emc, "服务商承担电费成本（正数，扣减）")
    el_l = _prov_series(prov_laas, "服务商承担电费成本（正数，扣减）")
    om_e = _prov_series(prov_emc, "运维成本（正数，扣减）")
    om_l = _prov_series(prov_laas, "运维成本（正数，扣减）")
    pl_e = _prov_series(prov_emc, "平台/管理成本（正数，扣减）")
    pl_l = _prov_series(prov_laas, "平台/AI升级成本（正数，扣减）") if "平台/AI升级成本（正数，扣减）" in prov_laas else _prov_series(prov_laas, "平台/管理成本（正数，扣减）")
    sp_e = _prov_series(prov_emc, "备件/小改造储备（正数，扣减）")
    sp_l = _prov_series(prov_laas, "升级/备件储备（正数，扣减）") if "升级/备件储备（正数，扣减）" in prov_laas else _prov_series(prov_laas, "备件/小改造储备（正数，扣减）")
    tc_e = _prov_series(prov_emc, "年度总现金成本")
    tc_l = _prov_series(prov_laas, "年度总现金成本")
    net_e = _prov_series(prov_emc, "年度净现金流")
    net_l = _prov_series(prov_laas, "年度净现金流")

    d_rev_y110 = _sum_y1_y10(rev_l) - _sum_y1_y10(rev_e)
    d_elec_y110 = _sum_y1_y10(el_e) - _sum_y1_y10(el_l)
    d_non_elec_y110 = 0.0
    for i in range(1, 11):
        ne = float(om_e[i] or 0.0) + float(pl_e[i] or 0.0) + float(sp_e[i] or 0.0)
        nl = float(om_l[i] or 0.0) + float(pl_l[i] or 0.0) + float(sp_l[i] or 0.0)
        d_non_elec_y110 += ne - nl
    # Cross-check: non-electric cash cost = total cash cost - electricity (same rows as the wide table).
    d_non_elec_from_tc = 0.0
    for i in range(1, 11):
        d_non_elec_from_tc += float(tc_e[i] - el_e[i]) - float(tc_l[i] - el_l[i])
    if abs(d_non_elec_from_tc - d_non_elec_y110) > 1.0:
        d_non_elec_y110 = float(d_non_elec_from_tc)

    d_net_y110 = _sum_y1_y10(net_l) - _sum_y1_y10(net_e)
    bridge_check = float(d_rev_y110 + d_elec_y110 + d_non_elec_y110)

    y0_delta = float((net_l[0] or 0.0) - (net_e[0] or 0.0)) if len(net_l) > 0 and len(net_e) > 0 else 0.0

    drv_rows: list[dict[str, object]] = [
        {"驱动": "订阅/服务收入差（LaaS−EMC）", "金额": d_rev_y110},
        {"驱动": "电费成本差（EMC−LaaS，服务商侧）", "金额": d_elec_y110},
        {"驱动": "非电费运维差（EMC−LaaS）", "金额": d_non_elec_y110},
        {"驱动": "核对：Y1–Y10 净现金流差（LaaS−EMC）", "金额": d_net_y110},
    ]
    if abs(y0_delta) > 1e-3:
        drv_rows.insert(
            0,
            {"驱动": "Y0：CAPEX现金流差（LaaS−EMC）", "金额": y0_delta},
        )

    df_drv = pd.DataFrame(drv_rows)
    df_drv["金额"] = pd.to_numeric(df_drv["金额"], errors="coerce").fillna(0.0)

    # Use explicit `go.Bar` traces (one category per trace) + value labels.
    # Some Plotly/Streamlit layouts made mid-axis bars hard to see on dense Chinese x labels.
    fig_drv = go.Figure()
    bar_colors = ["#64748B", "#2563EB", "#F59E0B", "#94A3B8", "#10B981"]
    for i, row in df_drv.reset_index(drop=True).iterrows():
        val = float(row["金额"])
        fig_drv.add_trace(
            go.Bar(
                x=[str(row["驱动"])],
                y=[val],
                name=str(row["驱动"]),
                showlegend=False,
                marker_color=bar_colors[i % len(bar_colors)],
                text=[f"{val:,.0f}"],
                textposition="outside",
                cliponaxis=False,
                width=0.55,
            )
        )
    fig_drv.update_layout(
        template="plotly_white",
        height=420,
        title="10年经营侧拆解（Y1–Y10 合计；服务商视角）",
        font_color=NAVY,
        yaxis_title="元（Y1–Y10 合计）",
        xaxis=dict(tickangle=-25),
        bargap=0.25,
        margin=dict(t=60, b=120),
    )
    st.plotly_chart(fig_drv, use_container_width=True)
    st.caption(
        "拆解数值（便于核对 `05_Annual_Model` 行17–20 / 行45–47）：\n"
        f"- 订阅/服务收入差（LaaS−EMC，Y1–Y10合计）：{d_rev_y110:,.0f}\n"
        f"- 电费成本差（EMC−LaaS，服务商侧，Y1–Y10合计）：{d_elec_y110:,.0f}\n"
        f"- 非电费运维差（EMC−LaaS，Y1–Y10合计）：{d_non_elec_y110:,.0f}\n"
        f"- 核对：净现金流差（LaaS−EMC，Y1–Y10合计）：{d_net_y110:,.0f}"
    )
    _non_rows = df_drv.loc[df_drv["驱动"].astype(str).str.contains("非电费", na=False), "金额"]
    if len(_non_rows) and abs(float(_non_rows.iloc[0])) < 1.0:
        st.caption(
            "「非电费运维差」为 **0 或≈0**：表示在当前侧边栏假设下，两模式 **人工+平台+备件** 的 **10 年合计** 对服务商几乎相同 "
            "（或已被产品 what-if 对齐）；此时该柱会贴近 0 轴线。"
        )
    if abs(bridge_check - d_net_y110) > 1.0:
        st.caption(f"内部核对：分项之和与净现金流差相差 {bridge_check - d_net_y110:,.0f} 元（若出现请反馈）。")

    tab_owner, tab_provider, tab_cost, tab_story, tab_trace = st.tabs(
        ["业主视角（年度表）", "服务商视角（回报）", "成本拆解（OPEX/CAPEX）", "为什么更好（证据）", "可追溯性"],
    )

    with tab_owner:
        st.subheader("B) 业主视角：年度总支出与年度净节省（对应 05_Annual_Model）")
        fig1 = go.Figure()
        fig1.add_trace(
            go.Scatter(x=df_t["year"], y=trust_spend_y, name="EMC 业主年度总支出", mode="lines+markers", line=dict(color="#94A3B8", width=3))
        )
        fig1.add_trace(
            go.Scatter(x=df_t["year"], y=laas_spend_y, name="LaaS 业主年度总支出", mode="lines+markers", line=dict(color=ACCENT, width=3))
        )
        fig1.update_layout(template="plotly_white", height=360, font_color=NAVY, xaxis_title="年", yaxis_title="元/年")
        st.plotly_chart(fig1, use_container_width=True)

        fig2 = go.Figure()
        fig2.add_trace(
            go.Scatter(x=df_t["year"], y=trust_save_y, name="EMC 业主年度净节省", mode="lines+markers", line=dict(color=RED, width=3))
        )
        fig2.add_trace(
            go.Scatter(x=df_t["year"], y=laas_save_y, name="LaaS 业主年度净节省", mode="lines+markers", line=dict(color=GREEN, width=3))
        )
        fig2.update_layout(template="plotly_white", height=360, font_color=NAVY, xaxis_title="年", yaxis_title="元/年")
        st.plotly_chart(fig2, use_container_width=True)

        if show_all_years_table:
            st.subheader("年度表格（Y1..Y10，像Excel）")
            st.dataframe(
                df_t[
                    [
                        "year",
                        "owner_spend_emc",
                        "owner_spend_laas",
                        "owner_save_emc",
                        "owner_save_laas",
                        "laas_fee",
                    ]
                ],
                use_container_width=True,
            )

    with tab_provider:
        st.subheader("C) 服务商视角：回报指标（对应 01_Dashboard）")
        c1, c2 = st.columns(2)
        with c1:
            st.metric("NPV（EMC）", _money(dsel.get("emc_npv")))
            st.metric("IRR（EMC）", _pct(dsel.get("emc_irr")))
        with c2:
            st.metric("NPV（LaaS）", _money(dsel.get("laas_npv")))
            st.metric("IRR（LaaS）", _pct(dsel.get("laas_irr")))

        fig = go.Figure()
        fig.add_trace(go.Bar(name="EMC", x=["NPV", "IRR"], y=[dsel.get("emc_npv") or 0.0, (dsel.get("emc_irr") or 0.0) * 1e7], marker_color="#94A3B8"))
        fig.add_trace(go.Bar(name="LaaS", x=["NPV", "IRR"], y=[dsel.get("laas_npv") or 0.0, (dsel.get("laas_irr") or 0.0) * 1e7], marker_color=ACCENT))
        fig.update_layout(
            barmode="group",
            template="plotly_white",
            height=360,
            font_color=NAVY,
            title="服务商KPI对比（IRR为展示缩放）",
        )
        st.plotly_chart(fig, use_container_width=True)

    with tab_cost:
        st.subheader("D) 成本拆解：OPEX包含什么 + CAPEX规模（来自 02_Inputs / 03_Baseline）")
        st.caption(
            "说明：本页的OPEX饼图（人工/维修/平台/备件）是**服务商支付的运维成本**。"
            "电费是否由服务商承担：工作簿里 EMC 对应 `02_Inputs!D25`，LaaS 对应 `02_Inputs!D35`；"
            "本页侧边栏可分别做 what-if（不写回工作簿）。"
        )
        lamps = dsel_view.get("lamps") or 0.0
        capex_emc = (dsel_view.get("capex_emc_per_lamp") or 0.0) * float(lamps)
        capex_laas = (dsel_view.get("capex_laas_per_lamp") or 0.0) * float(lamps)
        baseline_elec = dsel_view.get("baseline_electricity_y1")
        if baseline_elec is None or baseline_elec != baseline_elec:
            price = float(dsel_view.get("electricity_price_per_kwh") or 0.0)
            watts = float(dsel_view.get("watts_per_lamp") or 0.0)
            hpd = float(dsel_view.get("hours_per_day") or 0.0)
            dpy = float(dsel_view.get("days_per_year") or 0.0)
            baseline_elec = float(lamps) * watts / 1000.0 * hpd * dpy * price if (lamps and watts and hpd and dpy and price) else None
        st.markdown(
            f"- **灯具数量**：{_money(lamps)}\n"
            f"- **CAPEX（EMC vs LaaS）**：{_money(capex_emc)} → {_money(capex_laas)}\n"
            f"- **基准电费(Y1)**：{_money(baseline_elec)}\n"
            f"- **节电率（EMC vs LaaS）**：{_pct(dsel_view.get('emc_saving_rate'))} → {_pct(dsel_view.get('laas_saving_rate'))}\n"
        )
        opex_breakdown = pd.DataFrame(
            {
                "项目": ["人工/维修（按灯）", "平台费", "备件"],
                "金额": [
                    float(laas_om_per_lamp) * float(lamps),
                    float(laas_platform),
                    float(laas_spares),
                ],
            }
        )
        fig_o = px.pie(opex_breakdown, names="项目", values="金额", title="非电费运维结构（模板输入拆解）", template="plotly_white")
        fig_o.update_layout(height=360, font_color=NAVY)
        st.plotly_chart(fig_o, use_container_width=True)
        st.caption("电费在模板中通过“基准电费 × (1-节电率)”单独计算；此处饼图仅展示非电费运维（默认由服务商承担）。")

    with tab_story:
        st.subheader("E) 为什么更好（证据 + 映射到单元格）")
        cards = cards_for_selected_tier(
            has_upfront=float(dsel.get("upfront") or 0.0) > 0,
            has_tail_discount=float(dsel.get("tail_discount") or 0.0) > 0,
            laas_saving_rate=dsel.get("laas_saving_rate"),
            product_key=str(dsel_view.get("product_key_active") or dsel.get("product_key") or ""),
        )
        for c in cards:
            with st.expander(c.title, expanded=(view_mode != "只看关键结论（1分钟版）")):
                st.markdown(f"**为什么**：{c.why}")
                st.markdown(f"**证据/参考**：`{c.evidence_url}`")
                st.markdown("**对应模型单元格**：\n" + "\n".join([f"- `{x}`" for x in c.maps_to_cells]))
                if c.notes:
                    st.markdown(f"**备注**：{c.notes}")

    with tab_trace:
        st.subheader("可追溯性（该方案工作簿的读取范围）")
        td = next((x for x in tiers_dict if x["tier_name"] == sel), None)
        if td is None:
            st.warning("Tier not found in extracted bundle.")
        else:
            st.code(json.dumps(td, ensure_ascii=False, indent=2), language="json")

if __name__ == "__main__":
    main()

