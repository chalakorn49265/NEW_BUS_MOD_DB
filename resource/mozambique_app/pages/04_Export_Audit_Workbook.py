from __future__ import annotations

import io
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook

_ROOT = Path(__file__).resolve().parents[2]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from mozambique.baseline import DEFAULT_BASELINES  # noqa: E402
from mozambique.deal import DealCharge, DealSplits, DealSubscription  # noqa: E402
from mozambique.distribution import annualize_monthly  # noqa: E402
from mozambique.scenario import MozambiqueInputs, run_mozambique_scenario  # noqa: E402


NAVY = "#1F3864"
MUTED = "#6B7280"


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    # headers
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=str(col))
    # rows
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, v in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=float(v) if isinstance(v, (int, float)) else (None if v is None else str(v)))


def _build_workbook_bytes(res) -> bytes:
    wb = Workbook()

    ws_in = wb.active
    ws_in.title = "Inputs"
    inputs_rows = [
        ("baseline_name", res.baseline.name),
        ("baseline_watt_per_light", res.baseline.watt_per_light),
        ("baseline_maint_usd_per_light_year", res.baseline.maintenance_usd_per_light_year),
        ("number_of_lights", res.inputs.number_of_lights),
        ("operating_hours_per_night", res.inputs.operating_hours_per_night),
        ("days_per_year", res.inputs.days_per_year),
        ("electricity_price_usd_per_kwh", res.inputs.electricity_price_usd_per_kwh),
        ("term_years", res.subscription.term_years),
        ("annual_fee_usd", res.subscription.annual_fee_usd),
        ("upfront_usd", res.subscription.upfront_usd),
        ("escalation_pct_annual", res.subscription.escalation_pct_annual),
        ("capex_usd_total", res.inputs.capex_usd_total),
        ("provider_opex_annual_usd", res.inputs.provider_opex_annual_usd),
    ]
    for i, (k, v) in enumerate(inputs_rows, start=1):
        ws_in.cell(row=i, column=1, value=str(k))
        ws_in.cell(row=i, column=2, value=float(v) if isinstance(v, (int, float)) else str(v))

    ws_base = wb.create_sheet("Baseline_vs_After")
    base_total = float(res.baseline_energy_annual_usd + res.baseline_maintenance_annual_usd)
    after_total = float(res.subscription.annual_fee_usd)
    df_cmp = pd.DataFrame(
        [
            {"item": "Baseline electricity (USD/yr)", "value": res.baseline_energy_annual_usd},
            {"item": "Baseline maintenance (USD/yr)", "value": res.baseline_maintenance_annual_usd},
            {"item": "Baseline total (USD/yr)", "value": base_total},
            {"item": "After: subscription (USD/yr)", "value": after_total},
        ]
    )
    _write_df(ws_base, df_cmp)

    ws_cust = wb.create_sheet("Customer_incremental")
    cust_ann = annualize_monthly(res.customer_incremental_monthly)
    df_c = pd.DataFrame({"year": list(range(0, len(cust_ann))), "incremental_benefit_usd": cust_ann})
    _write_df(ws_cust, df_c)

    ws_prov = wb.create_sheet("Provider_net")
    df_p = pd.DataFrame(
        {
            "year": list(range(0, len(res.provider_net_annual_y0_to_yN))),
            "net_cashflow_usd": res.provider_net_annual_y0_to_yN,
            "cumulative_usd": res.provider_net_cumulative_annual,
        }
    )
    _write_df(ws_prov, df_p)

    ws_split = wb.create_sheet("Stakeholder_revenue")
    df_s = pd.DataFrame(res.stakeholder_revenue_annual)
    if df_s.empty:
        df_s = pd.DataFrame([{"year": 0, "stakeholder": "None", "kind": "none", "cash_in_usd": 0.0}])
    _write_df(ws_split, df_s.sort_values(["year", "kind", "stakeholder"]))

    ws_kpi = wb.create_sheet("KPIs")
    kpi_rows = [
        ("customer_irr_annual", res.customer_irr_annual),
        ("customer_payback_month", res.customer_payback_month),
        ("provider_net_irr_annual", res.provider_net_irr_annual),
        ("provider_net_payback_month", res.provider_net_payback_month),
    ]
    for i, (k, v) in enumerate(kpi_rows, start=1):
        ws_kpi.cell(row=i, column=1, value=str(k))
        if isinstance(v, (int, float)):
            ws_kpi.cell(row=i, column=2, value=float(v))
        else:
            ws_kpi.cell(row=i, column=2, value=str(v))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def main() -> None:
    st.set_page_config(page_title="Mozambique — Export", layout="wide", initial_sidebar_state="expanded")
    st.markdown(
        f"<h2 style='color:{NAVY};margin-bottom:0.2rem;'>Export audit workbook (.xlsx)</h2>"
        f"<p style='color:{MUTED};margin-top:0;'>Generates an Excel file with inputs, baseline vs after, customer/provider cashflows, and stakeholder revenues.</p>",
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.subheader("Project")
        baseline_type = st.selectbox("Existing baseline type", options=["LED", "HPS"], index=1)
        n = st.number_input("Number of lights", min_value=1, value=1000, step=50)
        hours = st.slider("Operating hours per night", 4.0, 14.0, 11.0, 0.25)
        days = st.number_input("Days per year", min_value=300.0, max_value=366.0, value=365.0, step=1.0)
        price = st.number_input("Electricity price (USD/kWh)", min_value=0.0, value=0.10, step=0.01, format="%.2f")

        st.divider()
        st.subheader("Subscription")
        term = st.number_input("Term (years)", min_value=1, max_value=20, value=10, step=1)
        annual_fee = st.number_input("Annual subscription fee (USD/year)", min_value=0.0, value=600_000.0, step=25_000.0)
        upfront = st.number_input("Upfront payment (USD, month 0)", min_value=0.0, value=0.0, step=25_000.0)
        esc = st.slider("Annual escalation (%)", -5.0, 20.0, 3.0, 0.5) / 100.0

        st.divider()
        st.subheader("Provider economics")
        capex_total = st.number_input("Provider CAPEX total (USD)", min_value=0.0, value=3_000_000.0, step=100_000.0)
        opex_annual = st.number_input("Provider annual OPEX (USD/year)", min_value=0.0, value=120_000.0, step=10_000.0)

        st.divider()
        st.subheader("Splits (simple)")
        gov_pct = st.slider("Government (% of subscription)", 0.0, 40.0, 5.0, 0.5) / 100.0
        interm_pct = st.slider("Intermediaries (% of subscription)", 0.0, 40.0, 8.0, 0.5) / 100.0
        undertable = st.number_input("Undertable fixed (USD/year)", min_value=0.0, value=0.0, step=10_000.0)

    baseline = DEFAULT_BASELINES[baseline_type]  # type: ignore[index]
    subscription = DealSubscription(term_years=int(term), annual_fee_usd=float(annual_fee), upfront_usd=float(upfront), escalation_pct_annual=float(esc))
    charges = [
        DealCharge(recipient="Government", kind="government", timing="annual", pct_of_subscription=float(gov_pct)),
        DealCharge(recipient="Intermediaries", kind="intermediary", timing="annual", pct_of_subscription=float(interm_pct)),
    ]
    if undertable > 0:
        charges.append(DealCharge(recipient="Undertable", kind="undertable", timing="annual", fixed_usd=float(undertable)))
    splits = DealSplits(charges=charges)

    res = run_mozambique_scenario(
        baseline=baseline,
        subscription=subscription,
        splits=splits,
        inputs=MozambiqueInputs(
            number_of_lights=int(n),
            operating_hours_per_night=float(hours),
            days_per_year=float(days),
            electricity_price_usd_per_kwh=float(price),
            capex_usd_total=float(capex_total),
            provider_opex_annual_usd=float(opex_annual),
        ),
    )

    st.subheader("Preview")
    st.markdown(
        f"- **Baseline annual cost (USD/yr)**: {res.baseline_energy_annual_usd + res.baseline_maintenance_annual_usd:,.0f}\n"
        f"- **Subscription (USD/yr)**: {res.subscription.annual_fee_usd:,.0f}\n"
        f"- **Provider IRR (net)**: {res.provider_net_irr_annual if isinstance(res.provider_net_irr_annual, str) else f'{res.provider_net_irr_annual:.1%}'}\n"
        f"- **Provider payback (months)**: {res.provider_net_payback_month}"
    )

    b = _build_workbook_bytes(res)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    fn = f"mozambique_ai_solar_laas_audit_{ts}.xlsx"
    st.download_button("Download audit workbook", data=b, file_name=fn, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")


if __name__ == "__main__":
    main()

