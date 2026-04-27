"""
Fill the reference CN workbook template with Roadlight baseline + 10 LaaS tiers.

- Copies `通用版_能源托管_vs_LaaS_财务模型_v3_逻辑复核版.xlsx` and writes calibrated `02_Inputs`
  (plus optional `05_Annual_Model!C41` upfront) so template formulas stay intact.
- KPI anchors + key Dashboard cells get **cached numeric <v>** for WPS (no recalc on open).

Run from repo root:
  python -m Dashboard_LaaS_vs_NYTG.excel_template_fill_wps
"""

from __future__ import annotations

import io
import shutil
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import numpy_financial as npf
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from business_model_comparison.models import build_baseline_energy_trust
from business_model_comparison.roadlight_data import load_roadlight_all
from Dashboard_LaaS_vs_NYTG.laas_feasible import LaaSScenario, evaluate_laas_scenario
from Dashboard_LaaS_vs_NYTG.product_profiles import capex_scale_vs_reference, get_product_profile, routine_om_scale

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "Dashboard_LaaS_vs_NYTG" / "通用版_能源托管_vs_LaaS_财务模型_v3_逻辑复核版.xlsx"
OUT_DIR = ROOT / "Dashboard_LaaS_vs_NYTG" / "new_models"
DEBUG_LOG_PATH = ROOT / ".cursor" / "debug-9617d5.log"

# User-provided: EMC provider service revenue (Y1) fixed across tiers.
EMC_FEE_Y1_RMB_FIXED = 2_342_300.0  # 234.23 万元
# Owner budget cap (基础参数：业主年度服务费预算/付费上限) should be above EMC fee.
OWNER_SERVICE_BUDGET_Y1_RMB = 2_600_000.0

# User-provided EMC baseline vs after-retrofit electricity (Y1), used as the template-grounded reference.
EMC_BASELINE_KWH_Y1 = 3_073_266.0
EMC_AFTER_KWH_Y1 = 1_489_860.0
EMC_BASELINE_ELEC_FEE_Y1_RMB = 2_120_554.0
EMC_AFTER_ELEC_FEE_Y1_RMB = 1_028_003.0
# Implied saving fraction (≈51.5%, commonly stated as ~52%).
EMC_SAVING_FRAC = 1.0 - (EMC_AFTER_ELEC_FEE_Y1_RMB / max(1e-9, EMC_BASELINE_ELEC_FEE_Y1_RMB))

# User-provided baseline table (lights/poles counts) used for `03_Baseline` display and for
# deriving the per-lamp wattage consistent with `EMC_BASELINE_KWH_Y1`.
EMC_LIGHTS_COUNT_Y1 = 4568.0
EMC_POLES_COUNT_Y1 = 2345.0
# Use stable, human-readable operating assumptions for baseline physics.
EMC_HOURS_PER_DAY_Y1 = 11.0
EMC_DAYS_PER_YEAR_Y1 = 365.0

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _q(tag: str) -> str:
    return f"{{{NS_MAIN}}}{tag}"

# #region agent log
def _dlog(location: str, message: str, data: dict, *, run_id: str = "pre-fix", hypothesis_id: str = "H_cashflow") -> None:
    import json
    import time

    payload = {
        "sessionId": "9617d5",
        "runId": run_id,
        "hypothesisId": hypothesis_id,
        "location": location,
        "message": message,
        "data": data,
        "timestamp": int(time.time() * 1000),
    }
    try:
        DEBUG_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with DEBUG_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        return

# #endregion agent log

@dataclass(frozen=True)
class TierSpec:
    name: str
    term_years: int
    opex_mode: Literal["uniform_pct", "electricity_only_pct", "ai_plus_solar"]
    product_key: Literal["AI_lightning_grid", "AI_battery_integrated_grid", "AI_plus_solar_offgrid"]
    annual_fee_rmb: float
    upfront_rmb: float
    tail_reduction_rmb: float
    ai_reduction_pct: float


def _owner_net_savings_min_rmb(
    *,
    baseline_pre_elec_y1: float,
    baseline_other_y1: float,
    term_years: int,
    saving_rate: float,
    annual_payments_rmb: list[float],
) -> float:
    """Owner metric consistent with template: baseline D8 (elec+om) minus (elec*(1-saving)+fee)."""
    base = float(baseline_pre_elec_y1 + baseline_other_y1)
    out = []
    for y in range(1, int(term_years) + 1):
        fee = float(annual_payments_rmb[y - 1]) if y - 1 < len(annual_payments_rmb) else 0.0
        elec = float(baseline_pre_elec_y1) * (1.0 - float(saving_rate))
        out.append(base - (elec + fee))
    return float(min(out)) if out else 0.0


def select_10_tiers(*, baseline) -> list[TierSpec]:
    """Pick 10 scenarios that satisfy BOTH provider constraints and owner net savings > 0 under template metric."""
    # Inputs for owner metric (see fill_one_tier comments).
    emc_saving = float(EMC_SAVING_FRAC)
    elec_post_y1 = float(EMC_AFTER_ELEC_FEE_Y1_RMB)
    cash_y1 = float(baseline.cash_opex_rmb_y.get(1))
    other_y1 = max(0.0, cash_y1 - elec_post_y1)
    elec_pre_y1 = float(elec_post_y1 / max(1e-9, (1.0 - emc_saving)))

    def saving_rate_for(mode: str, red: float) -> float:
        if mode == "ai_plus_solar":
            return 0.999
        return max(emc_saving + 0.03, max(0.0, min(0.95, float(red))))

    # Mechanism templates for 10 diversified tiers (we solve annual_fee_rmb via search).
    mech: list[tuple[str, int, str, float, float, float]] = [
        ("Tier01_UniformAI_80", 10, "uniform_pct", 0.80, 0.0, 0.0),
        ("Tier02_UniformAI_85", 10, "uniform_pct", 0.85, 0.0, 0.0),
        ("Tier03_ElecOnly_85", 10, "electricity_only_pct", 0.85, 0.0, 0.0),
        # Keep an “upfront” style tier, but cap it lower so provider cumulative still beats EMC under fee_cap_vs_emc.
        ("Tier04_ElecOnly_85_Upfront100k", 10, "electricity_only_pct", 0.85, 100_000.0, 0.0),
        # Tier05 no longer uses a dedicated “AI+solar mode”; solar becomes a product type (see product_key).
        ("Tier05_UniformAI_75", 10, "uniform_pct", 0.75, 0.0, 0.0),
        ("Tier06_Upfront_Prepay", 10, "uniform_pct", 0.85, 1_000_000.0, 0.0),
        ("Tier07_TailDiscount", 10, "uniform_pct", 0.85, 0.0, 250_000.0),
        ("Tier08_HigherFee_StrongerSLA", 10, "uniform_pct", 0.85, 0.0, 350_000.0),
        ("Tier09_MidTerm_8y", 8, "uniform_pct", 0.85, 500_000.0, 0.0),
        # Conservative but must still beat EMC on provider cumulative under fee cap.
        ("Tier10_Conservative", 10, "uniform_pct", 0.75, 0.0, 0.0),
    ]

    picked: list[TierSpec] = []
    product_cycle: list[Literal["AI_lightning_grid", "AI_battery_integrated_grid", "AI_plus_solar_offgrid"]] = [
        "AI_lightning_grid",
        "AI_battery_integrated_grid",
        "AI_plus_solar_offgrid",
    ]
    for i, (name, term, mode, red, upfront, tail) in enumerate(mech):
        product_key = product_cycle[i % len(product_cycle)]
        # Ensure the “upfront-style” tier still beats EMC even under fee_cap_vs_emc by giving it
        # the strongest OPEX advantage (solar/off-grid eliminates electricity OPEX in our product overlay).
        if name.startswith("Tier04_"):
            product_key = "AI_plus_solar_offgrid"
        s_rate = saving_rate_for(mode, red)
        # Owner-positivity implies an upper bound on annual fee (ignoring retained mgmt which we set to 0):
        fee_max_owner = float(other_y1 + elec_pre_y1 * s_rate) - 1.0
        fee_min = 200_000.0
        # IMPORTANT: if user toggles `02_Inputs!D35=1` (服务商承担电费), then for BOTH EMC and LaaS
        # the owner electricity outflow becomes 0 in the template. In that case, owner's annual spend
        # is dominated by the annual service fee. To preserve the narrative goal “LaaS reduces owner spend”
        # even under D35=1, enforce LaaS annual fee < EMC annual fee.
        fee_cap_vs_emc = float(EMC_FEE_Y1_RMB_FIXED) - 1.0
        fee_max = max(fee_min, min(3_000_000.0, fee_max_owner, fee_cap_vs_emc))

        best = None
        # Scan feasible fee range with 10k step; keep best provider NPV.
        step = 10_000.0
        f = fee_min
        while f <= fee_max + 1e-6:
            scen = LaaSScenario(
                term_years=int(term),
                annual_service_fee_rmb=float(f),
                upfront_rmb=float(upfront),
                ai_opex_reduction_pct=float(red),
                last_four_year_fee_reduction_rmb=float(tail),
                opex_mode=mode,  # type: ignore[arg-type]
                product_key=product_key,
            )
            r = evaluate_laas_scenario(baseline, scen, discount_rate_annual=0.12)
            if not r.provider_feasible:
                f += step
                continue
            if not (isinstance(r.irr_project_annual, float) and r.irr_project_annual > 0):
                f += step
                continue
            if float(r.npv_project_rmb) <= 0:
                f += step
                continue
            annual_payments = [float(r.client_payment_rmb_y.get(y, 0.0)) for y in range(1, int(term) + 1)]
            owner_min = _owner_net_savings_min_rmb(
                baseline_pre_elec_y1=elec_pre_y1,
                baseline_other_y1=other_y1,
                term_years=int(term),
                saving_rate=s_rate,
                annual_payments_rmb=annual_payments,
            )
            if owner_min <= 0:
                f += step
                continue
            score = float(r.npv_project_rmb)
            if best is None or score > best[0]:
                best = (score, float(f))
            f += step

        if best is None:
            raise RuntimeError(f"No feasible fee found for {name}. Try adjusting mechanism or search bounds.")
        picked.append(
            TierSpec(
                name=name,
                term_years=int(term),
                opex_mode=mode,  # type: ignore[arg-type]
                product_key=product_key,
                annual_fee_rmb=float(best[1]),
                upfront_rmb=float(upfront),
                tail_reduction_rmb=float(tail),
                ai_reduction_pct=float(red),
            )
        )

    return picked


def _baseline_elec_om_y1(
    *,
    lamps: float,
    price_per_kwh: float,
    watts_per_lamp: float,
    hours_per_day: float,
    days_per_year: float,
    owner_om_per_lamp: float,
) -> tuple[float, float, float]:
    hours_y = float(hours_per_day) * float(days_per_year)
    kwh = float(lamps) * float(watts_per_lamp) / 1000.0 * hours_y
    elec = kwh * float(price_per_kwh)
    om = float(lamps) * float(owner_om_per_lamp)
    return elec, om, elec + om


def _calibrate_physics_to_baseline(
    *,
    fee_y1: float,
    elec_target: float,
    om_target: float,
    lamps: float = 5000.0,
    price_per_kwh: float = 0.5,
    hours_per_day: float = 11.0,
    days_per_year: float = 365.0,
) -> dict[str, float]:
    denom = float(lamps) / 1000.0 * hours_per_day * days_per_year * price_per_kwh
    watts = float(elec_target) / denom if denom > 0 else 150.0
    watts = max(30.0, min(400.0, watts))
    owner_om_per_lamp = float(om_target) / float(lamps) if lamps > 0 else 100.0
    owner_om_per_lamp = max(0.0, owner_om_per_lamp)
    elec2, om2, _tot2 = _baseline_elec_om_y1(
        lamps=lamps,
        price_per_kwh=price_per_kwh,
        watts_per_lamp=watts,
        hours_per_day=hours_per_day,
        days_per_year=days_per_year,
        owner_om_per_lamp=owner_om_per_lamp,
    )
    if abs(elec2 - elec_target) > 1.0:
        watts = float(elec_target) / denom if denom > 0 else watts
        watts = max(30.0, min(400.0, watts))
        elec2, om2, _tot2 = _baseline_elec_om_y1(
            lamps=lamps,
            price_per_kwh=price_per_kwh,
            watts_per_lamp=watts,
            hours_per_day=hours_per_day,
            days_per_year=days_per_year,
            owner_om_per_lamp=owner_om_per_lamp,
        )
    if abs(om2 - om_target) > 1.0 and lamps > 0:
        owner_om_per_lamp = float(om_target) / float(lamps)
        elec2, om2, _tot2 = _baseline_elec_om_y1(
            lamps=lamps,
            price_per_kwh=price_per_kwh,
            watts_per_lamp=watts,
            hours_per_day=hours_per_day,
            days_per_year=days_per_year,
            owner_om_per_lamp=owner_om_per_lamp,
        )
    return {
        "D6": float(lamps),
        "D8": float(price_per_kwh),
        "D9": float(watts),
        "D10": float(hours_per_day),
        "D11": float(days_per_year),
        "D12": float(owner_om_per_lamp),
        "D7": float(fee_y1),
        "D19": float(fee_y1),
    }


def _split_non_elec_opex_components(
    *,
    lamps: float,
    target_total: float,
    default_per_lamp: float,
    default_plat: float,
    default_spare: float,
) -> tuple[float, float, float]:
    base = float(lamps) * default_per_lamp + default_plat + default_spare
    if target_total <= 0:
        return 0.0, 0.0, 0.0
    if base <= 0:
        return target_total / max(lamps, 1.0), 0.0, 0.0
    s = target_total / base
    return default_per_lamp * s, default_plat * s, default_spare * s


def _laa_s_inputs_for_mode(
    *,
    mode: Literal["uniform_pct", "electricity_only_pct", "ai_plus_solar"],
    reduction_pct: float,
    baseline_elec_y1: float,
    baseline_other_y1: float,
    lamps: float,
) -> tuple[float, float, float, float, int]:
    r = max(0.0, min(0.95, float(reduction_pct)))
    if mode == "ai_plus_solar":
        return 0.999, 0.0, 0.0, 0.0, 0
    if mode == "electricity_only_pct":
        d31 = r
        per_lamp, plat, spare = _split_non_elec_opex_components(
            lamps=lamps,
            target_total=max(0.0, baseline_other_y1),
            default_per_lamp=60.0,
            default_plat=350_000.0,
            default_spare=150_000.0,
        )
        return d31, per_lamp, plat, spare, 0
    new_other = max(0.0, baseline_other_y1 * (1.0 - r))
    d31 = r if baseline_elec_y1 > 1e-6 else 0.0
    per_lamp, plat, spare = _split_non_elec_opex_components(
        lamps=lamps,
        target_total=new_other,
        default_per_lamp=60.0,
        default_plat=350_000.0,
        default_spare=150_000.0,
    )
    return d31, per_lamp, plat, spare, 0


def _simulate_provider_cashflows(
    *,
    lamps: float,
    model_years: int,
    d14: float,
    d15: float,
    d16: float,
    baseline_elec_y1: float,
    baseline_om_y1: float,
    trust_saving: float,
    trust_per_lamp_om: float,
    trust_plat: float,
    trust_spare: float,
    trust_pays_elec: int,
    laa_saving: float,
    laa_per_lamp_om: float,
    laa_plat: float,
    laa_spare: float,
    laa_pays_elec: int,
    trust_fee_y1: float,
    laa_fee_y1: float,
    trust_other_y1: float,
    laa_other_y1: float,
    trust_capex: float,
    laa_capex: float,
    laa_upfront_y0: float,
) -> tuple[list[float], list[float]]:
    n = int(model_years)

    def grow(base: float, rate: float, y: int) -> float:
        return float(base) * (1.0 + float(rate)) ** max(0, y - 1)

    def baseline_elec(y: int) -> float:
        return grow(baseline_elec_y1, d14, y)

    def baseline_om(y: int) -> float:
        return grow(baseline_om_y1, d15, y)

    trust_cf: list[float] = []
    laa_cf: list[float] = []
    for y in range(0, n + 1):
        if y == 0:
            trust_cf.append(0.0 - float(trust_capex))
            laa_cf.append(float(laa_upfront_y0) - float(laa_capex))
            continue

        be = baseline_elec(y)
        post_t = be * (1.0 - float(trust_saving))
        t_elec_cost = post_t * float(trust_pays_elec)
        t_om_cost = grow(float(lamps) * float(trust_per_lamp_om), d15, y)
        t_plat = grow(float(trust_plat), d15, y)
        t_spare = grow(float(trust_spare), d15, y)
        t_rev = grow(float(trust_fee_y1), d16, y) + grow(float(trust_other_y1), d16, y)
        t_cost = t_elec_cost + t_om_cost + t_plat + t_spare
        trust_cf.append(t_rev - t_cost)

        post_l = be * (1.0 - float(laa_saving))
        l_elec_cost = post_l * float(laa_pays_elec)
        l_om_cost = grow(float(lamps) * float(laa_per_lamp_om), d15, y)
        l_plat = grow(float(laa_plat), d15, y)
        l_spare = grow(float(laa_spare), d15, y)
        l_rev = grow(float(laa_fee_y1), d16, y) + grow(float(laa_other_y1), d16, y)
        l_cost = l_elec_cost + l_om_cost + l_plat + l_spare
        laa_cf.append(l_rev - l_cost)

    return trust_cf, laa_cf


def _npv_irr_cumulative(*, disc: float, flows: list[float]) -> tuple[float, float, float]:
    if len(flows) < 2:
        return float(flows[0]), float("nan"), float(flows[0])
    npv_v = float(npf.npv(disc, flows[1:]) + flows[0])
    irr_v = float(npf.irr(flows))
    cum = float(sum(flows))
    return npv_v, irr_v, cum


def _roi_operating_over_capex(*, flows: list[float], capex_abs: float) -> float:
    if capex_abs <= 0:
        return float("nan")
    return float(sum(flows[1:])) / float(capex_abs)


def _sheet_paths_from_zip(zin: zipfile.ZipFile) -> dict[str, str]:
    root_wb = ET.fromstring(zin.read("xl/workbook.xml"))
    root_rel = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
    rid_to_target: dict[str, str] = {}
    for rel in root_rel:
        if rel.tag.endswith("Relationship"):
            rid = rel.attrib.get("Id")
            tgt = rel.attrib.get("Target")
            if rid and tgt:
                rid_to_target[rid] = tgt
    out: dict[str, str] = {}
    for sh in root_wb.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
        name = sh.attrib.get("name")
        rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        tgt = rid_to_target.get(rid or "", "")
        if name and tgt:
            t = tgt.replace("\\", "/").lstrip("/")
            if t.startswith("xl/"):
                out[name] = t
            else:
                out[name] = "xl/" + t
    return out


def _set_cell_v(root: ET.Element, cell_ref: str, value: float) -> None:
    sheet_data = root.find(f".//{_q('sheetData')}")
    if sheet_data is None:
        raise RuntimeError("sheetData missing")
    target = None
    for row in sheet_data.findall(_q("row")):
        for c in row.findall(_q("c")):
            if c.attrib.get("r") == cell_ref:
                target = c
                break
        if target is not None:
            break
    if target is None:
        raise RuntimeError(f"cell {cell_ref} not found")
    for child in list(target):
        if child.tag == _q("v") or str(child.tag).endswith("}v"):
            target.remove(child)
    v_el = ET.Element(_q("v"))
    if value != value:  # NaN
        v_el.text = "0"
    else:
        v_el.text = str(float(value))
    target.append(v_el)


def patch_workbook_cached_values(xlsx_path: Path, updates: dict[str, dict[str, float]]) -> None:
    """Apply many cached <v> patches in one rewrite of the xlsx zip."""
    zin = zipfile.ZipFile(xlsx_path, "r")
    paths = _sheet_paths_from_zip(zin)
    files: dict[str, bytes] = {n: zin.read(n) for n in zin.namelist()}
    zin.close()

    for sheet_name, cells in updates.items():
        sp = paths[sheet_name]
        root = ET.fromstring(files[sp])
        for addr, val in cells.items():
            _set_cell_v(root, addr, val)
        files[sp] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    xlsx_path.write_bytes(buf.getvalue())


def _validate_cached_kpis(xlsx_path: Path) -> None:
    z = zipfile.ZipFile(xlsx_path, "r")
    paths = _sheet_paths_from_zip(z)

    def snip(sheet: str, addr: str) -> str:
        p = paths[sheet]
        xml = z.read(p).decode("utf-8", errors="ignore")
        i = xml.find(f'r="{addr}"')
        assert i >= 0, f"missing {sheet}!{addr}"
        return xml[i : i + 400]

    def _has_f_v(blob: str) -> bool:
        # openpyxl/ET may serialize as <f>… or <s:f>… depending on namespace prefixes
        has_f = (":f>" in blob) or ("<f>" in blob)
        has_v = ("</v>" in blob) or ("</s:v>" in blob)
        return has_f and has_v

    for addr in ("C25", "C26", "M24", "C52", "C53", "M51", "C27", "C54"):
        s = snip("05_Annual_Model", addr)
        assert _has_f_v(s), addr
    for addr in ("C18", "D18", "C19", "D19", "C20", "D20", "C21", "D21"):
        s = snip("01_Dashboard", addr)
        assert _has_f_v(s), addr
    # Product-driven inputs in 02_Inputs are formula cells and must have cached values for WPS.
    for addr in ("D28", "D31", "D32", "D33", "D34"):
        s = snip("02_Inputs", addr)
        assert _has_f_v(s), addr
    # Sensitivity (row19 B..H) should have cached values so WPS won't show same number everywhere.
    for addr in ("B19", "C19", "D19", "E19", "F19", "G19", "H19"):
        s = snip("06_Sensitivity", addr)
        assert _has_f_v(s), addr
    # First sheet savings narrative (C15:C17, D15:D17) should not be blank in WPS.
    for addr in ("C15", "C16", "C17", "D15", "D16", "D17"):
        s = snip("00_LaaS收益来源", addr)
        assert _has_f_v(s), addr
    # Incremental bridge table (cached for WPS): EMC/LaaS columns + delta column.
    for addr in (
        "C6",
        "C7",
        "C8",
        "C9",
        "C10",
        "C11",
        "C18",
        "D18",
        "E18",
        "C19",
        "D19",
        "E19",
        "C20",
        "D20",
        "E20",
        "C21",
        "D21",
        "E21",
        "C22",
        "D22",
        "E22",
        "C23",
        "D23",
        "E23",
        "C24",
        "D24",
        "E24",
        "C25",
        "D25",
        "E25",
        "C26",
        "D26",
        "E26",
        "C27",
        "D27",
        "E27",
        "C28",
        "D28",
        "E28",
    ):
        s = snip("00_LaaS收益来源", addr)
        assert _has_f_v(s), addr
    z.close()


def fill_one_tier(
    *,
    tier: TierSpec,
    baseline,
    out_path: Path,
    emc_fee_y1_tuned: float,
    discount_annual: float = 0.12,
) -> None:
    if not TEMPLATE.is_file():
        raise FileNotFoundError(str(TEMPLATE))

    # `baseline.electricity_opex_rmb_y` comes from opex.csv “改造后电费” (post-retrofit / post-saving).
    # The reference template’s baseline rows (`05_Annual_Model` row6/8) represent the *pre-saving* baseline,
    # and apply saving rates afterwards (see `04_Mode_Params` row9). Therefore we back-calculate:
    # baseline_pre_elec = post_elec / (1 - emc_saving).
    emc_saving = float(EMC_SAVING_FRAC)
    elec_post_y1 = float(EMC_AFTER_ELEC_FEE_Y1_RMB)
    cash_y1 = float(baseline.cash_opex_rmb_y.get(1))
    other_y1 = max(0.0, cash_y1 - elec_post_y1)
    capex0 = float(baseline.capex_y0_rmb)
    # `D7` is an owner-side budget cap (not used as provider revenue in our KPIs).
    # Per user requirement, assume it is above EMC fee.
    fee_y1 = float(OWNER_SERVICE_BUDGET_Y1_RMB)
    # Owner baseline physical spend in the template uses *pre-saving* electricity.
    # Use user-provided baseline as anchor.
    elec_pre_y1 = float(EMC_BASELINE_ELEC_FEE_Y1_RMB)
    baseline_owner_cost_y1 = float(elec_pre_y1 + other_y1)

    scenario = LaaSScenario(
        term_years=int(tier.term_years),
        annual_service_fee_rmb=float(tier.annual_fee_rmb),
        upfront_rmb=float(tier.upfront_rmb),
        ai_opex_reduction_pct=float(tier.ai_reduction_pct),
        last_four_year_fee_reduction_rmb=float(tier.tail_reduction_rmb),
        opex_mode=tier.opex_mode,
        product_key=tier.product_key,  # type: ignore[arg-type]
    )
    laa_res = evaluate_laas_scenario(baseline, scenario, discount_rate_annual=float(discount_annual))

    _dlog(
        "excel_template_fill_wps.py:fill_one_tier",
        "Provider KPI sanity (Python model)",
        {
            "tier": tier.name,
            "fee_y1": float(scenario.annual_service_fee_rmb),
            "upfront": float(tier.upfront_rmb),
            "tail": float(tier.tail_reduction_rmb),
            "opex_mode": str(tier.opex_mode),
            "ai_red_pct": float(tier.ai_reduction_pct),
            "capex_y0": float(capex0),
            "npv_project_rmb": float(laa_res.npv_project_rmb),
            "irr_project_annual": float(laa_res.irr_project_annual) if isinstance(laa_res.irr_project_annual, float) else None,
            "payback_months": laa_res.payback_months,
        },
        hypothesis_id="H_npv_irr_negative",
    )

    # ---------------------------------------------------------------------
    # Baseline physics: force to user-provided baseline table anchors
    # so `03_Baseline` matches the “baseline vs EMC” table exactly in WPS.
    #
    # Template formulas:
    #   03_Baseline!D10 = lamps * watts/1000 * (hours_per_day * days_per_year)
    #   03_Baseline!D12 = D10 * price_per_kwh
    # We therefore derive watts from (baseline_kwh, lamps, hours/day, days/year).
    # ---------------------------------------------------------------------
    lamps = float(EMC_LIGHTS_COUNT_Y1)
    hours = float(EMC_HOURS_PER_DAY_Y1)
    days = float(EMC_DAYS_PER_YEAR_Y1)
    baseline_kwh_y1 = float(EMC_BASELINE_KWH_Y1)
    watts = float(baseline_kwh_y1 * 1000.0 / max(1e-9, (lamps * hours * days)))
    price_per_kwh = float(EMC_BASELINE_ELEC_FEE_Y1_RMB / max(1e-9, baseline_kwh_y1))
    owner_om_per_lamp = float(other_y1 / max(1e-9, lamps))

    # `02_Inputs` cells referenced by `03_Baseline` formulas.
    phys = {
        "D6": float(lamps),
        "D8": float(price_per_kwh),
        "D9": float(watts),
        "D10": float(hours),
        "D11": float(days),
        "D12": float(owner_om_per_lamp),
    }

    shutil.copyfile(TEMPLATE, out_path)
    wb = load_workbook(out_path)
    ws_in = wb["02_Inputs"]
    ws_in["D5"] = int(tier.term_years)
    ws_in["D13"] = float(discount_annual)
    ws_in["D14"] = 0.0
    ws_in["D15"] = 0.0
    ws_in["D16"] = 0.0
    ws_in["D17"] = 0.0
    for k, v in phys.items():
        ws_in[k] = v

    # Add “poles count” for reporting (template does not use it in formulas).
    ws_in["A49"] = "灯杆数量(根)"
    ws_in["D49"] = float(EMC_POLES_COUNT_Y1)

    ws_in["D18"] = capex0 / lamps if lamps > 0 else 0.0
    tp, tpl, tsp = _split_non_elec_opex_components(
        # EMC cost components should represent non-electricity O&M (electricity handled separately by saving-rate rows).
        lamps=lamps, target_total=other_y1, default_per_lamp=80.0, default_plat=150_000.0, default_spare=50_000.0
    )
    ws_in["D22"] = float(tp)
    ws_in["D23"] = float(tpl)
    ws_in["D24"] = float(tsp)
    ws_in["D20"] = 0.0
    # Keep EMC savings at template default (0.50); enforce LaaS savings > EMC.
    ws_in["D21"] = float(emc_saving)
    # EMC assumption (fixed): owner always pays electricity under EMC.
    # Keep EMC electricity-payer toggle independent from LaaS.
    ws_in["D25"] = 0

    capex_ref_per_lamp = capex0 / lamps if lamps > 0 else 0.0
    ws_in["D29"] = float(scenario.annual_service_fee_rmb)
    ws_in["D30"] = 0.0
    d31, d32, d33, d34, d35 = _laa_s_inputs_for_mode(
        mode=tier.opex_mode,
        reduction_pct=float(tier.ai_reduction_pct),
        baseline_elec_y1=elec_pre_y1,
        baseline_other_y1=other_y1,
        lamps=lamps,
    )

    # Show poles count on `03_Baseline` (without shifting template rows/columns).
    try:
        ws_base = wb["03_Baseline"]
        ws_base["F5"] = "灯杆数量(根)"
        ws_base["G5"] = "='02_Inputs'!D49"
    except Exception:
        pass
    # Saving-rate baseline from product physics (PRODUCT_ROWS), then keep the existing narrative guardrails.
    try:
        d31 = max(float(d31), float(get_product_profile(tier.product_key).implied_grid_saving_rate()))  # type: ignore[arg-type]
    except Exception:
        pass
    # Enforce “AI > EMC” narrative: LaaS saving rate must be higher than EMC's saving rate.
    d31 = max(float(d31), float(emc_saving) + 0.03)
    # Owner-facing sensitivity assumes owner pays electricity so it varies with saving rate.
    # Keep LaaS electricity allocation at owner (0) unless a tier explicitly requires otherwise.
    d35 = 0
    laa_opex_y1 = float(laa_res.provider_cash_opex_rmb_y.get(1))
    # Provider does not bear electricity under this template fill assumption.
    elec_cost_y1 = 0.0
    non_e_target = max(0.0, laa_opex_y1 - elec_cost_y1)
    d32, d33, d34 = _split_non_elec_opex_components(
        lamps=lamps,
        target_total=non_e_target,
        default_per_lamp=float(d32) if d32 > 0 else 60.0,
        default_plat=float(d33) if d33 > 0 else 350_000.0,
        default_spare=float(d34) if d34 > 0 else 150_000.0,
    )
    # Product table (used for Excel/WPS dropdown-driven lookup).
    # Columns: product_key | saving_floor | capex_per_lamp | om_per_lamp | platform | spares | note
    ws_in["A50"] = "产品对比（来自 defaults.py PRODUCT_ROWS；切换 D48 将联动 D28/D31/D32:D34）"
    ws_in["A51"] = "产品(product_key)"
    ws_in["B51"] = "节电率下限(推导)"
    ws_in["C51"] = "LaaS CAPEX/盏"
    ws_in["D51"] = "人工/维修(元/盏/年)"
    ws_in["E51"] = "平台费(元/年)"
    ws_in["F51"] = "备件(元/年)"
    ws_in["G51"] = "说明"

    product_keys: list[str] = [
        "AI_lightning_grid",
        "AI_battery_integrated_grid",
        "AI_plus_solar_offgrid",
    ]
    product_rows: dict[str, dict[str, float]] = {}
    for ridx, pk in enumerate(product_keys, start=52):
        # Evaluate scenario under each product key to get a consistent provider non-electric OPEX target.
        scen_pk = LaaSScenario(
            term_years=int(tier.term_years),
            annual_service_fee_rmb=float(tier.annual_fee_rmb),
            upfront_rmb=float(tier.upfront_rmb),
            ai_opex_reduction_pct=float(tier.ai_reduction_pct),
            last_four_year_fee_reduction_rmb=float(tier.tail_reduction_rmb),
            opex_mode=tier.opex_mode,
            product_key=pk,  # type: ignore[arg-type]
        )
        res_pk = evaluate_laas_scenario(baseline, scen_pk, discount_rate_annual=float(discount_annual))
        opex_y1 = float(res_pk.provider_cash_opex_rmb_y.get(1))
        # Template split cells D32:D34 are NON-electric service-provider OPEX components.
        # We treat provider electricity as 0 in the template fill, so non-electric target == total cash OPEX.
        non_e_target_pk = max(0.0, opex_y1)
        per_lamp_pk, plat_pk, spare_pk = _split_non_elec_opex_components(
            lamps=lamps,
            target_total=non_e_target_pk,
            default_per_lamp=float(d32) if d32 > 0 else 60.0,
            default_plat=float(d33) if d33 > 0 else 350_000.0,
            default_spare=float(d34) if d34 > 0 else 150_000.0,
        )
        prof_pk = get_product_profile(pk)  # type: ignore[arg-type]
        saving_floor = float(prof_pk.implied_grid_saving_rate())
        capex_per_lamp_pk = float(capex_ref_per_lamp) * float(capex_scale_vs_reference(pk))  # type: ignore[arg-type]

        ws_in[f"A{ridx}"] = pk
        ws_in[f"B{ridx}"] = float(saving_floor)
        ws_in[f"C{ridx}"] = float(capex_per_lamp_pk)
        ws_in[f"D{ridx}"] = float(per_lamp_pk)
        ws_in[f"E{ridx}"] = float(plat_pk)
        ws_in[f"F{ridx}"] = float(spare_pk)
        ws_in[f"G{ridx}"] = "离网太阳能：电费≈0" if pk == "AI_plus_solar_offgrid" else ("含电池：应急/出勤更少" if pk == "AI_battery_integrated_grid" else "基础并网AI灯")

        product_rows[pk] = {
            "saving_floor": float(saving_floor),
            "capex_per_lamp": float(capex_per_lamp_pk),
            "per_lamp": float(per_lamp_pk),
            "plat": float(plat_pk),
            "spare": float(spare_pk),
        }

    # Active product values (used by Python-side cached values + provider cashflow simulation)
    active_pk = str(tier.product_key)
    pr = product_rows.get(active_pk) or product_rows.get("AI_lightning_grid") or {}
    capex_per_lamp_active = float(pr.get("capex_per_lamp", capex_ref_per_lamp))
    d31 = max(float(emc_saving) + 0.03, float(pr.get("saving_floor", 0.0)), float(d31))
    d32 = float(pr.get("per_lamp", d32))
    d33 = float(pr.get("plat", d33))
    d34 = float(pr.get("spare", d34))

    # Wire key inputs to product dropdown so changing D48 updates these cells automatically.
    lookup = "$A$52:$G$54"
    ws_in["D28"] = f"=VLOOKUP($D$48,{lookup},3,FALSE)"
    ws_in["D31"] = f"=MAX($D$21+0.03,VLOOKUP($D$48,{lookup},2,FALSE))"
    ws_in["D32"] = f"=VLOOKUP($D$48,{lookup},4,FALSE)"
    ws_in["D33"] = f"=VLOOKUP($D$48,{lookup},5,FALSE)"
    ws_in["D34"] = f"=VLOOKUP($D$48,{lookup},6,FALSE)"
    ws_in["D35"] = int(d35)

    # Style the product block to match template section headers (same palette as A18/A28 in 02_Inputs).
    header_fill = PatternFill(patternType="solid", fgColor="00E8F0F8")
    header_font = Font(name="Arial", size=11, bold=True, color="00111827")
    header_align = Alignment(horizontal="left", vertical="center")
    subhead_font = Font(name="Arial", size=10, bold=True, color="00111827")
    subhead_fill = PatternFill(patternType="solid", fgColor="00F8FAFC")
    thin = Side(style="thin", color="00D1D5DB")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row (merge across A..G)
    try:
        ws_in.merge_cells("A50:G50")
    except Exception:
        pass
    tcell = ws_in["A50"]
    tcell.fill = header_fill
    tcell.font = header_font
    tcell.alignment = header_align
    tcell.border = thin_border

    # Header row (A51..G51)
    for col in "ABCDEFG":
        c = ws_in[f"{col}51"]
        c.fill = subhead_fill
        c.font = subhead_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border

    # Data rows (A52..G54)
    for r in range(52, 55):
        for col in "ABCDEFG":
            c = ws_in[f"{col}{r}"]
            c.border = thin_border
            if col in ("A", "G"):
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    # EMC fee (provider revenue under EMC) is fixed per user requirement.
    ws_in["D19"] = float(EMC_FEE_Y1_RMB_FIXED)
    # Ensure owner budget is above EMC fee.
    ws_in["D7"] = float(OWNER_SERVICE_BUDGET_Y1_RMB)

    # Section header for commercial terms block (row 44, just above D45/D46).
    # Apply the same highlight style as the template's section headers.
    ws_in["A44"] = "商业条款/收费机制（可调整）"
    try:
        ws_in.merge_cells("A44:G44")
    except Exception:
        pass
    h44 = ws_in["A44"]
    h44.fill = PatternFill(patternType="solid", fgColor="00E8F0F8")
    h44.font = Font(name="Arial", size=11, bold=True, color="00111827")
    h44.alignment = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="00D1D5DB")
    h44.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_in["A45"] = "首期款(元)/客户预付"
    ws_in["D45"] = float(tier.upfront_rmb)
    ws_in["A46"] = "后4年年费减免(元/年)"
    ws_in["D46"] = float(tier.tail_reduction_rmb)
    ws_in["A47"] = "方案标识"
    ws_in["D47"] = tier.name
    ws_in["A48"] = "产品类型(product_key)"
    ws_in["D48"] = str(tier.product_key)
    # Data-validation dropdown for product selection (Excel/WPS).
    dv = DataValidation(
        type="list",
        formula1='"AI_lightning_grid,AI_battery_integrated_grid,AI_plus_solar_offgrid"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="无效产品类型",
        error="请从下拉框选择：AI_lightning_grid / AI_battery_integrated_grid / AI_plus_solar_offgrid",
        showInputMessage=True,
        promptTitle="产品类型(product_key)",
        prompt="用于区分不同产品的CAPEX/OPEX假设（来自 defaults.py PRODUCT_ROWS）。",
    )
    ws_in.add_data_validation(dv)
    dv.add(ws_in["D48"])

    # Sanity check: owner Y1 net savings should be positive for both EMC and LaaS (like the reference template).
    # EMC owner outflow Y1 = elec*(1-emc_saving) + fee (owner pays elec, D25=0, D17=0)
    emc_fee_y1 = float(ws_in["D19"].value or ws_in["D7"].value or 0.0)
    laa_fee_y1_local = float(ws_in["D29"].value or 0.0)
    # Owner outflow uses baseline_pre_elec * (1 - saving) per template wiring (04_Mode_Params row9).
    emc_owner_outflow_y1 = float(elec_pre_y1 * (1.0 - float(emc_saving)) + emc_fee_y1)
    laa_owner_outflow_y1 = float(elec_pre_y1 * (1.0 - float(d31)) + laa_fee_y1_local)
    emc_net_save_y1 = float(baseline_owner_cost_y1 - emc_owner_outflow_y1)
    laa_net_save_y1 = float(baseline_owner_cost_y1 - laa_owner_outflow_y1)
    _dlog(
        "excel_template_fill_wps.py:fill_one_tier",
        "Owner net savings sanity (Y1)",
        {
            "tier": tier.name,
            "baseline_owner_cost_y1": baseline_owner_cost_y1,
            "elec_pre_y1": elec_pre_y1,
            "elec_post_y1": elec_post_y1,
            "owner_om_y1": other_y1,
            "emc_saving": float(emc_saving),
            "laa_saving": float(d31),
            "emc_fee_y1": emc_fee_y1,
            "laa_fee_y1": laa_fee_y1_local,
            "emc_net_save_y1": emc_net_save_y1,
            "laa_net_save_y1": laa_net_save_y1,
        },
        hypothesis_id="H_savings_sign",
    )

    trust_saving = float(ws_in["D21"].value or 0.0)
    trust_per_lamp_om = float(ws_in["D22"].value or 0.0)
    trust_plat = float(ws_in["D23"].value or 0.0)
    trust_spare = float(ws_in["D24"].value or 0.0)
    trust_fee_y1 = float(ws_in["D19"].value or 0.0)
    laa_fee_y1 = float(ws_in["D29"].value or 0.0)
    trust_other_y1 = float(ws_in["D20"].value or 0.0)
    laa_other_y1 = float(ws_in["D30"].value or 0.0)
    trust_capex = float(ws_in["D18"].value or 0.0) * lamps
    laa_capex = float(capex_per_lamp_active) * lamps
    upfront = float(tier.upfront_rmb)

    ws_ann = wb["05_Annual_Model"]
    # Upfront: show as Y0 "other income" so it is visible in annual CF table.
    ws_ann["C41"] = "='02_Inputs'!$D$45"

    # Reflect prepay amortization + tail discount directly in the service fee revenue row (row 40, cols D..M).
    # Mirror Python convention:
    # - prepaid_per_year = upfront / term
    # - tail discount (absolute RMB) applies in years 7-10, with floor at 20% of gross annual fee
    # - service fee escalation still supported via 02_Inputs!D16 (default 0)
    prepaid = "('02_Inputs'!$D$45/'02_Inputs'!$D$5)"
    fee0 = "('04_Mode_Params'!$D$6*(1+'02_Inputs'!$D$16)^(Y-1))"
    tail = "'02_Inputs'!$D$46"
    floor = "(('04_Mode_Params'!$D$6*(1+'02_Inputs'!$D$16)^(Y-1))*0.2)"
    # For each year y=1..10 map to columns D..M
    cols = list("DEFGHIJKLM")
    for idx, col in enumerate(cols, start=1):
        y = idx
        gross = f"('04_Mode_Params'!$D$6*(1+'02_Inputs'!$D$16)^({y}-1))"
        if y >= 7:
            gross = f"MAX({floor.replace('Y', str(y))},{gross}-{tail})"
        net = f"MAX(0,{gross}-{prepaid})"
        ws_ann[f"{col}40"] = f"={net}"

    # IMPORTANT: keep reference template baseline definition:
    # `05_Annual_Model!D8 = D6 + D7` (do not add D9).

    _dlog(
        "excel_template_fill_wps.py:fill_one_tier",
        "Updated annual CF formulas for upfront+prepay+tail",
        {
            "tier": tier.name,
            "D45_upfront": float(tier.upfront_rmb),
            "D46_tail": float(tier.tail_reduction_rmb),
            "example_D40": ws_ann["D40"].value,
            "example_J40": ws_ann["J40"].value,
        },
        hypothesis_id="H1_formulas",
    )

    # Fix Dashboard IRR formatting (template shows IRR cells as money in some copies)
    ws_dash = wb["01_Dashboard"]
    ws_dash["C20"].number_format = "0.00%"
    ws_dash["D20"].number_format = "0.00%"
    ws_ann["C26"].number_format = "0.00%"
    ws_ann["C53"].number_format = "0.00%"

    # Escalation / growth knobs used by `_simulate_provider_cashflows` (defaults are usually 0 in template).
    d14 = float(ws_in["D14"].value or 0.0)
    d15 = float(ws_in["D15"].value or 0.0)
    d16 = float(ws_in["D16"].value or 0.0)

    wb.save(out_path)
    wb.close()

    # Cache the new annual row (D40:M40) so WPS shows it immediately.
    # Provider revenue series already exists in laa_res (net of prepay + tail discount), so use that as truth.
    annual_rev = [float(laa_res.provider_revenue_rmb_y.get(y, 0.0)) for y in range(1, 11)]
    # Owner-side annual cashflows for LaaS (template rows 58/59/61/62):
    # - owner pays electricity (D35=0) so owner elec outflow = baseline_elec*(1-savings)
    # - owner subscription fee outflow = annual_rev (net of prepay/tail)
    # - owner retained mgmt cost = 0 (we set 02_Inputs!D17=0)
    owner_elec = [float(elec_pre_y1) * (1.0 - float(d31)) for _y in range(1, 11)]
    owner_fee = list(annual_rev)
    owner_outflow = [owner_elec[i] + owner_fee[i] for i in range(10)]
    baseline_spend_y1 = float(baseline_owner_cost_y1)  # reference: D8 = D6 + D7
    owner_net_save = [baseline_spend_y1 - owner_outflow[i] for i in range(10)]

    # EMC owner-side annual cashflows (template rows 31/32/34/35):
    # EMC assumption fixed: owner pays electricity (D25=0).
    trust_owner_pays_elec = 1
    emc_elec = [float(elec_pre_y1) * (1.0 - float(emc_saving)) * float(trust_owner_pays_elec) for _y in range(1, 11)]
    emc_fee = [float(ws_in["D19"].value or ws_in["D7"].value or 0.0) for _y in range(1, 11)]
    emc_outflow = [emc_elec[i] + emc_fee[i] for i in range(10)]
    emc_net_save = [baseline_spend_y1 - emc_outflow[i] for i in range(10)]

    t_cf, l_cf = _simulate_provider_cashflows(
        lamps=lamps,
        model_years=int(tier.term_years),
        d14=0.0,
        d15=0.0,
        d16=0.0,
        baseline_elec_y1=elec_post_y1,
        baseline_om_y1=other_y1,
        trust_saving=trust_saving,
        trust_per_lamp_om=trust_per_lamp_om,
        trust_plat=trust_plat,
        trust_spare=trust_spare,
        trust_pays_elec=0,
        laa_saving=float(d31),
        laa_per_lamp_om=float(d32),
        laa_plat=float(d33),
        laa_spare=float(d34),
        laa_pays_elec=int(d35),
        trust_fee_y1=trust_fee_y1,
        laa_fee_y1=laa_fee_y1,
        trust_other_y1=trust_other_y1,
        laa_other_y1=laa_other_y1,
        trust_capex=trust_capex,
        laa_capex=laa_capex,
        laa_upfront_y0=upfront,
    )

    disc = float(discount_annual)
    trust_npv, trust_irr, trust_cum = _npv_irr_cumulative(disc=disc, flows=t_cf)
    _laa_npv_a, laa_irr_t, _laa_cum_a = _npv_irr_cumulative(disc=disc, flows=l_cf)

    laa_npv = float(laa_res.npv_project_rmb)
    laa_irr = float(laa_res.irr_project_annual) if isinstance(laa_res.irr_project_annual, float) else laa_irr_t
    laa_cum = float(sum(laa_res.project_cashflows_month0))

    trust_roi = _roi_operating_over_capex(flows=t_cf, capex_abs=abs(t_cf[0]) if t_cf else 0.0)
    laa_roi = _roi_operating_over_capex(flows=l_cf, capex_abs=abs(l_cf[0]) if l_cf else 0.0)

    # Owner Y1 net savings under reference baseline definition (D8 = D6 + D7):
    baseline_spend_y1 = float(baseline_owner_cost_y1)
    laas_spend_y1 = float(elec_pre_y1 * (1.0 - float(d31)) + float(annual_rev[0] if annual_rev else 0.0))
    owner_net_save_y1_ref = float(baseline_spend_y1 - laas_spend_y1)
    _dlog(
        "excel_template_fill_wps.py:fill_one_tier",
        "Owner savings sanity (reference baseline def, Y1)",
        {
            "tier": tier.name,
            "baseline_owner_cost_y1": baseline_owner_cost_y1,
            "baseline_spend_y1": baseline_spend_y1,
            "laas_spend_y1": laas_spend_y1,
            "owner_net_save_y1_ref": owner_net_save_y1_ref,
        },
        hypothesis_id="H_owner_savings_baseline",
    )

    def _fin(x: float) -> float:
        return 0.0 if x != x else float(x)

    # ------------------------------------------------------------------
    # `00_LaaS收益来源` narrative table: cache key “delta bridge” cells for WPS.
    # These rows are formula-driven in the template; without cached <v>, openpyxl
    # `data_only=True` reads blanks and the Streamlit viewer can't match Excel/WPS.
    # ------------------------------------------------------------------
    def _y1_provider_ops_cash_cost(*, pays_elec: int) -> tuple[float, float, float, float]:
        """
        Mirror `_simulate_provider_cashflows` cost split at Y=1:
          cost = elec_post * pays_elec + om + plat + spare
        where `baseline_elec_y1` passed into `_simulate_provider_cashflows` is the **post-retrofit**
        baseline electricity OPEX anchor (see `fill_one_tier`: `baseline_elec_y1=elec_post_y1`).
        """
        y = 1
        be = float(elec_post_y1) * (1.0 + float(d14)) ** max(0, y - 1)
        post = be * (1.0 - float(d31))
        elec_cost = post * float(pays_elec)
        om_cost = float(lamps) * float(d32) * (1.0 + float(d15)) ** max(0, y - 1)
        plat_cost = float(d33) * (1.0 + float(d15)) ** max(0, y - 1)
        spare_cost = float(d34) * (1.0 + float(d15)) ** max(0, y - 1)
        return float(elec_cost), float(om_cost), float(plat_cost), float(spare_cost)

    def _y1_trust_ops_cash_cost(*, pays_elec: int) -> tuple[float, float, float, float]:
        y = 1
        be = float(elec_post_y1) * (1.0 + float(d14)) ** max(0, y - 1)
        post = be * (1.0 - float(trust_saving))
        elec_cost = post * float(pays_elec)
        om_cost = float(lamps) * float(trust_per_lamp_om) * (1.0 + float(d15)) ** max(0, y - 1)
        plat_cost = float(trust_plat) * (1.0 + float(d15)) ** max(0, y - 1)
        spare_cost = float(trust_spare) * (1.0 + float(d15)) ** max(0, y - 1)
        return float(elec_cost), float(om_cost), float(plat_cost), float(spare_cost)

    emc_owner_spend_y1 = float(emc_owner_outflow_y1)
    laas_owner_spend_y1 = float(laa_owner_outflow_y1)
    emc_net_save_y1_tbl = float(baseline_owner_cost_y1 - emc_owner_spend_y1)
    laas_net_save_y1_tbl = float(baseline_owner_cost_y1 - laas_owner_spend_y1)

    emc_rev_y1 = float(trust_fee_y1) + float(trust_other_y1)
    laas_rev_y1 = float(laa_fee_y1) + float(laa_other_y1)

    emc_e1, emc_om1, emc_plat1, emc_spare1 = _y1_trust_ops_cash_cost(pays_elec=int(0))
    laas_e1, laas_om1, laas_plat1, laas_spare1 = _y1_provider_ops_cash_cost(pays_elec=int(d35))

    emc_cost_y1 = float(emc_e1 + emc_om1 + emc_plat1 + emc_spare1)
    laas_cost_y1 = float(laas_e1 + laas_om1 + laas_plat1 + laas_spare1)
    emc_nopat_cf_y1 = float(emc_rev_y1 - emc_cost_y1)
    laas_nopat_cf_y1 = float(laas_rev_y1 - laas_cost_y1)

    same_fee = bool(abs(float(laa_fee_y1) - float(trust_fee_y1)) < 1e-6)

    patch_workbook_cached_values(
        out_path,
        {
            "02_Inputs": {
                # Product-driven inputs are formula cells; cache active product values for WPS.
                "D28": float(capex_per_lamp_active),
                "D31": float(d31),
                "D32": float(d32),
                "D33": float(d33),
                "D34": float(d34),
                # Electricity allocation toggles:
                # - EMC fixed owner-pays: D25 = 0
                # - LaaS toggle: D35 = d35
                "D25": 0.0,
                "D35": float(d35),
            },
            "03_Baseline": {
                # Baseline sheet key rows are formula cells; cache so WPS shows them immediately.
                "D10": float(EMC_BASELINE_KWH_Y1),
                "D11": float(EMC_BASELINE_ELEC_FEE_Y1_RMB / max(1e-9, EMC_BASELINE_KWH_Y1)),
                "D12": float(EMC_BASELINE_ELEC_FEE_Y1_RMB),
                # D13 is owner O&M unit price; cache based on the calibrated D12 (owner_om_per_lamp) and lamps.
                "D13": float(other_y1 / max(1e-9, lamps)),
                "D14": float(other_y1),
                "D15": float(EMC_BASELINE_ELEC_FEE_Y1_RMB + other_y1),
            },
            "00_LaaS收益来源": {
                # Prevent WPS blanks in Tier5 (and others): cache key savings cells used by the first sheet narrative.
                "C15": float(emc_saving),
                "C16": float(baseline_kwh_y1 * float(emc_saving)),
                "C17": float(elec_pre_y1 * float(emc_saving)),
                "D15": float(d31),
                "D16": float(baseline_kwh_y1 * float(d31)),
                "D17": float(elec_pre_y1 * float(d31)),
                # Snapshot table (template uses **column C** for the numeric “数值/状态”; column D is explanatory text).
                "C6": float(lamps),
                "C7": float(elec_pre_y1 / max(1e-9, float(EMC_BASELINE_KWH_Y1))),
                "C8": float(EMC_BASELINE_ELEC_FEE_Y1_RMB),
                "C9": float(other_y1),
                "C10": float(OWNER_SERVICE_BUDGET_Y1_RMB),
                "C11": 1.0 if same_fee else 0.0,
                # Incremental bridge (template: C=EMC, D=LaaS, E=delta with mixed sign conventions per row)
                "C18": float(emc_owner_spend_y1),
                "D18": float(laas_owner_spend_y1),
                "E18": float(emc_owner_spend_y1 - laas_owner_spend_y1),
                "C19": float(emc_net_save_y1_tbl),
                "D19": float(laas_net_save_y1_tbl),
                "E19": float(laas_net_save_y1_tbl - emc_net_save_y1_tbl),
                "C20": float(emc_rev_y1),
                "D20": float(laas_rev_y1),
                "E20": float(laas_rev_y1 - emc_rev_y1),
                "C21": float(emc_cost_y1),
                "D21": float(laas_cost_y1),
                "E21": float(emc_cost_y1 - laas_cost_y1),
                "C22": float(emc_om1),
                "D22": float(laas_om1),
                "E22": float(emc_om1 - laas_om1),
                "C23": float(emc_plat1),
                "D23": float(laas_plat1),
                "E23": float(emc_plat1 - laas_plat1),
                "C24": float(emc_nopat_cf_y1),
                "D24": float(laas_nopat_cf_y1),
                "E24": float(laas_nopat_cf_y1 - emc_nopat_cf_y1),
                "C25": float(trust_capex),
                "D25": float(laa_capex),
                "E25": float(laa_capex - trust_capex),
                "C26": _fin(trust_cum),
                "D26": _fin(laa_cum),
                "E26": _fin(laa_cum - trust_cum),
                "C27": _fin(trust_npv),
                "D27": _fin(laa_npv),
                "E27": _fin(laa_npv - trust_npv),
                "C28": _fin(trust_irr),
                "D28": _fin(laa_irr),
                "E28": _fin((laa_irr - trust_irr) if (isinstance(laa_irr, float) and isinstance(trust_irr, float)) else float("nan")),
            },
            "06_Sensitivity": {
                # Row 19: LaaS owner net savings under electricity saving sensitivity (0.40..0.70)
                # Columns B..H correspond to 0.40,0.45,0.50,0.55,0.60,0.65,0.70 in the template.
                # Note: even if the contract allocates electricity to provider (D10=1), this sensitivity is meant
                # to show the owner's savings impact vs different saving rates; we cache assuming owner pays electricity.
                # Use net-of-prepay/tail Y1 payment (annual_rev[0]) so the row matches the annual model.
                "B19": float((elec_pre_y1 + other_y1) - (elec_pre_y1 * (1 - 0.40) + float(annual_rev[0]) + 0.0)),
                "C19": float((elec_pre_y1 + other_y1) - (elec_pre_y1 * (1 - 0.45) + float(annual_rev[0]) + 0.0)),
                "D19": float((elec_pre_y1 + other_y1) - (elec_pre_y1 * (1 - float(emc_saving)) + float(annual_rev[0]) + 0.0)),
                "E19": float((elec_pre_y1 + other_y1) - (elec_pre_y1 * (1 - 0.55) + float(annual_rev[0]) + 0.0)),
                "F19": float((elec_pre_y1 + other_y1) - (elec_pre_y1 * (1 - 0.60) + float(annual_rev[0]) + 0.0)),
                "G19": float((elec_pre_y1 + other_y1) - (elec_pre_y1 * (1 - 0.65) + float(annual_rev[0]) + 0.0)),
                "H19": float((elec_pre_y1 + other_y1) - (elec_pre_y1 * (1 - 0.70) + float(annual_rev[0]) + 0.0)),
            },
            "01_Dashboard": {
                "C18": _fin(trust_cum),
                "D18": _fin(laa_cum),
                "C19": _fin(trust_npv),
                "D19": _fin(laa_npv),
                "C20": _fin(trust_irr),
                "D20": _fin(laa_irr),
                "C21": _fin(trust_roi),
                "D21": _fin(laa_roi),
            },
            "05_Annual_Model": {
                # KPI anchors (Dashboard wiring)
                "C25": _fin(trust_npv),
                "C26": _fin(trust_irr),
                "M24": _fin(trust_cum),
                "C52": _fin(laa_npv),
                "C53": _fin(laa_irr),
                "M51": _fin(laa_cum),
                "C27": _fin(trust_roi),
                "C54": _fin(laa_roi),
                # Cache LaaS service-fee revenue row (D40..M40) as net-of-prepay+tail.
                "D40": float(annual_rev[0]),
                "E40": float(annual_rev[1]),
                "F40": float(annual_rev[2]),
                "G40": float(annual_rev[3]),
                "H40": float(annual_rev[4]),
                "I40": float(annual_rev[5]),
                "J40": float(annual_rev[6]),
                "K40": float(annual_rev[7]),
                "L40": float(annual_rev[8]),
                "M40": float(annual_rev[9]),
                # Do NOT cache row 8 (baseline total cost). It is formula-driven in the template (D8=D6+D7),
                # and caching a conflicting value causes confusion.
                # Cache LaaS owner-side rows so WPS shows net savings correctly.
                # Row 58: owner electricity outflow
                "D58": float(owner_elec[0]),
                "E58": float(owner_elec[1]),
                "F58": float(owner_elec[2]),
                "G58": float(owner_elec[3]),
                "H58": float(owner_elec[4]),
                "I58": float(owner_elec[5]),
                "J58": float(owner_elec[6]),
                "K58": float(owner_elec[7]),
                "L58": float(owner_elec[8]),
                "M58": float(owner_elec[9]),
                # Row 59: owner subscription fee outflow
                "D59": float(owner_fee[0]),
                "E59": float(owner_fee[1]),
                "F59": float(owner_fee[2]),
                "G59": float(owner_fee[3]),
                "H59": float(owner_fee[4]),
                "I59": float(owner_fee[5]),
                "J59": float(owner_fee[6]),
                "K59": float(owner_fee[7]),
                "L59": float(owner_fee[8]),
                "M59": float(owner_fee[9]),
                # Row 61: owner total outflow
                "D61": float(owner_outflow[0]),
                "E61": float(owner_outflow[1]),
                "F61": float(owner_outflow[2]),
                "G61": float(owner_outflow[3]),
                "H61": float(owner_outflow[4]),
                "I61": float(owner_outflow[5]),
                "J61": float(owner_outflow[6]),
                "K61": float(owner_outflow[7]),
                "L61": float(owner_outflow[8]),
                "M61": float(owner_outflow[9]),
                # Row 62: owner net savings (baseline spend - outflow)
                "D62": float(owner_net_save[0]),
                "E62": float(owner_net_save[1]),
                "F62": float(owner_net_save[2]),
                "G62": float(owner_net_save[3]),
                "H62": float(owner_net_save[4]),
                "I62": float(owner_net_save[5]),
                "J62": float(owner_net_save[6]),
                "K62": float(owner_net_save[7]),
                "L62": float(owner_net_save[8]),
                "M62": float(owner_net_save[9]),

                # Cache EMC owner-side rows so WPS shows EMC net savings too.
                # Row 31: EMC owner electricity outflow
                "D31": float(emc_elec[0]),
                "E31": float(emc_elec[1]),
                "F31": float(emc_elec[2]),
                "G31": float(emc_elec[3]),
                "H31": float(emc_elec[4]),
                "I31": float(emc_elec[5]),
                "J31": float(emc_elec[6]),
                "K31": float(emc_elec[7]),
                "L31": float(emc_elec[8]),
                "M31": float(emc_elec[9]),
                # Row 32: EMC fee outflow
                "D32": float(emc_fee[0]),
                "E32": float(emc_fee[1]),
                "F32": float(emc_fee[2]),
                "G32": float(emc_fee[3]),
                "H32": float(emc_fee[4]),
                "I32": float(emc_fee[5]),
                "J32": float(emc_fee[6]),
                "K32": float(emc_fee[7]),
                "L32": float(emc_fee[8]),
                "M32": float(emc_fee[9]),
                # Row 34: EMC total outflow
                "D34": float(emc_outflow[0]),
                "E34": float(emc_outflow[1]),
                "F34": float(emc_outflow[2]),
                "G34": float(emc_outflow[3]),
                "H34": float(emc_outflow[4]),
                "I34": float(emc_outflow[5]),
                "J34": float(emc_outflow[6]),
                "K34": float(emc_outflow[7]),
                "L34": float(emc_outflow[8]),
                "M34": float(emc_outflow[9]),
                # Row 35: EMC net savings
                "D35": float(emc_net_save[0]),
                "E35": float(emc_net_save[1]),
                "F35": float(emc_net_save[2]),
                "G35": float(emc_net_save[3]),
                "H35": float(emc_net_save[4]),
                "I35": float(emc_net_save[5]),
                "J35": float(emc_net_save[6]),
                "K35": float(emc_net_save[7]),
                "L35": float(emc_net_save[8]),
                "M35": float(emc_net_save[9]),
            },
        },
    )

    _dlog(
        "excel_template_fill_wps.py:fill_one_tier",
        "Cached owner rows (EMC+LaaS) for WPS",
        {
            "tier": tier.name,
            "baseline_spend_y1": baseline_spend_y1,
            "emc_outflow_y1": emc_outflow[0],
            "emc_net_save_y1": emc_net_save[0],
            "owner_outflow_y1": owner_outflow[0],
            "owner_net_save_y1": owner_net_save[0],
        },
        hypothesis_id="H_owner_cache",
    )

    _dlog(
        "excel_template_fill_wps.py:fill_one_tier",
        "Cached annual fee row for WPS",
        {"tier": tier.name, "rev_y1": annual_rev[0], "rev_y7": annual_rev[6], "rev_y10": annual_rev[9]},
        hypothesis_id="H2_cached_row",
    )

    _validate_cached_kpis(out_path)


def main() -> None:
    parsed = load_roadlight_all(ROOT / "data")
    baseline = build_baseline_energy_trust(parsed, analysis_years=10, discount_rate_annual=0.12)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    tiers = select_10_tiers(baseline=baseline)
    emc_fee_y1 = float(EMC_FEE_Y1_RMB_FIXED)

    for i, tier in enumerate(tiers, start=1):
        suffix = f"{i:02d}_{tier.name}"
        out = OUT_DIR / f"通用版_能源托管_vs_LaaS_财务模型_v3_逻辑复核版__FILLED_{suffix}_WPS.xlsx"
        fill_one_tier(tier=tier, baseline=baseline, out_path=out, emc_fee_y1_tuned=emc_fee_y1)
        print("Wrote", out)


if __name__ == "__main__":
    main()
