from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from business_model_comparison.laas_feasible import ClientValueAssumptions
from business_model_comparison.models import build_baseline_energy_trust
from business_model_comparison.roadlight_data import load_roadlight_all


ROOT = Path(__file__).resolve().parents[1]
OUT_CN = ROOT / "Dashboard_LaaS_vs_NYTG" / "Trust_to_LaaS_Investment_Model_CN.xlsx"
OUT_EN = ROOT / "Dashboard_LaaS_vs_NYTG" / "Trust_to_LaaS_Investment_Model_EN.xlsx"

# Style palette (consulting-ish)
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color=NAVY, bold=True, size=18)
SUBTITLE_FONT = Font(color="374151", size=11)
SECTION_FONT = Font(color=NAVY, bold=True, size=12)
BODY_FONT = Font(color="111827", size=10)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header_bar(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:L1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A2:L2")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")


def _section(ws, r: int, text: str, c1: int = 1, c2: int = 12) -> int:
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return r + 1


def _write_row(ws, r: int, values: list, *, bold: bool = False, fill: PatternFill | None = None) -> None:
    for c, v in enumerate(values, start=1):
        cell = ws.cell(r, c, v)
        cell.font = Font(color="111827", size=10, bold=bold)
        cell.alignment = WRAP_TOP
        if fill is not None:
            cell.fill = fill


def _write_table(ws, start_row: int, headers: list[str], rows: list[list]) -> int:
    _write_row(ws, start_row, headers, bold=True, fill=HEADER_FILL)
    for j in range(1, len(headers) + 1):
        ws.cell(start_row, j).font = HEADER_FONT
        ws.cell(start_row, j).alignment = CENTER
    r = start_row + 1
    for row in rows:
        _write_row(ws, r, row)
        r += 1
    return r


def _add_story_sources(ws, *, is_cn: bool) -> None:
    _set_col_widths(ws, {1: 22, 2: 22, 3: 56, 4: 22, 5: 52})
    title = "Story_and_Sources（故事与引用）" if is_cn else "Story_and_Sources"
    _header_bar(
        ws,
        title,
        ("Each assumption is mapped to a model lever and backed by a public source link."
         if not is_cn
         else "每个关键假设都映射到模型杠杆，并提供公开来源链接。"),
    )
    r = 4
    r = _section(ws, r, "Benchmarks and how they map into the model" if not is_cn else "外部基准与如何映射到模型")

    headers = ["Driver", "Metric (benchmark)", "How used in model", "Assumption cell", "Source (URL)"]
    if is_cn:
        headers = ["驱动因素", "指标（基准）", "如何进入模型", "假设单元格", "来源链接（URL）"]

    rows = [
        [
            "Predictive maintenance / remote diagnostics" if not is_cn else "预测性维护/远程诊断",
            "Avoided patrols / optimized repairs (qualitative + routing savings)" if not is_cn else "减少巡检、优化维修路径（定性+节省空间）",
            "Maps to reductions in staff / vehicles / materials OPEX (line-item factors)" if not is_cn else "映射到薪酬/车辆/材料等运维项的降本系数（分项）",
            "Assumptions!B8:B12",
            "https://oxmaint.com/industries/government/street-lighting-lamp-out-ai-detection",
        ],
        [
            "Adaptive controls / dimming" if not is_cn else "智能调光/自适应控制",
            "Energy savings can be large depending on baseline and dimming strategy (literature)" if not is_cn else "节能幅度与基线/调光策略强相关（文献）",
            "Maps to electricity OPEX reduction factor (electricity_only_pct mode)" if not is_cn else "映射到电费降本系数（仅电费降本模式）",
            "Assumptions!B6",
            "https://mdpi-res.com/d_attachment/smartcities/smartcities-03-00071/article_deploy/smartcities-03-00071-v2.pdf?version=1607517901",
        ],
        [
            "City case study" if not is_cn else "城市案例",
            "Energy costs cut by 61% (case; not universal)" if not is_cn else "电费降低61%（案例，不代表普遍）",
            "Used as an upper-bound sense check; model uses conservative default" if not is_cn else "作为上限 sanity check；模型默认采用保守假设",
            "Assumptions!B6",
            "https://insights.acuitybrands.com/casestudies/city-of-west-richland-case-study",
        ],
        [
            "Utility case study" if not is_cn else "电力公司案例",
            "Avoided truck rolls (evidence of O&M reduction potential)" if not is_cn else "减少出车次数（运维降本证据）",
            "Supports reductions in vehicles / staff OPEX lines (dispatch, truck rolls)" if not is_cn else "支撑车辆/人员项降本（出车/派单）",
            "Assumptions!B9:B11",
            "https://lam.itron.com/documents/d/asset-library-120736/itron-tampa-electric-case-study-1",
        ],
        [
            "Solarization (street lights)" if not is_cn else "路灯光伏化",
            "System-level feasibility context (LCC / solarization programs)" if not is_cn else "系统可行性背景（全生命周期/政策项目）",
            "Used only to justify the AI+solar tier narrative; electricity=0 is a contract/engineering assumption" if not is_cn else "仅用于支撑AI+光伏方案叙事；电费为0属于工程/合同假设",
            "Assumptions!B6 (ai_plus_solar tier)",
            "https://www.saarcenergy.org/wp-content/uploads/2022/08/Draft-report-of-EE-Solarisation-of-Street-Lighting_Rev4_04.07.2022_No-Markup.pdf",
        ],
    ]
    r_end = _write_table(ws, r, headers, rows)
    # add hyperlinks
    for rr in range(r + 1, r_end):
        url = ws.cell(rr, 5).value
        if isinstance(url, str) and url.startswith("http"):
            ws.cell(rr, 5).hyperlink = url
            ws.cell(rr, 5).style = "Hyperlink"


def _build_inputs_data(wb: openpyxl.Workbook, baseline, *, is_cn: bool) -> dict[str, str]:
    ws = wb.create_sheet("Inputs_Data" if not is_cn else "输入数据", 1)
    _set_col_widths(ws, {1: 10, 2: 18, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18})
    _header_bar(ws, "Inputs (from /data)" if not is_cn else "输入（来自/data）", "All values are RMB/year unless noted." if not is_cn else "所有金额默认为元/年（除特别说明）。")

    years = baseline.years
    headers = ["Year", "Trust_fee", "Elec_OPEX", "Staff+Other_OPEX", "Cash_OPEX", "Depreciation", "Debt_service", "CAPEX_y0"]
    if is_cn:
        headers = ["年", "托管费", "电费OPEX", "非电费OPEX", "现金OPEX", "折旧", "债务服务", "CAPEX(期初)"]

    rows = []
    for y in years:
        fee = float(baseline.revenue_rmb_y.get(y))
        elec = float(baseline.electricity_opex_rmb_y.get(y))
        cash = float(baseline.cash_opex_rmb_y.get(y))
        other = cash - elec
        dep = float(baseline.depreciation_rmb_y.get(y))
        debt = float(baseline.debt_service_rmb_y.get(y))
        rows.append([y, fee, elec, other, cash, dep, debt, 0.0])
    # add CAPEX in year0 row separate at top
    rows.insert(0, [0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, float(baseline.capex_y0_rmb)])
    r = 4
    _write_table(ws, r, headers, rows)

    # return key cell refs for formulas
    # fee for year1 in column B row 6 (since header at row4, year0 at row5, year1 at row6)
    return {
        "fee_y1": f"'{ws.title}'!B6",
        "capex_y0": f"'{ws.title}'!H5",
        "years_start": 6,
        "ws_title": ws.title,
    }


def _build_assumptions(wb: openpyxl.Workbook, *, is_cn: bool) -> dict[str, str]:
    ws = wb.create_sheet("Assumptions" if not is_cn else "假设输入", 2)
    _set_col_widths(ws, {1: 34, 2: 20, 3: 60, 4: 24})
    _header_bar(ws, "Assumptions (editable)" if not is_cn else "假设（可调整）", "Highlighted cells drive all outputs via formulas." if not is_cn else "高亮单元格驱动所有输出（全公式）。")

    r = 4
    r = _section(ws, r, "OPEX reduction factors (by line item)" if not is_cn else "分项OPEX降本系数")

    # Conservative defaults; story sheet provides citations.
    # Electricity reduction factor is used for electricity_only_pct mode (and as 100% for ai_plus_solar tiers).
    rows = [
        ["Electricity OPEX reduction %", 0.40, "Used in electricity_only_pct tier (conservative vs case studies).", "Story_and_Sources"],
        ["Staff OPEX reduction %", 0.35, "Remote monitoring + optimized dispatch; applied to non-electric OPEX.", "Story_and_Sources"],
        ["Vehicles OPEX reduction %", 0.30, "Truck-roll reduction; applied to non-electric OPEX.", "Story_and_Sources"],
        ["Materials OPEX reduction %", 0.15, "Predictive grouping reduces waste; applied to non-electric OPEX.", "Story_and_Sources"],
        ["Management OPEX reduction %", 0.10, "Automation reduces overhead; applied to non-electric OPEX.", "Story_and_Sources"],
        ["Max AI OPEX reduction cap %", 0.85, "Model guardrail: avoid unrealistic improvements; requires evidence for >85%.", "Story_and_Sources"],
    ]
    if is_cn:
        rows = [
            ["电费降本比例", 0.40, "用于“仅电费降本”方案（保守值）。", "Story_and_Sources"],
            ["人员薪酬降本比例", 0.35, "远程监控+派单优化，作用于非电费OPEX。", "Story_and_Sources"],
            ["车辆费用降本比例", 0.30, "减少出车次数，作用于非电费OPEX。", "Story_and_Sources"],
            ["维修材料降本比例", 0.15, "预测性维护降低浪费，作用于非电费OPEX。", "Story_and_Sources"],
            ["管理费用降本比例", 0.10, "平台化降低管理成本，作用于非电费OPEX。", "Story_and_Sources"],
            ["AI降本上限（防夸大）", 0.85, "护栏：>85%需额外证据。", "Story_and_Sources"],
        ]
    start = r
    _write_table(ws, r, ["Item", "Value", "Notes", "Ref"], rows)
    # style input col B
    input_fill = PatternFill("solid", fgColor="E5E7EB")
    for rr in range(start + 1, start + 1 + len(rows)):
        ws.cell(rr, 2).fill = input_fill
        ws.cell(rr, 2).number_format = "0.00%"

    # return cells
    return {
        "elec_red": f"'{ws.title}'!B{start+1}",
        "staff_red": f"'{ws.title}'!B{start+2}",
        "veh_red": f"'{ws.title}'!B{start+3}",
        "mat_red": f"'{ws.title}'!B{start+4}",
        "mgmt_red": f"'{ws.title}'!B{start+5}",
        "cap_red": f"'{ws.title}'!B{start+6}",
        "ws_title": ws.title,
    }


def _build_scenarios_10(wb: openpyxl.Workbook, inputs_ref: dict[str, str], *, is_cn: bool) -> dict[str, str]:
    ws = wb.create_sheet("Scenarios_10" if not is_cn else "10个方案", 3)
    _set_col_widths(ws, {1: 16, 2: 14, 3: 18, 4: 16, 5: 14, 6: 14, 7: 18, 8: 28})
    _header_bar(ws, "10 win-win tiers" if not is_cn else "10个共赢方案分层", "Each tier uses a distinct mechanism; all outputs are formula-driven." if not is_cn else "每个方案机制不同；输出均由公式生成。")

    base_fee_cell = inputs_ref["fee_y1"]
    # Store scenario inputs only; Model sheet computes outputs.
    headers = [
        "Tier",
        "Term_y",
        "OPEX_mode",
        "Base_fee_RMB",
        "Upfront_RMB",
        "TailDisc_RMB",
        "AI_red_pct",
        "Story",
    ]
    if is_cn:
        headers = ["方案", "年限", "降本模式", "基准服务费(元/年)", "首期款(元)", "后4年优惠(元/年)", "AI降本比例", "一句话故事"]

    tiers = [
        ["Tier1_UniformAI_20", 10, "uniform_pct", f"={base_fee_cell}*0.95", 0, 0, 0.20, "AI dispatch + predictive maintenance reduces non-electric OPEX"],
        ["Tier2_UniformAI_35", 10, "uniform_pct", f"={base_fee_cell}*0.97", 0, 0, 0.35, "Higher AI ops maturity; still within cited ranges"],
        ["Tier3_ElecOnly_40", 10, "electricity_only_pct", f"={base_fee_cell}*0.96", 0, 0, 0.40, "Adaptive dimming/controls reduce electricity cost"],
        ["Tier4_ElecOnly_60", 10, "electricity_only_pct", f"={base_fee_cell}*0.98", 0, 0, 0.60, "Aggressive controls (case-like), needs stronger acceptance"],
        ["Tier5_AISolar", 10, "ai_plus_solar", f"={base_fee_cell}*1.05", 0, 0, 0.00, "AI+Solar: electricity OPEX to 0 (engineering/contract assumption)"],
        ["Tier6_Upfront_Prepay", 10, "uniform_pct", f"={base_fee_cell}*1.00", 1_000_000, 0, 0.25, "Upfront as prepayment reduces annual fee; improves payback"],
        ["Tier7_TailDiscount", 10, "uniform_pct", f"={base_fee_cell}*1.02", 0, 250_000, 0.25, "SLA-style tail discount in last 4 years"],
        ["Tier8_HigherFee_StrongerSLA", 10, "uniform_pct", f"={base_fee_cell}*1.08", 0, 350_000, 0.30, "Higher fee but more SLA credit; keep client whole"],
        ["Tier9_MidTerm_8y", 8, "uniform_pct", f"={base_fee_cell}*1.00", 500_000, 0, 0.30, "Shorter term with upfront to keep 36m payback"],
        ["Tier10_Conservative", 10, "uniform_pct", f"={base_fee_cell}*0.92", 0, 0, 0.10, "Low fee + modest AI improvements (very conservative)"],
    ]
    if is_cn:
        # translate story
        for row in tiers:
            story = row[7]
            row[7] = {
                "AI dispatch + predictive maintenance reduces non-electric OPEX": "AI派单+预测性维护降低非电费运维",
                "Higher AI ops maturity; still within cited ranges": "更成熟的AI运维，但仍在文献可解释范围",
                "Adaptive dimming/controls reduce electricity cost": "智能调光/控制降低电费",
                "Aggressive controls (case-like), needs stronger acceptance": "更激进的调光（接近案例），需更强可行性支撑",
                "AI+Solar: electricity OPEX to 0 (engineering/contract assumption)": "AI+光伏：电费为0（工程/合同假设）",
                "Upfront as prepayment reduces annual fee; improves payback": "首期款作为预付，摊销降低年费，提升回本",
                "SLA-style tail discount in last 4 years": "后4年优惠（SLA/信用返还机制）",
                "Higher fee but more SLA credit; keep client whole": "费率更高但SLA返还更强，保持客户整体受益",
                "Shorter term with upfront to keep 36m payback": "缩短期限+首期款，保持36个月回本",
                "Low fee + modest AI improvements (very conservative)": "低费率+小幅AI降本（最保守）",
            }.get(story, story)

    r = 4
    r_end = _write_table(ws, r, headers, tiers)
    # format pct column
    for rr in range(r + 1, r_end):
        ws.cell(rr, 7).number_format = "0.00%"
    return {"ws_title": ws.title}


def _build_model_calc(wb: openpyxl.Workbook, inputs_title: str, assumptions_title: str, scenarios_title: str, *, is_cn: bool) -> None:
    ws = wb.create_sheet("Model_Calc" if not is_cn else "模型计算", 4)
    _set_col_widths(ws, {1: 12, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16, 9: 16, 10: 18, 11: 18, 12: 18})
    _header_bar(ws, "Model (formula-driven)" if not is_cn else "模型（全公式驱动）", "Each scenario computed via formulas from Inputs_Data + Assumptions." if not is_cn else "每个方案由输入数据+假设用公式推导。")

    # We build a block per tier (rows). Keep it simple: compute Year1 only + summary metrics.
    r = 4
    r = _section(ws, r, "Scenario summary metrics (computed)" if not is_cn else "方案汇总指标（计算得到）")

    headers = [
        "Tier",
        "ClientPay_Y1",
        "BaselinePay_Y1",
        "ClientSavings_Y1",
        "LaaS_CashOPEX_Y1",
        "Baseline_CashOPEX_Y1",
        "ProviderGP_Uplift_Y1",
        "Payback_months(approx)",
        "WinWin?",
    ]
    if is_cn:
        headers = ["方案", "客户Y1支付", "基线Y1支付", "客户Y1节约", "LaaS现金OPEX(Y1)", "基线现金OPEX(Y1)", "华普Y1毛利提升", "回本(月,近似)", "是否共赢"]

    # References
    # Inputs_Data: year1 row is row6
    fee_y1 = f"'{inputs_title}'!B6"
    elec_y1 = f"'{inputs_title}'!C6"
    other_y1 = f"'{inputs_title}'!D6"
    cash_opex_y1 = f"'{inputs_title}'!E6"
    dep_y1 = f"'{inputs_title}'!F6"
    capex_y0 = f"'{inputs_title}'!H5"

    # Assumptions
    elec_red = f"'{assumptions_title}'!B5"
    staff_red = f"'{assumptions_title}'!B6"
    veh_red = f"'{assumptions_title}'!B7"
    mat_red = f"'{assumptions_title}'!B8"
    mgmt_red = f"'{assumptions_title}'!B9"
    cap_red = f"'{assumptions_title}'!B10"

    rows = []
    # Scenarios_10 starts at row5 with first tier in row5? our table header at row4, so first tier row5.
    for i in range(10):
        rr = 5 + i
        tier = f"'{scenarios_title}'!A{rr}"
        term = f"'{scenarios_title}'!B{rr}"
        mode = f"'{scenarios_title}'!C{rr}"
        base_fee = f"'{scenarios_title}'!D{rr}"
        upfront = f"'{scenarios_title}'!E{rr}"
        taildisc = f"'{scenarios_title}'!F{rr}"
        ai_red = f"'{scenarios_title}'!G{rr}"

        # Annual fee net of prepay (year1 includes prepay allocation)
        fee_net_expr = f"MAX(0,{base_fee}-{upfront}/{term})"
        # Apply tail discount only affects last 4 years; year1 unaffected -> keep fee_net for y1.

        # OPEX for LaaS year1 depending on mode:
        # uniform_pct: reduce total cash opex by min(ai_red, cap_red)
        # electricity_only_pct: reduce electricity by min(ai_red, cap_red); other unchanged
        # ai_plus_solar: electricity=0; other reduced by a blended factor using staff/veh/mat/mgmt (approx)
        ai_capped_expr = f"MIN({ai_red},{cap_red})"
        opex_uniform_expr = f"{cash_opex_y1}*(1-{ai_capped_expr})"
        opex_elec_only_expr = f"({elec_y1}*(1-{ai_capped_expr}))+{other_y1}"
        # non-electric blended reduction: average of line-item factors (simple, traceable)
        non_elec_red_expr = f"AVERAGE({staff_red},{veh_red},{mat_red},{mgmt_red})"
        opex_ai_solar_expr = f"0+({other_y1}*(1-{non_elec_red_expr}))"
        laas_opex_formula = (
            f"=IF({mode}=\"uniform_pct\",{opex_uniform_expr},"
            f"IF({mode}=\"electricity_only_pct\",{opex_elec_only_expr},{opex_ai_solar_expr}))"
        )

        # Provider gross profit uplift (accounting): (fee_net - (laas_opex + dep)) - (baseline_fee - (baseline_cash_opex + dep))
        # depreciation cancels if same; keep explicit
        gp_uplift_formula = f"=(({fee_net_expr}-({laas_opex_formula[1:]}+{dep_y1}))-({fee_y1}-({cash_opex_y1}+{dep_y1})))"

        # Approx payback: capex / monthly net cash (fee_net - laas_opex)/12
        payback_m_formula = f"=IF(({fee_net_expr}-({laas_opex_formula[1:]}))<=0,999,ROUNDUP({capex_y0}/(({fee_net_expr}-({laas_opex_formula[1:]}))/12),0))"

        winwin_formula = f"=AND({fee_net_expr}<={fee_y1},{gp_uplift_formula[1:]}>=0,{payback_m_formula[1:]}<=36)"

        rows.append(
            [
                f"={tier}",
                f"={fee_net_expr}",
                f"={fee_y1}",
                f"=({fee_y1}-({fee_net_expr}))",
                laas_opex_formula,
                f"={cash_opex_y1}",
                gp_uplift_formula,
                payback_m_formula,
                winwin_formula,
            ]
        )

    r_end = _write_table(ws, r, headers, rows)

    # format money columns
    for rr in range(r + 1, r_end):
        for cc in [2, 3, 4, 5, 6, 7]:
            ws.cell(rr, cc).number_format = "#,##0"
        ws.cell(rr, 8).number_format = "0"

    # A simple chart: client payment vs baseline (Year1 across tiers)
    r = r_end + 2
    r = _section(ws, r, "Chart: Client payment vs baseline (Year 1)" if not is_cn else "图表：客户Y1支付 vs 基线", c1=1, c2=12)
    # Create a small table for charting
    ws.cell(r, 1, "Tier").fill = HEADER_FILL
    ws.cell(r, 2, "ClientPay_Y1").fill = HEADER_FILL
    ws.cell(r, 3, "BaselinePay_Y1").fill = HEADER_FILL
    for c in range(1, 4):
        ws.cell(r, c).font = HEADER_FONT
        ws.cell(r, c).alignment = CENTER
    for i in range(10):
        rr = r + 1 + i
        ws.cell(rr, 1, f"={ws.title}!A{5+i}").font = BODY_FONT
        ws.cell(rr, 2, f"={ws.title}!B{5+i}").number_format = "#,##0"
        ws.cell(rr, 3, f"={ws.title}!C{5+i}").number_format = "#,##0"

    chart = LineChart()
    chart.title = "Client payment vs baseline (Y1)"
    chart.y_axis.title = "RMB"
    chart.x_axis.title = "Tier"
    data = Reference(ws, min_col=2, min_row=r, max_col=3, max_row=r + 10)
    cats = Reference(ws, min_col=1, min_row=r + 1, max_row=r + 10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    ws.add_chart(chart, f"E{r}")


def _build_scenario_tables(
    wb: openpyxl.Workbook,
    *,
    inputs_title: str,
    assumptions_title: str,
    scenarios_title: str,
    is_cn: bool,
    horizon_years: int = 10,
) -> None:
    ws = wb.create_sheet("Scenario_Tables" if not is_cn else "方案明细表", 5)
    _set_col_widths(
        ws,
        {
            1: 8,
            2: 16,
            3: 16,
            4: 16,
            5: 16,
            6: 18,
            7: 16,
            8: 16,
            9: 16,
            10: 16,
            11: 16,
            12: 18,
        },
    )
    _header_bar(
        ws,
        "Scenario cashflow tables (all formulas)" if not is_cn else "10个方案现金流明细（全公式）",
        "Baseline vs LaaS year-by-year cashflow; same structure for all 10 tiers." if not is_cn else "逐年对比：基线 vs LaaS；10个方案结构一致。",
    )

    # Inputs row mapping: header at row4, year0 at row5, year1 at row6...
    def in_row(y: int) -> int:
        return 5 + int(y)

    fee = lambda y: f"'{inputs_title}'!B{in_row(y)}"
    elec = lambda y: f"'{inputs_title}'!C{in_row(y)}"
    other = lambda y: f"'{inputs_title}'!D{in_row(y)}"
    cash_opex = lambda y: f"'{inputs_title}'!E{in_row(y)}"

    capex_y0 = f"'{inputs_title}'!H5"
    cap_red = f"'{assumptions_title}'!B10"

    r = 4
    for i in range(10):
        scen_row = 5 + i  # in Scenarios_10
        tier = f"'{scenarios_title}'!A{scen_row}"
        term = f"'{scenarios_title}'!B{scen_row}"
        mode = f"'{scenarios_title}'!C{scen_row}"
        base_fee = f"'{scenarios_title}'!D{scen_row}"
        upfront = f"'{scenarios_title}'!E{scen_row}"
        taildisc = f"'{scenarios_title}'!F{scen_row}"
        ai_red = f"'{scenarios_title}'!G{scen_row}"

        r = _section(ws, r + 1, f"Tier: ={tier}" if not is_cn else f"方案：={tier}", c1=1, c2=12)

        headers = [
            "year",
            "trust_capex",
            "trust_fee",
            "trust_cash_opex",
            "trust_net_cf",
            "trust_cum_cf",
            "laas_capex",
            "laas_fee",
            "laas_upfront",
            "laas_cash_opex",
            "laas_net_cf",
            "laas_cum_cf",
        ]
        if is_cn:
            headers = [
                "年",
                "能源托管_CAPEX",
                "能源托管_服务费",
                "能源托管_OPEX(现金)",
                "能源托管_净现金流",
                "能源托管_累计净现金流",
                "LaaS_CAPEX",
                "LaaS服务费",
                "LaaS首期款",
                "LaaS_OPEX(现金)",
                "LaaS净现金流",
                "LaaS累计净现金流",
            ]

        _write_row(ws, r, headers, bold=True, fill=HEADER_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).font = HEADER_FONT
            ws.cell(r, c).alignment = CENTER
        r += 1

        # Common scenario expressions (no leading '=')
        ai_capped = f"MIN({ai_red},{cap_red})"
        fee_net = f"MAX(0,{base_fee}-{upfront}/{term})"

        trust_cum_prev = "0"
        laas_cum_prev = "0"
        for y in range(0, horizon_years + 1):
            # Baseline
            trust_capex = f"IF({y}=0,-{capex_y0},0)"
            trust_fee = f"IF({y}=0,0,{fee(y)})"
            trust_opex = f"IF({y}=0,0,{cash_opex(y)})"
            trust_net = f"=({trust_capex})+({trust_fee})-({trust_opex})"
            trust_cum = f"=({trust_cum_prev})+({ws.cell(r, 5).coordinate})"

            # LaaS annual fee with term cutoff + tail discount (last 4 years)
            # year within term? -> apply net fee; else 0
            # tail discount applies when y > term-4
            within_term = f"AND({y}>=1,{y}<={term})"
            tail_applies = f"AND({y}>({term}-4),{y}<={term})"
            laas_fee_y = f"IF({within_term},MAX(0,({fee_net})-IF({tail_applies},{taildisc},0)),0)"
            laas_upfront_y = f"IF({y}=0,{upfront},0)"

            # LaaS opex by mode
            opex_uniform = f"{cash_opex(y)}*(1-({ai_capped}))"
            opex_elec_only = f"({elec(y)}*(1-({ai_capped})))+{other(y)}"
            opex_ai_solar = f"0+({other(y)}*(1-AVERAGE('{assumptions_title}'!B6,'{assumptions_title}'!B7,'{assumptions_title}'!B8,'{assumptions_title}'!B9)))"
            laas_opex_y = f"IF({y}=0,0,IF({mode}=\"uniform_pct\",{opex_uniform},IF({mode}=\"electricity_only_pct\",{opex_elec_only},{opex_ai_solar})))"

            laas_capex = f"IF({y}=0,-{capex_y0},0)"
            laas_net = f"=({laas_capex})+({laas_fee_y})+({laas_upfront_y})-({laas_opex_y})"
            laas_cum = f"=({laas_cum_prev})+({ws.cell(r, 11).coordinate})"

            row_vals = [
                y,
                f"={trust_capex}",
                f"={trust_fee}",
                f"={trust_opex}",
                trust_net,
                trust_cum,
                f"={laas_capex}",
                f"={laas_fee_y}",
                f"={laas_upfront_y}",
                f"={laas_opex_y}",
                laas_net,
                laas_cum,
            ]
            _write_row(ws, r, row_vals)

            # formats
            for cc in range(2, 13):
                ws.cell(r, cc).number_format = "#,##0"

            trust_cum_prev = ws.cell(r, 6).coordinate
            laas_cum_prev = ws.cell(r, 12).coordinate
            r += 1

        r += 1


def build_workbook(*, is_cn: bool) -> Path:
    parsed = load_roadlight_all(ROOT / "data")
    baseline = build_baseline_energy_trust(parsed, analysis_years=10, discount_rate_annual=0.12)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Mirror reference workbook ordering for readability
    s00 = "00_LaaS收益来源" if is_cn else "00_Value_Drivers"
    s01 = "01_Dashboard"
    s02 = "02_Inputs"
    s03 = "03_Baseline"
    s04 = "04_Mode_Params"
    s05 = "05_Annual_Model"
    s06 = "06_Sensitivity"
    s07 = "07_Model_Checks"
    s08 = "08_Calc_Logic"
    s09 = "09_Glossary"

    # Reference-style sheets first (lightweight, formula-linked to our backbone sheets)
    ws_v = wb.create_sheet(s00, 0)
    ws_d = wb.create_sheet(s01, 1)
    ws_inputs = wb.create_sheet(s02, 2)
    ws_base = wb.create_sheet(s03, 3)
    ws_mode = wb.create_sheet(s04, 4)
    ws_annual = wb.create_sheet(s05, 5)
    ws_sens = wb.create_sheet(s06, 6)
    ws_chk = wb.create_sheet(s07, 7)
    ws_logic = wb.create_sheet(s08, 8)
    ws_gl = wb.create_sheet(s09, 9)

    # Backbone sheets
    ws_story = wb.create_sheet("Story_and_Sources", 10)
    _add_story_sources(ws_story, is_cn=is_cn)

    inputs_ref = _build_inputs_data(wb, baseline, is_cn=is_cn)
    ass_ref = _build_assumptions(wb, is_cn=is_cn)
    scen_ref = _build_scenarios_10(wb, inputs_ref, is_cn=is_cn)
    _build_model_calc(wb, inputs_ref["ws_title"], ass_ref["ws_title"], scen_ref["ws_title"], is_cn=is_cn)
    _build_scenario_tables(
        wb,
        inputs_title=inputs_ref["ws_title"],
        assumptions_title=ass_ref["ws_title"],
        scenarios_title=scen_ref["ws_title"],
        is_cn=is_cn,
        horizon_years=10,
    )

    # Populate reference-style sheets with a minimal readable skeleton linking to backbone
    # 00 Value drivers
    _set_col_widths(ws_v, {1: 28, 2: 22, 3: 22, 4: 50})
    _header_bar(ws_v, "同等条件下：LaaS 相比能源托管的收益体现在哪里" if is_cn else "Under equal conditions: where does LaaS value come from?", "增量收益拆解与回报压力提示" if is_cn else "Decompose incremental value drivers and payback pressure.")
    ws_v["A4"] = "同等条件快照" if is_cn else "Snapshot"
    ws_v["A4"].font = SECTION_FONT
    ws_v["A6"] = "基准年电费" if is_cn else "Baseline electricity (Y1)"
    ws_v["B6"] = f"='{inputs_ref['ws_title']}'!C6"
    ws_v["A7"] = "业主年度预算" if is_cn else "Owner annual budget"
    ws_v["B7"] = f"='{inputs_ref['ws_title']}'!B6"
    ws_v["A8"] = "初始CAPEX" if is_cn else "Initial CAPEX"
    ws_v["B8"] = f"='{inputs_ref['ws_title']}'!H5"

    # 01 Dashboard: headline + tier list links to Model_Calc
    _set_col_widths(ws_d, {1: 24, 2: 20, 3: 20, 4: 40})
    _header_bar(ws_d, "通用版财务模型｜能源托管 vs LaaS/AI订阅" if is_cn else "Generic model | Trust vs LaaS/AI", f"生成日期 {date.today().isoformat()}" if is_cn else f"Generated {date.today().isoformat()}")
    ws_d["A4"] = "10个方案（来自模型计算）" if is_cn else "10 tiers (from Model_Calc)"
    ws_d["A4"].font = SECTION_FONT
    model_title = "模型计算" if is_cn else "Model_Calc"
    for j, h in enumerate(["方案", "客户Y1支付", "客户Y1节约", "LaaS现金OPEX(Y1)"] if is_cn else ["Tier", "ClientPay_Y1", "Savings_Y1", "LaaS_OPEX_Y1"], start=1):
        cell = ws_d.cell(5, j, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for i in range(10):
        rr = 6 + i
        ws_d.cell(rr, 1, f"='{model_title}'!A{6+i}")
        ws_d.cell(rr, 2, f"='{model_title}'!B{6+i}")
        ws_d.cell(rr, 3, f"='{model_title}'!D{6+i}")
        ws_d.cell(rr, 4, f"='{model_title}'!E{6+i}")

    # 02 Inputs (curated) – keep simple for now
    _set_col_widths(ws_inputs, {1: 14, 2: 28, 3: 18, 4: 10, 5: 40, 6: 30})
    _header_bar(ws_inputs, "通用输入假设｜能源托管 vs LaaS/AI订阅" if is_cn else "Inputs | Trust vs LaaS/AI", "蓝色数字为可修改输入" if is_cn else "Editable inputs (blue).")

    # 03 Baseline
    _header_bar(ws_base, "现状基准测算｜不改造情形" if is_cn else "Baseline | No retrofit", "所有节省额相对本页基准" if is_cn else "All savings relative to this baseline.")
    ws_base["A5"] = "基准年电费" if is_cn else "Baseline electricity (Y1)"
    ws_base["B5"] = f"='{inputs_ref['ws_title']}'!C6"
    ws_base["A6"] = "基准年运维费(现金OPEX)" if is_cn else "Baseline cash OPEX (Y1)"
    ws_base["B6"] = f"='{inputs_ref['ws_title']}'!E6"

    # 04 Mode params / 05 Annual model / 06 Sensitivity / 07 Checks / 08 Logic / 09 Glossary: placeholders for openpyxl version
    _header_bar(ws_mode, "模式参数对比" if is_cn else "Mode params", "（openpyxl版占位，详版见WPS导出版）" if is_cn else "(placeholder; see WPS export for full layout)")
    _header_bar(ws_annual, "年度现金流模型" if is_cn else "Annual model", "(placeholder)")
    _header_bar(ws_sens, "敏感性分析" if is_cn else "Sensitivity", "(placeholder)")
    _header_bar(ws_chk, "模型逻辑复核" if is_cn else "Model checks", "(placeholder)")
    _header_bar(ws_logic, "计算逻辑说明" if is_cn else "Calc logic", "(placeholder)")
    _header_bar(ws_gl, "术语与口径说明" if is_cn else "Glossary", "(placeholder)")

    # Finishing touches
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A3"

    # Force recalculation in Excel/WPS on open (avoid empty-looking sheets due to no calc pass).
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalculation = True
    except Exception:
        pass

    out = OUT_CN if is_cn else OUT_EN
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    return out


def main() -> None:
    cn = build_workbook(is_cn=True)
    en = build_workbook(is_cn=False)
    print(f"Wrote: {cn}")
    print(f"Wrote: {en}")


if __name__ == "__main__":
    main()

