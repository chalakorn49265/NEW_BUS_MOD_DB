from __future__ import annotations

from dataclasses import asdict
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from business_model_comparison.laas_feasible import (
    ClientValueAssumptions,
    LaaSScenario,
    default_fee_grid_from_baseline,
    evaluate_laas_scenario,
    grid_search_feasible_envelope,
)
from business_model_comparison.models import build_baseline_energy_trust
from business_model_comparison.report import (
    baseline_summary_table,
    envelope_table,
    provenance_bundle,
    rank_recommended_offers,
    simple_cashflow_comparison_table,
)
from business_model_comparison.roadlight_data import load_roadlight_all


ROOT = Path(__file__).resolve().parents[1]
OUT_EN = ROOT / "Dashboard_LaaS_vs_NYTG" / "Trust_to_LaaS_Offer_Envelope.xlsx"

# Palette (corporate)
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color=NAVY, bold=True, size=18)
SUBTITLE_FONT = Font(color="374151", size=11)
SECTION_FONT = Font(color=NAVY, bold=True, size=12)
BODY_FONT = Font(color="111827", size=10)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header_bar(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")


def _section(ws, r: int, text: str, c1: int = 1, c2: int = 10) -> int:
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return r + 1


def _write_table(ws, start_row: int, start_col: int, df: pd.DataFrame) -> tuple[int, int]:
    r = start_row
    c = start_col
    # header
    for j, col in enumerate(df.columns, start=c):
        cell = ws.cell(r, j, str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    r += 1
    # body
    for _, row in df.iterrows():
        for j, col in enumerate(df.columns, start=c):
            cell = ws.cell(r, j, row[col])
            cell.font = BODY_FONT
            cell.alignment = WRAP_TOP
        r += 1
    return r, c + len(df.columns) - 1


def _pick_chart_sample(provider_df: pd.DataFrame, n_each: int = 1000) -> pd.DataFrame:
    d = provider_df[["client_gap_rmb", "npv_project_rmb"]].copy()
    d = d.sort_values("client_gap_rmb")
    neg = d[d["client_gap_rmb"] <= 0].tail(n_each)
    pos = d[d["client_gap_rmb"] > 0].head(n_each)
    out = pd.concat([neg, pos], axis=0).drop_duplicates()
    out.columns = ["client_gap_pv_rmb", "provider_npv_rmb"]
    return out


def build_workbook_en() -> Path:
    # Deterministic export defaults
    horizon_years = 10
    payback_constraint_months = 36
    discount_rate = 0.12

    fee_low = 0.35
    fee_high = 1.20
    fee_steps = 60

    upfront_grid = [0.0, 200_000.0, 500_000.0, 1_000_000.0, 2_000_000.0]
    opex_modes = ["uniform_pct", "electricity_only_pct", "ai_plus_solar"]
    reduction_grid = [0.0, 0.10, 0.20, 0.30, 0.40, 0.60, 0.80, 1.00]

    client_value = ClientValueAssumptions(
        baseline_outage_hours_per_year=30.0,
        laas_guaranteed_outage_hours_per_year=5.0,
        outage_cost_rmb_per_hour=10_000.0,
        sla_credit_share_to_client=1.0,
        client_discount_rate_annual=0.12,
    )

    parsed = load_roadlight_all(ROOT / "data")
    baseline = build_baseline_energy_trust(parsed, analysis_years=horizon_years, discount_rate_annual=discount_rate)
    fee_grid = default_fee_grid_from_baseline(baseline, pct_low=fee_low, pct_high=fee_high, steps=fee_steps)

    env = grid_search_feasible_envelope(
        baseline,
        term_years=list(range(1, horizon_years + 1)),
        annual_fee_rmb_grid=fee_grid,
        last_four_year_fee_reduction_rmb_grid=[0.0],
        upfront_rmb_grid=upfront_grid,
        ai_opex_reduction_grid=reduction_grid,
        discount_rate_annual=discount_rate,
        opex_modes=opex_modes,  # type: ignore[arg-type]
        client_value=client_value,
    )
    env_df = envelope_table(env)
    env_df["payback_ok"] = env_df["payback_months"].apply(
        lambda x: isinstance(x, int) and int(x) <= int(payback_constraint_months)
    )
    provider_df = env_df[(env_df["provider_feasible"]) & (env_df["payback_ok"])].copy()
    everyone_df = env_df[(env_df["feasible_everyone_better_off"]) & (env_df["payback_ok"])].copy()

    # Pick Top-N recommended offers (already ranked by business metrics)
    topN = rank_recommended_offers(everyone_df).head(9)
    best = None if topN.empty else topN.iloc[0].to_dict()

    best_result = None
    if best is not None:
        best_result = evaluate_laas_scenario(
            baseline,
            LaaSScenario(
                term_years=int(best["term_years"]),
                annual_service_fee_rmb=float(best["annual_service_fee_rmb"]),
                upfront_rmb=float(best["upfront_rmb"]),
                ai_opex_reduction_pct=float(best["ai_opex_reduction_pct"]),
                opex_mode=str(best["opex_mode"]),  # type: ignore[arg-type]
            ),
            discount_rate_annual=float(discount_rate),
            client_value=client_value,
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Dashboard
    ws = wb.create_sheet("Dashboard", 0)
    _set_col_widths(ws, {1: 26, 2: 22, 3: 22, 4: 22, 5: 22, 6: 18, 7: 18, 8: 18, 9: 18, 10: 18})
    _header_bar(
        ws,
        "Energy Trust → LaaS | Offer Envelope Dashboard",
        f"Roadlight case | Generated {date.today().isoformat()} | All figures trace to /data + explicit assumptions",
    )
    r = 4
    r = _section(ws, r, "Key KPIs")
    kpis = [
        ("Baseline CAPEX (RMB)", float(baseline.capex_y0_rmb)),
        ("Baseline payback (months)", str(baseline.payback_months)),
        ("Provider-feasible offers (#)", int(provider_df.shape[0])),
        ("Win-win offers (#)", int(everyone_df.shape[0])),
    ]
    for i, (k, v) in enumerate(kpis):
        rr = r + (i // 2)
        cc = 1 + (i % 2) * 5
        ws.merge_cells(start_row=rr, start_column=cc, end_row=rr, end_column=cc + 1)
        ws.cell(rr, cc, k).font = SECTION_FONT
        ws.merge_cells(start_row=rr, start_column=cc + 2, end_row=rr, end_column=cc + 4)
        val_cell = ws.cell(rr, cc + 2, v)
        val_cell.font = Font(color="111827", bold=True, size=12)
        if isinstance(v, (int, float)):
            val_cell.number_format = "#,##0"
    r += 3

    r = _section(ws, r, "Best win-win offer (max provider NPV)")
    if best is None:
        ws.cell(r, 1, "No win-win offers found under current assumptions.").font = BODY_FONT
        r += 2
    else:
        rows = [
            ("Term (years)", int(best["term_years"])),
            ("Annual service fee (RMB)", float(best["annual_service_fee_rmb"])),
            ("Upfront (RMB)", float(best["upfront_rmb"])),
            ("OPEX mode", str(best["opex_mode"])),
            ("Reduction (pct modes)", float(best["ai_opex_reduction_pct"])),
            ("Payback (months)", int(best["payback_months"])),
            ("Provider NPV (RMB)", float(best["npv_project_rmb"])),
            ("Client gap (PV RMB)", float(best["client_gap_rmb"])),
        ]
        for i, (k, v) in enumerate(rows):
            ws.cell(r + i, 1, k).font = BODY_FONT
            ws.cell(r + i, 2, v).font = Font(color="111827", bold=True, size=10)
            ws.cell(r + i, 2).number_format = "#,##0" if isinstance(v, (int, float)) else "General"
        r += len(rows) + 1

    if best_result is not None:
        r = _section(ws, r, "Trust vs LaaS yearly cashflow comparison (RMB/year)")
        comp = simple_cashflow_comparison_table(best_result, baseline)
        end_r, _ = _write_table(ws, r, 1, comp)
        r = end_r + 1

    # Client-facing: remove provider-NPV scatter chart; show a diversified tier list instead.
    r = _section(ws, r, "Win-win tiers (diversified mechanisms)")
    show = topN.copy()
    keep = [
        "term_years",
        "annual_service_fee_rmb",
        "upfront_rmb",
        "opex_mode",
        "ai_opex_reduction_pct",
        "payback_months",
        "client_gap_rmb",
        "npv_project_rmb",
    ]
    keep = [c for c in keep if c in show.columns]
    show = show[keep].head(10)
    show = show.rename(
        columns={
            "term_years": "Term (y)",
            "annual_service_fee_rmb": "Annual fee (RMB, first years)",
            "upfront_rmb": "Upfront (RMB)",
            "opex_mode": "OPEX mode",
            "ai_opex_reduction_pct": "AI reduction",
            "payback_months": "Payback (months)",
            "client_gap_rmb": "Client gap (PV RMB)",
            "npv_project_rmb": "Provider NPV (RMB)",
        }
    )
    end_r, _ = _write_table(ws, r, 1, show)
    for rr in range(r + 1, end_r):
        for cc in range(1, 1 + len(show.columns)):
            header = ws.cell(r, cc).value
            if header in ("Annual fee (RMB, first years)", "Upfront (RMB)", "Client gap (PV RMB)", "Provider NPV (RMB)"):
                ws.cell(rr, cc).number_format = "#,##0"
            if header == "AI reduction":
                ws.cell(rr, cc).number_format = "0.00%"

    # Exec summary
    ws2 = wb.create_sheet("Executive_Summary", 1)
    _set_col_widths(ws2, {1: 22, 2: 90, 3: 16})
    _header_bar(ws2, "Executive Summary", "Feasible offer envelope and trade-offs (Trust → LaaS)")
    rr = 4
    rr = _section(ws2, rr, "Objective", c1=1, c2=3)
    ws2.cell(rr, 2, "Demonstrate LaaS superiority by enumerating feasible commercial offers and highlighting win-win combinations (client benefit + provider upside), with full traceability.").alignment = WRAP_TOP

    # Assumptions
    ws3 = wb.create_sheet("Assumptions", 2)
    _set_col_widths(ws3, {1: 34, 2: 28, 3: 70})
    _header_bar(ws3, "Assumptions & Inputs", "All assumptions are explicit; inputs trace back to /data")
    rr = 4
    rr = _section(ws3, rr, "Search ranges")
    rows = [
        ("Horizon (years)", horizon_years, ""),
        ("Payback constraint (months)", payback_constraint_months, ""),
        ("Discount rate (annual)", discount_rate, ""),
        ("Service fee low (% baseline)", fee_low, ""),
        ("Service fee high (% baseline)", fee_high, ""),
        ("Fee steps", fee_steps, ""),
        ("Upfront grid (RMB)", ", ".join(f"{x:,.0f}" for x in upfront_grid), "Upfront treated as prepayment reducing annual fees."),
        ("OPEX modes", ", ".join(opex_modes), ""),
        ("Reduction grid (%)", ", ".join(f"{int(x*100)}%" for x in reduction_grid), ""),
    ]
    for i, (k, v, note) in enumerate(rows):
        ws3.cell(rr + i, 1, k).font = BODY_FONT
        ws3.cell(rr + i, 2, v).font = Font(color="111827", bold=True, size=10)
        ws3.cell(rr + i, 3, note).alignment = WRAP_TOP

    # Outputs
    ws4 = wb.create_sheet("Outputs", 3)
    _set_col_widths(ws4, {1: 18, 2: 20, 3: 14, 4: 18, 5: 16, 6: 14, 7: 14, 8: 18, 9: 18})
    _header_bar(ws4, "Outputs", "Top-N win-win offers (diversified) and offer universe sample")
    rr = 4
    rr = _section(ws4, rr, "Top-N win-win offers (ranked)")
    _write_table(ws4, rr, 1, topN)

    # Appendix
    ws5 = wb.create_sheet("Appendix_Data", 4)
    _set_col_widths(ws5, {1: 14, 2: 18, 3: 14, 4: 14, 5: 16, 6: 16, 7: 16, 8: 16, 9: 16, 10: 18})
    _header_bar(ws5, "Appendix", "Baseline yearly table + offer universe sample")
    rr = 4
    rr = _section(ws5, rr, "Baseline yearly table (RMB/year)")
    _write_table(ws5, rr, 1, baseline_summary_table(baseline))

    # Traceability
    ws6 = wb.create_sheet("Traceability", 5)
    _set_col_widths(ws6, {1: 24, 2: 110})
    _header_bar(ws6, "Traceability", "Provenance bundle (sources + transforms)")
    bundle = provenance_bundle(baseline, best_result)
    rr = 4
    rr = _section(ws6, rr, "Provenance JSON", c1=1, c2=2)
    ws6.cell(rr, 1, "bundle").font = SECTION_FONT
    ws6.cell(rr, 2, str(bundle)).alignment = WRAP_TOP

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A3"

    OUT_EN.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_EN)
    return OUT_EN


def main() -> None:
    out = build_workbook_en()
    print(f"Wrote Excel (EN): {out}")


if __name__ == "__main__":
    main()

