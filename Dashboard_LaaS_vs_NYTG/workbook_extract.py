from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _q(tag: str) -> str:
    return f"{{{NS_MAIN}}}{tag}"


def _sheet_name_to_path(z: zipfile.ZipFile) -> dict[str, str]:
    import xml.etree.ElementTree as ET

    wb = ET.fromstring(z.read("xl/workbook.xml"))
    rel = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid_to_target: dict[str, str] = {}
    for r in rel:
        if r.tag.endswith("Relationship"):
            rid = r.attrib.get("Id")
            tgt = r.attrib.get("Target")
            if rid and tgt:
                rid_to_target[rid] = tgt

    out: dict[str, str] = {}
    for sh in wb.findall(f".//{_q('sheet')}"):
        name = sh.attrib.get("name")
        rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        tgt = rid_to_target.get(rid or "")
        if not (name and tgt):
            continue
        t = tgt.replace("\\", "/").lstrip("/")
        if t.startswith("xl/"):
            out[name] = t
        else:
            out[name] = "xl/" + t
    return out


def _read_sheet_xml(z: zipfile.ZipFile, sheet_path: str) -> str:
    return z.read(sheet_path).decode("utf-8", errors="ignore")


_CELL_RE_CACHE: dict[str, re.Pattern[str]] = {}


def _cell_snip_re(addr: str) -> re.Pattern[str]:
    # Capture the whole <c ...>...</c> element for the cell.
    # We accept either <c> or <s:c> prefixes in serialized XML.
    if addr not in _CELL_RE_CACHE:
        _CELL_RE_CACHE[addr] = re.compile(
            rf'(<[^>]*\br="{re.escape(addr)}"[^>]*>.*?</[^>]*c>)',
            re.DOTALL,
        )
    return _CELL_RE_CACHE[addr]


def _extract_v(cell_xml: str) -> float | None:
    # Prefer <v> then <s:v>
    m = re.search(r"<v>(.*?)</v>", cell_xml, re.DOTALL)
    if not m:
        m = re.search(r"<s:v>(.*?)</s:v>", cell_xml, re.DOTALL)
    if not m:
        return None
    txt = (m.group(1) or "").strip()
    if txt == "":
        return None
    try:
        return float(txt)
    except Exception:
        return None


def read_cached_cell(sheet_xml: str, addr: str) -> float | None:
    m = _cell_snip_re(addr).search(sheet_xml)
    if not m:
        return None
    return _extract_v(m.group(1))


def _cols_d_to_m() -> list[str]:
    return list("DEFGHIJKLM")


def read_cached_row_c_to_m(sheet_xml: str, row: int) -> list[float | None]:
    return [read_cached_cell(sheet_xml, f"{col}{row}") for col in list("CDEFGHIJKLM")]


def _find_rows_by_labels(*, xlsx_path: Path, sheet_name: str, labels: list[str]) -> dict[str, int]:
    """
    Find row indices for label strings in a worksheet. This uses openpyxl only to
    locate static label cells (not to read numeric values).
    """
    wb = load_workbook(xlsx_path, data_only=False, read_only=True)
    ws = wb[sheet_name]
    wanted = {s.strip(): s.strip() for s in labels}
    found: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5, values_only=False):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            key = v.strip()
            if key in wanted and key not in found:
                found[key] = cell.row
        if len(found) == len(wanted):
            break
    wb.close()
    return found

def read_cached_row_d_to_m(sheet_xml: str, row: int) -> list[float | None]:
    return [read_cached_cell(sheet_xml, f"{col}{row}") for col in _cols_d_to_m()]


@dataclass(frozen=True)
class TierExtract:
    file_path: str
    tier_name: str

    # Provider KPIs (from 01_Dashboard)
    emc_npv: float | None
    emc_irr: float | None
    laas_npv: float | None
    laas_irr: float | None

    # Owner annual series (Y1..Y10) from 05_Annual_Model
    owner_spend_emc_y: list[float | None]   # row34
    owner_save_emc_y: list[float | None]    # row35
    owner_spend_laas_y: list[float | None]  # row61
    owner_save_laas_y: list[float | None]   # row62
    laas_fee_y: list[float | None]          # row40 (provider subscription fee schedule)

    # Key parameters from 02_Inputs (best-effort, cached)
    term_years: float | None
    laas_fee_y1_input: float | None
    upfront: float | None
    tail_discount: float | None
    emc_fee_y1: float | None
    emc_owner_pays_elec_flag: float | None

    # OPEX/CAPEX breakdown (CN template anchors)
    lamps: float | None  # 02_Inputs!D6
    capex_emc_per_lamp: float | None  # 02_Inputs!D18
    capex_laas_per_lamp: float | None  # 02_Inputs!D28
    baseline_electricity_y1: float | None  # 03_Baseline!D12
    electricity_price_per_kwh: float | None  # 02_Inputs!D8
    watts_per_lamp: float | None  # 02_Inputs!D9
    hours_per_day: float | None  # 02_Inputs!D10
    days_per_year: float | None  # 02_Inputs!D11
    emc_saving_rate: float | None  # 02_Inputs!D21
    laas_saving_rate: float | None  # 02_Inputs!D31
    opex_om_per_lamp: float | None  # 02_Inputs!D22
    opex_platform: float | None  # 02_Inputs!D23
    opex_spares: float | None  # 02_Inputs!D24

    # LaaS non-electric OPEX breakdown (02_Inputs!D32:D34)
    laas_opex_om_per_lamp: float | None  # 02_Inputs!D32
    laas_opex_platform: float | None  # 02_Inputs!D33
    laas_opex_spares: float | None  # 02_Inputs!D34

    # Product metadata
    product_key: str | None  # 02_Inputs!D48

    # Provider cashflow lines (Y0..Y10) from 05_Annual_Model (labels in column A/B)
    provider_lines: dict[str, list[float | None]]

    def as_dict(self) -> dict[str, Any]:
        return {
            "file_path": self.file_path,
            "tier_name": self.tier_name,
            "emc_npv": self.emc_npv,
            "emc_irr": self.emc_irr,
            "laas_npv": self.laas_npv,
            "laas_irr": self.laas_irr,
            "owner_spend_emc_y": self.owner_spend_emc_y,
            "owner_save_emc_y": self.owner_save_emc_y,
            "owner_spend_laas_y": self.owner_spend_laas_y,
            "owner_save_laas_y": self.owner_save_laas_y,
            "laas_fee_y": self.laas_fee_y,
            "term_years": self.term_years,
            "laas_fee_y1_input": self.laas_fee_y1_input,
            "upfront": self.upfront,
            "tail_discount": self.tail_discount,
            "emc_fee_y1": self.emc_fee_y1,
            "emc_owner_pays_elec_flag": self.emc_owner_pays_elec_flag,
            "lamps": self.lamps,
            "capex_emc_per_lamp": self.capex_emc_per_lamp,
            "capex_laas_per_lamp": self.capex_laas_per_lamp,
            "baseline_electricity_y1": self.baseline_electricity_y1,
            "electricity_price_per_kwh": self.electricity_price_per_kwh,
            "watts_per_lamp": self.watts_per_lamp,
            "hours_per_day": self.hours_per_day,
            "days_per_year": self.days_per_year,
            "emc_saving_rate": self.emc_saving_rate,
            "laas_saving_rate": self.laas_saving_rate,
            "opex_om_per_lamp": self.opex_om_per_lamp,
            "opex_platform": self.opex_platform,
            "opex_spares": self.opex_spares,
            "laas_opex_om_per_lamp": self.laas_opex_om_per_lamp,
            "laas_opex_platform": self.laas_opex_platform,
            "laas_opex_spares": self.laas_opex_spares,
            "product_key": self.product_key,
            "provider_lines": self.provider_lines,
        }


def extract_one_workbook(xlsx_path: Path) -> TierExtract:
    with zipfile.ZipFile(xlsx_path, "r") as z:
        name_to_path = _sheet_name_to_path(z)
        dash_xml = _read_sheet_xml(z, name_to_path["01_Dashboard"])
        ann_xml = _read_sheet_xml(z, name_to_path["05_Annual_Model"])
        inp_xml = _read_sheet_xml(z, name_to_path["02_Inputs"])
        base_xml = _read_sheet_xml(z, name_to_path["03_Baseline"])

    tier_name = xlsx_path.stem

    # KPIs
    emc_npv = read_cached_cell(dash_xml, "C19")
    emc_irr = read_cached_cell(dash_xml, "C20")
    laas_npv = read_cached_cell(dash_xml, "D19")
    laas_irr = read_cached_cell(dash_xml, "D20")

    # Annual series (Y1..Y10) in D..M
    owner_spend_emc_y = read_cached_row_d_to_m(ann_xml, 34)
    owner_save_emc_y = read_cached_row_d_to_m(ann_xml, 35)
    owner_spend_laas_y = read_cached_row_d_to_m(ann_xml, 61)
    owner_save_laas_y = read_cached_row_d_to_m(ann_xml, 62)
    laas_fee_y = read_cached_row_d_to_m(ann_xml, 40)

    # Inputs
    term_years = read_cached_cell(inp_xml, "D5")
    laas_fee_y1_input = read_cached_cell(inp_xml, "D29")
    upfront = read_cached_cell(inp_xml, "D45")
    tail_discount = read_cached_cell(inp_xml, "D46")
    emc_fee_y1 = read_cached_cell(inp_xml, "D19")
    emc_owner_pays_elec_flag = read_cached_cell(inp_xml, "D25")  # 0 means owner pays

    # OPEX/CAPEX inputs
    lamps = read_cached_cell(inp_xml, "D6")
    capex_emc_per_lamp = read_cached_cell(inp_xml, "D18")
    capex_laas_per_lamp = read_cached_cell(inp_xml, "D28")
    baseline_electricity_y1 = read_cached_cell(base_xml, "D12")
    electricity_price_per_kwh = read_cached_cell(inp_xml, "D8")
    watts_per_lamp = read_cached_cell(inp_xml, "D9")
    hours_per_day = read_cached_cell(inp_xml, "D10")
    days_per_year = read_cached_cell(inp_xml, "D11")
    emc_saving_rate = read_cached_cell(inp_xml, "D21")
    laas_saving_rate = read_cached_cell(inp_xml, "D31")
    opex_om_per_lamp = read_cached_cell(inp_xml, "D22")
    opex_platform = read_cached_cell(inp_xml, "D23")
    opex_spares = read_cached_cell(inp_xml, "D24")
    laas_opex_om_per_lamp = read_cached_cell(inp_xml, "D32")
    laas_opex_platform = read_cached_cell(inp_xml, "D33")
    laas_opex_spares = read_cached_cell(inp_xml, "D34")
    # Product key is written as an inline string in the template (no cached numeric v). We read via openpyxl for this one cell.
    try:
        wb_meta = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws_meta = wb_meta["02_Inputs"]
        pk = ws_meta["D48"].value
        product_key = str(pk).strip() if pk is not None else None
        wb_meta.close()
    except Exception:
        product_key = None

    # Provider cashflow block (mirror the workbook table you screenshot'ed)
    provider_labels = [
        "初始CAPEX支出（正数，扣减）",
        "固定服务费收入",
        "其他/第三方收入",
        "资产转让收入",
        "节电后电费总额",
        "服务商承担电费成本（正数，扣减）",
        "运维成本（正数，扣减）",
        "平台/管理成本（正数，扣减）",
        "备件/小改造储备（正数，扣减）",
        "年度总收入",
        "年度总现金成本",
        "年度净现金流",
        "累计现金流",
    ]
    label_to_row = _find_rows_by_labels(xlsx_path=xlsx_path, sheet_name="05_Annual_Model", labels=provider_labels)
    provider_lines: dict[str, list[float | None]] = {}
    for lab in provider_labels:
        r = label_to_row.get(lab)
        if r is None:
            provider_lines[lab] = [None] * 11
        else:
            provider_lines[lab] = read_cached_row_c_to_m(ann_xml, r)

    return TierExtract(
        file_path=str(xlsx_path),
        tier_name=tier_name,
        emc_npv=emc_npv,
        emc_irr=emc_irr,
        laas_npv=laas_npv,
        laas_irr=laas_irr,
        owner_spend_emc_y=owner_spend_emc_y,
        owner_save_emc_y=owner_save_emc_y,
        owner_spend_laas_y=owner_spend_laas_y,
        owner_save_laas_y=owner_save_laas_y,
        laas_fee_y=laas_fee_y,
        term_years=term_years,
        laas_fee_y1_input=laas_fee_y1_input,
        upfront=upfront,
        tail_discount=tail_discount,
        emc_fee_y1=emc_fee_y1,
        emc_owner_pays_elec_flag=emc_owner_pays_elec_flag,
        lamps=lamps,
        capex_emc_per_lamp=capex_emc_per_lamp,
        capex_laas_per_lamp=capex_laas_per_lamp,
        baseline_electricity_y1=baseline_electricity_y1,
        electricity_price_per_kwh=electricity_price_per_kwh,
        watts_per_lamp=watts_per_lamp,
        hours_per_day=hours_per_day,
        days_per_year=days_per_year,
        emc_saving_rate=emc_saving_rate,
        laas_saving_rate=laas_saving_rate,
        opex_om_per_lamp=opex_om_per_lamp,
        opex_platform=opex_platform,
        opex_spares=opex_spares,
        laas_opex_om_per_lamp=laas_opex_om_per_lamp,
        laas_opex_platform=laas_opex_platform,
        laas_opex_spares=laas_opex_spares,
        product_key=product_key,
        provider_lines=provider_lines,
    )


def discover_and_extract(*, new_models_dir: Path) -> list[TierExtract]:
    # WPS/Office temp files often start with .~ and are not valid xlsx zips.
    files = sorted(p for p in new_models_dir.glob("*.xlsx") if not p.name.startswith(".~"))
    out: list[TierExtract] = []
    for p in files:
        out.append(extract_one_workbook(p))
    return out

