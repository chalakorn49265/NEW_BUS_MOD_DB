from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "business_model_comparison" / "roadlight_four_models_owner_fixed_cn_yearly_diff_cn_backup.xlsx"
OUT = ROOT / "business_model_comparison" / "roadlight_four_models_owner_fixed_cn_yearly_diff_cn_v2.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # navy
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FONT = Font(color="1F3864", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


@dataclass(frozen=True)
class Row:
    model: str
    pays_for: str
    payment_basis: str
    ownership_residual: str
    o_and_m_risk: str
    best_fit: str
    best_fit_products: str


DIFF_ROWS: list[Row] = [
    Row(
        model="EMC（合同能源管理/EPC/ESPC）",
        pays_for="节能绩效（节能收益/节能量/节能费用）",
        payment_basis="通常与“节约的能源费用/节能量”挂钩（分享型/保证型等）",
        ownership_residual="视合同：可移交、回购或期末约定残值",
        o_and_m_risk="节能服务方承担较多技术与绩效风险；M&V要求更强",
        best_fit="有明确基线（改造前电费/能耗），且节能可被核算/验证的项目",
        best_fit_products="AI lighting；AI+Battery（若存在峰谷/容量约束，可扩展到韧性/调峰）",
    ),
    Row(
        model="能源托管（能源费用托管型）",
        pays_for="能源系统托管运营（能源费用+运维管理的打包服务）",
        payment_basis="更偏“托管费/服务费”口径（可固定/指数化；可带绩效条款）",
        ownership_residual="视合同：通常由服务方更深度运营管理；期末归属可协商",
        o_and_m_risk="服务方承担更完整的运维与费用波动管理责任（“省心+预算锁定”叙事）",
        best_fit="业主希望外包运维、锁定预算、减少管理复杂度的项目（对M&V严格度可低于EMC）",
        best_fit_products="AI lighting；AI+Battery（提升安全与应急价值，利于托管SLA）",
    ),
    Row(
        model="LaaS（Lighting as a Service/Pay-per-lux）",
        pays_for="“照明结果”作为服务（亮度/可用性/运维/平台能力）",
        payment_basis="订阅费/按服务SLA计费；可参考节能但不必以kWh为核心",
        ownership_residual="通常服务方持有设备；期末残值默认归服务方（循环回收/再部署）",
        o_and_m_risk="服务方承担O&M与SLA交付风险（故障、替换、性能达标）",
        best_fit="对“持续可用/安全合规/远程运维”更敏感的场景（加油站/防爆/矿区/私有园区）",
        best_fit_products="AI lighting（强调平台+运维）；AI+Solar+Battery（强调可用性而非电费节省）",
    ),
    Row(
        model="租赁（经营/融资/融资租赁等）",
        pays_for="资产使用权（融资工具属性更强）",
        payment_basis="租金（通常可分解为本金+利息或隐含利率）；是否含运维取决于附加服务",
        ownership_residual="出租人持有；期末可回收或按约定购买（融资租赁常见）",
        o_and_m_risk="通常承租方承担较多运维风险（除非另签托管/LaaS服务包）",
        best_fit="业主资金/预算约束、需要平滑现金流或资产获取路径的项目（但不自动带来“运营能力”）",
        best_fit_products="AI lighting / AI+Battery（当客户明确想“拿资产”且信用可接受）",
    ),
]


SOURCES = [
    ("US DOE — Energy Savings Performance Contracts (ESPC)", "https://www.energy.gov/eere/buildings/energy-savings-performance-contracts"),
    ("IEA — ESCO contracts (shared vs guaranteed savings)", "https://www.iea.org/reports/energy-service-companies-escos-2/esco-contracts"),
    ("GB/T 24915-2010 概览（含托管型/分享型/保证型等）", "https://www.waizi.org.cn/law/9788.html"),
    ("Philips + Thomas Rau “Pay per lux” (LaaS)", "https://www.architectura.be/nl/nieuws/philips-en-thomas-rau-verlichten-kantoor-duurzaam-met-pay-per-lux/"),
    ("IFRS — IFRS 16 Leases", "https://www.ifrs.org/content/dam/ifrs/publications/html-standards/english/2026/issued/ifrs16.html"),
]


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header(ws, r: int, c1: int, c2: int, text: str) -> None:
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _add_qualitative_sheet(wb: openpyxl.Workbook) -> None:
    name = "Qualitative_Comparison"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)

    _set_col_widths(
        ws,
        {
            1: 18,
            2: 22,
            3: 28,
            4: 26,
            5: 22,
            6: 30,
            7: 28,
        },
    )

    r = 1
    _header(ws, r, 1, 7, "四种商业模式：定性差异表（用于快速 sanity check）")
    r += 2
    ws.cell(r, 1, "口径提醒").font = SUBHEADER_FONT
    ws.cell(
        r,
        2,
        "当前工作簿主口径为“业主年总投入固定”。在该口径下，甲方前9年可能被设计成一致；"
        "要拉开差异，需要切换到“按合同口径计费/承担成本不同”的现金流口径，并做NPV/IRR。",
    ).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    headers = ["模式", "客户买的是什么", "计费/收入口径（模型必须区分）", "资产所有权/终值", "运维风险承担", "最适用场景", "更适配的2.0产品线"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    r += 1

    for row in DIFF_ROWS:
        ws.cell(r, 1, row.model).alignment = WRAP
        ws.cell(r, 2, row.pays_for).alignment = WRAP
        ws.cell(r, 3, row.payment_basis).alignment = WRAP
        ws.cell(r, 4, row.ownership_residual).alignment = WRAP
        ws.cell(r, 5, row.o_and_m_risk).alignment = WRAP
        ws.cell(r, 6, row.best_fit).alignment = WRAP
        ws.cell(r, 7, row.best_fit_products).alignment = WRAP
        r += 1

    r += 2
    _header(ws, r, 1, 7, "客户/项目类型 × 商业模式 × 产品（建议映射）")
    r += 1
    ws.cell(r, 1, "项目类型").font = SUBHEADER_FONT
    ws.cell(r, 2, "更推荐的模式").font = SUBHEADER_FONT
    ws.cell(r, 3, "更推荐的产品").font = SUBHEADER_FONT
    ws.cell(r, 4, "一句话理由（你可以直接拿去讲）").font = SUBHEADER_FONT
    for c in range(1, 8):
        ws.cell(r, c).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    r += 1

    scenarios = [
        ("城市照明（政府/市政局）", "EMC / 能源托管", "AI lighting；AI+Battery（若有应急诉求）", "基线与预算清晰：适合绩效合同或托管锁定预算；电费与运维可量化。"),
        ("电力公司型业主（国网/私营utility）", "EMC（保证/分享）+（可选）托管", "AI lighting；AI+Battery", "关注系统效率与负荷：可用保证节能/TOU与削峰填谷讲清价值。"),
        ("加油站/防爆场景", "LaaS", "AI lighting（强调运维/合规）；AI+Solar+Battery（无电缆/孤岛）", "客户更买“可靠与合规+省心运维”，不想自己养团队。"),
        ("矿区/大型园区（私有）", "LaaS / 托管", "AI lighting；AI+Battery", "追求 uptime、安全与运维外包；SLA与响应速度可成为溢价点。"),
        ("rural/离网/不铺地下电缆", "LaaS（优先）/ 租赁（次选）", "AI+Solar+Battery", "核心价值在“避免电缆CAPEX+持续可用”，而不是虚拟电费节省。"),
    ]
    for proj, model, product, why in scenarios:
        ws.cell(r, 1, proj).alignment = WRAP
        ws.cell(r, 2, model).alignment = WRAP
        ws.cell(r, 3, product).alignment = WRAP
        ws.cell(r, 4, why).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        r += 1

    r += 2
    _header(ws, r, 1, 7, "外部参考（用于校验概念，不构成会计/法律意见）")
    r += 1
    for title, url in SOURCES:
        ws.cell(r, 1, title).alignment = WRAP
        ws.cell(r, 2, url)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        r += 1


def _add_sensitivity_sheet(wb: openpyxl.Workbook) -> None:
    name = "Sensitivity_Check"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    _set_col_widths(ws, {1: 30, 2: 18, 3: 18, 4: 18, 5: 18, 6: 38})

    r = 1
    _header(ws, r, 1, 6, "敏感性/DCF 检查（快速判断：四模式是否真的不同）")
    r += 2

    ws.cell(r, 1, "说明").font = SUBHEADER_FONT
    ws.cell(
        r,
        2,
        "本表用“逐年利润对比”中相同口径的数据做NPV对比。若四模式NPV仍一致，说明模型仍未引入关键差异（计费口径、O&M承担、残值归属、租赁利息/结构等）。",
    ).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 2

    # Inputs
    ws.cell(r, 1, "折现率（年，建议用于NPV）").font = SUBHEADER_FONT
    ws.cell(r, 2, "0.12")
    ws.cell(r, 2).number_format = "0.00%"
    ws.cell(r, 6, "可改为业主WACC/资金方要求回报/项目贴现率").alignment = WRAP
    r += 2

    # NPV table headers
    headers = ["指标（NPV基于年序列）", "EMC", "能源托管", "LaaS", "租赁", "取数说明"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    r += 1

    rate = "$B$5"
    # Owner incremental benefit NPV: 从逐年利润对比 sheet rows 6-15 columns B-E
    ws.cell(r, 1, "甲方增量收益 NPV（Y1–Y10）").alignment = WRAP
    ws.cell(r, 2, f"=NPV({rate},'逐年利润对比'!B6:B15)")
    ws.cell(r, 3, f"=NPV({rate},'逐年利润对比'!C6:C15)")
    ws.cell(r, 4, f"=NPV({rate},'逐年利润对比'!D6:D15)")
    ws.cell(r, 5, f"=NPV({rate},'逐年利润对比'!E6:E15)")
    ws.cell(r, 6, "表一：甲方逐年增量收益对比").alignment = WRAP
    r += 1

    # Core service net operating cash NPV: rows 21-30 columns B-E
    ws.cell(r, 1, "核心服务方净经营现金 NPV（Y1–Y10）").alignment = WRAP
    ws.cell(r, 2, f"=NPV({rate},'逐年利润对比'!B21:B30)")
    ws.cell(r, 3, f"=NPV({rate},'逐年利润对比'!C21:C30)")
    ws.cell(r, 4, f"=NPV({rate},'逐年利润对比'!D21:D30)")
    ws.cell(r, 5, f"=NPV({rate},'逐年利润对比'!E21:E30)")
    ws.cell(r, 6, "表二：核心服务方逐年净经营现金对比（未扣建设期股权）").alignment = WRAP
    r += 2

    # Checklist: must-differ drivers
    _header(ws, r, 1, 6, "Sanity checklist（若做不到，这两列就会‘看起来一样’）")
    r += 1
    bullets = [
        "EMC vs 能源托管：至少要在【计费口径】（节能分享 vs 托管费）或【风险承担】（能耗/运维）上有硬差异。",
        "LaaS vs 租赁：至少要在【SLA+运维打包】、【残值默认归属】、【支付结构=服务费 vs 本金/利息】上有硬差异。",
        "AI+Solar：不要只用“虚拟电费节省”，要显式建模【避免电缆/土建CAPEX】与【可用性/运维服务】价值。",
        "所有模式：用同一折现率做NPV对比；再做 ±10–20% 电价/运维/利用小时/残值 的 tornado。",
    ]
    for b in bullets:
        ws.cell(r, 1, "•")
        ws.cell(r, 2, b).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1


def _add_improvement_sheet(wb: openpyxl.Workbook) -> None:
    name = "Model_Improvements"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 2)
    _set_col_widths(ws, {1: 26, 2: 34, 3: 40, 4: 40})

    r = 1
    _header(ws, r, 1, 4, "模型改进清单（把“四模式差异”落到可计算的变量上）")
    r += 2

    ws.cell(r, 1, "现状诊断（为何曲线会一样）").font = SUBHEADER_FONT
    ws.cell(
        r,
        2,
        "在 `假设` 表中当前锁定了“业主年总投入（恒定）=B16”。因此甲方年度流出在四模式下被强制相同，"
        "导致 `逐年利润对比` 的甲方前9年天然一致；同时若托管与LaaS的费用口径/承担方设定接近，就会出现两列相同。",
    ).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 2

    _header(ws, r, 1, 4, "必须引入的“硬差异驱动”（否则只是换名字）")
    r += 1

    items = [
        (
            "EMC vs 能源托管",
            "至少 2 个硬差异",
            "1) 计费口径：EMC = 节能分享/保证（随kWh/费用变化）；托管 = 托管费（可固定/指数化）\n"
            "2) 风险承担：托管承担能耗与运维波动更完整；EMC更偏节能绩效风险\n"
            "3) M&V：EMC显式设置基线、核算、M&V成本；托管可弱化或打包",
            "在 `假设` 新增：\n- EMC节能分享比例（%）\n- 托管费（$/年）或 托管费=基线×(1-保证节省率)\n- M&V成本（$/年）\n并在各模式表把 K列“收入”改为相应口径。",
        ),
        (
            "LaaS vs 租赁",
            "至少 2 个硬差异",
            "1) SLA/可用性：LaaS按可用性/亮度交付（含罚则）；租赁不自带SLA\n"
            "2) O&M承担：LaaS必须服务方承担运维与更换；租赁通常承租方承担（除非另签托管）\n"
            "3) 支付结构：租赁 = 本金+利息/隐含利率；LaaS = 服务费（可指数化/按性能）\n"
            "4) 残值默认：LaaS残值默认归服务方（循环回收），租赁按合同回收/回购",
            "在 `假设` 新增：\n- LaaS SLA罚则（%收入或$/年）\n- LaaS运维/更换预算（$/年，随寿命/故障率）\n- 租赁隐含利率/租金结构（可引用 `融资摊还` ）\n并在 LaaS/租赁 模式表分别反映。",
        ),
        (
            "AI+Solar（含电池）",
            "避免“虚拟电费”",
            "若客户本来就不打算用市电，‘节省电费’是虚拟基线；应转为：\n"
            "- 避免电缆/土建CAPEX（对比市电方案）\n"
            "- 可用性/运维服务（尤其离网/偏远）\n"
            "- 远程AI运维带来的人员/响应成本下降",
            "新增一套“对比基线选择器”:\n- Baseline=A: 市电方案（含电缆）\n- Baseline=B: 纯光伏方案（无电缆）\n并为两种baseline分别定义 CAPEX/OPEX，再决定哪类场景用哪个baseline讲故事。",
        ),
        (
            "DCF/敏感性",
            "把‘利润’变成可决策指标",
            "对每个参与方（甲方/服务方/资金方）输出：NPV、IRR、回收期，并对关键输入做±10–20%敏感性。",
            "已在 `Sensitivity_Check` 提供NPV骨架。下一步把：\n- 折现率\n- 电价/运维/利用小时\n- 残值/更换周期\n- SLA罚则\n做成可切换情景 + tornado 图。",
        ),
    ]

    # headers
    ws.cell(r, 1, "模块").fill = HEADER_FILL
    ws.cell(r, 2, "目标").fill = HEADER_FILL
    ws.cell(r, 3, "要引入的差异（定性→定量）").fill = HEADER_FILL
    ws.cell(r, 4, "落地到工作簿（建议改哪里）").fill = HEADER_FILL
    for c in range(1, 5):
        ws.cell(r, c).font = HEADER_FONT
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    r += 1

    for mod, goal, diff, how in items:
        ws.cell(r, 1, mod).alignment = WRAP
        ws.cell(r, 2, goal).alignment = WRAP
        ws.cell(r, 3, diff).alignment = WRAP
        ws.cell(r, 4, how).alignment = WRAP
        r += 1


def main() -> None:
    if not SRC.is_file():
        raise SystemExit(f"Missing source workbook: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=False)
    _add_qualitative_sheet(wb)
    _add_sensitivity_sheet(wb)
    _add_improvement_sheet(wb)
    wb.save(OUT)
    print(f"Wrote: {OUT}")


if __name__ == "__main__":
    main()

