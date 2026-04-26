from __future__ import annotations

from dataclasses import asdict
from datetime import date
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from business_model_comparison.laas_feasible import (
    ClientValueAssumptions,
    LaaSScenario,
    MAX_AI_OPEX_REDUCTION_PCT,
    MIN_LAST_FOUR_YEAR_FEE_PCT,
    default_fee_grid_from_baseline,
    evaluate_laas_scenario,
    grid_search_feasible_envelope,
)
from business_model_comparison.models import build_baseline_energy_trust
from business_model_comparison.report import (
    baseline_summary_table,
    envelope_table,
    provenance_bundle,
    rank_recommended_offers,
    simple_cashflow_comparison_table,
)
from business_model_comparison.roadlight_data import load_roadlight_all


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Dashboard_LaaS_vs_NYTG" / "Trust_to_LaaS_Offer_Envelope.xlsx"
OUT_CN = ROOT / "Dashboard_LaaS_vs_NYTG" / "Trust_to_LaaS_Offer_Envelope_CN.xlsx"

# Style palette (matches existing repo navy theme + excel_instruction.md)
NAVY = "1F3864"
TEAL = "0F766E"
LIGHT_GRAY = "F3F4F6"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color=NAVY, bold=True, size=18)
SUBTITLE_FONT = Font(color="374151", size=11)
SECTION_FONT = Font(color=NAVY, bold=True, size=12)
BODY_FONT = Font(color="111827", size=10)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

OPEX_MODE_CN = {
    "uniform_pct": "整体运维降本",
    "electricity_only_pct": "仅电费降本",
    "ai_plus_solar": "AI+光伏",
}

BOOL_CN = {True: "是", False: "否"}


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header_bar(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")


def _section(ws, r: int, text: str, c1: int = 1, c2: int = 10) -> int:
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return r + 1


def _write_table(ws, start_row: int, start_col: int, df: pd.DataFrame, *, header_fill: PatternFill = HEADER_FILL) -> tuple[int, int]:
    r = start_row
    c = start_col

    # Header row
    for j, col in enumerate(df.columns, start=c):
        cell = ws.cell(r, j, str(col))
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    r += 1

    # Body
    for _, row in df.iterrows():
        for j, col in enumerate(df.columns, start=c):
            val = row[col]
            cell = ws.cell(r, j, val)
            cell.font = BODY_FONT
            cell.alignment = WRAP_TOP
        r += 1

    return r, c + len(df.columns) - 1


def _rename_columns(df: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
    return df.rename(columns={k: v for k, v in mapping.items() if k in df.columns})


def _localize_simple_comparison(df: pd.DataFrame) -> pd.DataFrame:
    return _rename_columns(
        df,
        {
            "year": "年份",
            "trust_capex_rmb": "能源托管_CAPEX",
            "trust_service_fee_rmb": "能源托管_服务费",
            "trust_upfront_rmb": "能源托管_首期款",
            "trust_cash_opex_rmb": "能源托管_OPEX(现金)",
            "trust_net_cashflow_rmb": "能源托管_净现金流",
            "trust_net_cashflow_cumulative_rmb": "能源托管_累计净现金流",
            "laas_capex_rmb": "LaaS_CAPEX",
            "laas_service_fee_rmb": "LaaS服务费",
            "laas_upfront_rmb": "LaaS首期款",
            "laas_cash_opex_rmb": "LaaS_OPEX(现金)",
            "laas_net_cashflow_rmb": "LaaS净现金流",
            "laas_net_cashflow_cumulative_rmb": "LaaS累计净现金流",
        },
    )


def _localize_baseline_cashflow(df: pd.DataFrame) -> pd.DataFrame:
    return _rename_columns(
        df,
        {
            "year": "年份",
            "capex_rmb": "CAPEX",
            "service_fee_rmb": "服务费",
            "upfront_rmb": "首期款",
            "net_cashflow_cumulative_rmb": "累计净现金流",
        },
    )


def _localize_envelope_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "opex_mode" in out.columns:
        out["opex_mode"] = out["opex_mode"].map(lambda x: OPEX_MODE_CN.get(str(x), str(x)))
    for col in [
        "meets_pay_less_each_year",
        "meets_provider_gross_profit_each_year",
        "meets_payback_36m",
        "meets_payback_faster_than_baseline",
        "provider_feasible",
        "client_benefit_pass",
        "feasible_everyone_better_off",
    ]:
        if col in out.columns:
            out[col] = out[col].map(lambda x: BOOL_CN.get(bool(x), str(x)) if pd.notna(x) else x)
    return _rename_columns(
        out,
        {
            "term_years": "合同年限",
            "annual_service_fee_rmb": "前6年年服务费",
            "average_client_payment_rmb_per_year": "平均年服务费",
            "last_four_year_fee_reduction_rmb": "后4年年降幅",
            "upfront_rmb": "首期款",
            "ai_opex_reduction_pct": "AI降本比例",
            "opex_mode": "降本模式",
            "payback_months": "回本周期(月)",
            "irr_project_annual": "项目IRR",
            "npv_project_rmb": "项目NPV",
            "dscr_min": "最低DSCR",
            "meets_pay_less_each_year": "客户逐年更优",
            "meets_provider_gross_profit_each_year": "华普逐年毛利更优",
            "meets_payback_36m": "回本≤36个月",
            "meets_payback_faster_than_baseline": "较基线回本更快",
            "provider_feasible": "华普可行",
            "client_benefit_pass": "客户受益通过",
            "client_gap_rmb": "客户价值缺口(PV)",
            "baseline_client_npv_cost_rmb": "基线客户成本PV",
            "laas_client_npv_cost_rmb": "LaaS客户成本PV",
            "guarantees_npv_value_rmb": "保障价值PV",
            "min_client_savings_rmb_per_year": "客户最少年节省",
            "min_provider_gross_profit_uplift_rmb_per_year": "华普最少年毛利提升",
            "payback_improvement_months": "回本改善(月)",
            "feasible_everyone_better_off": "双方共赢",
        },
    )


def pareto_frontier(df: pd.DataFrame, *, x_col: str, y_col: str) -> pd.DataFrame:
    """Return nondominated points where we maximize y and minimize x.

    Dominance: A dominates B if (y_A >= y_B and x_A <= x_B) and at least one strict.
    """
    d = df[[x_col, y_col] + [c for c in df.columns if c not in (x_col, y_col)]].copy()
    d = d.replace([np.inf, -np.inf], np.nan).dropna(subset=[x_col, y_col])
    d = d.sort_values([x_col, y_col], ascending=[True, False]).reset_index(drop=True)

    frontier_idx: list[int] = []
    best_y = None
    for i, row in d.iterrows():
        x = float(row[x_col])
        y = float(row[y_col])
        if best_y is None or y > best_y + 1e-9:
            frontier_idx.append(int(i))
            best_y = y
        else:
            # dominated by a previous point with <=x and >=y
            continue
    return d.loc[frontier_idx].reset_index(drop=True)


def _top_n_diversified(
    df: pd.DataFrame,
    *,
    n: int = 9,
    min_per_mode: int = 2,
    pct_cap: float = 0.40,
) -> pd.DataFrame:
    """Pick Top-N rows by provider NPV with diversification.

    - Enforce at least `min_per_mode` per opex_mode where possible.
    - Cap ai_opex_reduction_pct for pct modes (uniform_pct / electricity_only_pct) to avoid only 100% solutions.
    - Allow ai_plus_solar separately (cap not applied).
    """
    if df.empty:
        return df

    d = df.copy()
    d["is_pct_mode"] = d["opex_mode"].isin(["uniform_pct", "electricity_only_pct"])
    d = d[(~d["is_pct_mode"]) | (d["ai_opex_reduction_pct"] <= float(pct_cap) + 1e-9)]

    # Rank by provider NPV
    d = d.sort_values("npv_project_rmb", ascending=False)

    picked = []
    for mode in ["uniform_pct", "electricity_only_pct", "ai_plus_solar"]:
        dm = d[d["opex_mode"] == mode].head(int(min_per_mode))
        picked.append(dm)
    base = pd.concat(picked, axis=0).drop_duplicates()

    remaining = d[~d.index.isin(base.index)].head(max(0, int(n) - len(base)))
    out = pd.concat([base, remaining], axis=0).drop_duplicates().head(int(n)).reset_index(drop=True)
    return out


def _summary_block(
    baseline_fee_rmb_y1: float,
    baseline_gp_rmb_y1: float,
    baseline_payback_months: Any,
    offer: dict[str, Any],
) -> pd.DataFrame:
    """Build a screenshot-style comparison block (values in 万元 unless noted)."""
    # Convert RMB → 万元 for display
    base_fee_10k = float(baseline_fee_rmb_y1) / 10_000.0
    rec_fee_10k = float(offer.get("average_client_payment_rmb_per_year", offer.get("net_client_payment_y1_rmb", offer["annual_service_fee_rmb"]))) / 10_000.0
    owner_save_10k = base_fee_10k - rec_fee_10k

    base_gp_10k = float(baseline_gp_rmb_y1) / 10_000.0
    rec_gp_10k = float(offer.get("gp_year1_rmb", 0.0)) / 10_000.0

    base_pb_y = (float(baseline_payback_months) / 12.0) if isinstance(baseline_payback_months, int) else None
    rec_pb_y = (float(offer["payback_months"]) / 12.0) if isinstance(offer["payback_months"], int) else None

    rows = [
        ("业主年度总支出（万元）", base_fee_10k, rec_fee_10k, rec_fee_10k - base_fee_10k, "业主", "下降" if owner_save_10k > 0 else "上升"),
        ("业主年度节约（万元）", 0.0, owner_save_10k, owner_save_10k, "业主", "放大收益" if owner_save_10k > 0 else "减少"),
        ("华普年毛空间（万元）", base_gp_10k, rec_gp_10k, rec_gp_10k - base_gp_10k, "华普", "提升" if rec_gp_10k >= base_gp_10k else "下降"),
        ("华普静态回本周期（年）", base_pb_y if base_pb_y is not None else "不适用", rec_pb_y if rec_pb_y is not None else "不适用", (rec_pb_y - base_pb_y) if (base_pb_y is not None and rec_pb_y is not None) else "不适用", "华普", "缩短" if (base_pb_y is not None and rec_pb_y is not None and rec_pb_y < base_pb_y) else "延长"),
        ("华普NPV（万元）", float(offer.get("baseline_npv_rmb", 0.0)) / 10_000.0, float(offer["npv_project_rmb"]) / 10_000.0, (float(offer["npv_project_rmb"]) - float(offer.get("baseline_npv_rmb", 0.0))) / 10_000.0, "华普", "提升"),
        ("合同定位", "能源托管", f"AI-LaaS（{OPEX_MODE_CN.get(str(offer['opex_mode']), str(offer['opex_mode']))}）", "重构", "双方", "重构"),
    ]
    df = pd.DataFrame(rows, columns=["结论", "原能源托管", "AI-LaaS推荐方案", "变化", "影响对象", "判断"])
    df["备注"] = ""
    return df


def _select_representative_points(frontier: pd.DataFrame, n: int = 8) -> pd.DataFrame:
    if frontier.empty:
        return frontier
    # Pick points across client_gap quantiles.
    frontier = frontier.sort_values("client_gap_rmb").reset_index(drop=True)
    if len(frontier) <= n:
        return frontier
    qs = np.linspace(0, 1, n)
    idx = sorted({int(round(q * (len(frontier) - 1))) for q in qs})
    return frontier.loc[idx].reset_index(drop=True)


def build_workbook() -> Path:
    # Defaults aligned with Streamlit page but export is deterministic.
    horizon_years = 10
    payback_constraint_months = 36
    discount_rate = 0.12

    fee_low = 0.35
    fee_high = 1.20
    fee_steps = 36

    last_four_year_fee_reduction_grid = [0.0, 400_000.0, 800_000.0, 1_200_000.0]
    upfront_grid = [0.0, 200_000.0, 500_000.0, 1_000_000.0]
    opex_modes = ["uniform_pct", "electricity_only_pct", "ai_plus_solar"]
    reduction_grid = [0.0, 0.10, 0.20, 0.30, 0.40, 0.60, 0.80, 0.85]

    client_value = ClientValueAssumptions(
        baseline_outage_hours_per_year=30.0,
        laas_guaranteed_outage_hours_per_year=5.0,
        outage_cost_rmb_per_hour=10_000.0,
        sla_credit_share_to_client=1.0,
        client_discount_rate_annual=0.12,
    )

    parsed = load_roadlight_all(ROOT / "data")
    baseline = build_baseline_energy_trust(parsed, analysis_years=horizon_years, discount_rate_annual=discount_rate)
    fee_grid = default_fee_grid_from_baseline(baseline, pct_low=fee_low, pct_high=fee_high, steps=fee_steps)

    env = grid_search_feasible_envelope(
        baseline,
        term_years=list(range(1, horizon_years + 1)),
        annual_fee_rmb_grid=fee_grid,
        last_four_year_fee_reduction_rmb_grid=last_four_year_fee_reduction_grid,
        upfront_rmb_grid=upfront_grid,
        ai_opex_reduction_grid=reduction_grid,
        discount_rate_annual=discount_rate,
        opex_modes=opex_modes,  # type: ignore[arg-type]
        client_value=client_value,
    )
    env_df = envelope_table(env)

    # Enforce payback constraint as in dashboard filtering
    env_df["payback_ok"] = env_df["payback_months"].apply(lambda x: isinstance(x, int) and int(x) <= int(payback_constraint_months))

    provider_df = env_df[(env_df["provider_feasible"]) & (env_df["payback_ok"])].copy()
    everyone_df = env_df[(env_df["feasible_everyone_better_off"]) & (env_df["payback_ok"])].copy()

    # Add baseline NPV for delta columns in summary
    env_df["baseline_npv_rmb"] = float(baseline.npv_project_rmb)

    # Pareto frontier on provider-feasible universe
    frontier = pareto_frontier(provider_df, x_col="client_gap_rmb", y_col="npv_project_rmb")
    rep = _select_representative_points(frontier, n=8)

    # Top-N diversified everyone-feasible offers
    recommended_df = rank_recommended_offers(everyone_df)
    topN = _top_n_diversified(recommended_df, n=9, min_per_mode=2, pct_cap=MAX_AI_OPEX_REDUCTION_PCT)
    best = None if recommended_df.empty else recommended_df.head(1).iloc[0].to_dict()
    best_result = None
    if best is not None:
        best_result = evaluate_laas_scenario(
            baseline,
            LaaSScenario(
                term_years=int(best["term_years"]),
                annual_service_fee_rmb=float(best["annual_service_fee_rmb"]),
                last_four_year_fee_reduction_rmb=float(best["last_four_year_fee_reduction_rmb"]),
                upfront_rmb=float(best["upfront_rmb"]),
                ai_opex_reduction_pct=float(best["ai_opex_reduction_pct"]),
                opex_mode=str(best["opex_mode"]),
            ),
            discount_rate_annual=discount_rate,
            client_value=client_value,
        )

    # Build workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1) Cover / Dashboard
    ws = wb.create_sheet("仪表盘", 0)
    _set_col_widths(ws, {1: 24, 2: 20, 3: 20, 4: 20, 5: 20, 6: 18, 7: 18, 8: 18, 9: 18, 10: 18})
    _header_bar(
        ws,
        "能源托管转LaaS可行方案总览",
        f"Roadlight项目 | 生成日期：{date.today().isoformat()} | 全部结果均可追溯至底层数据与显式假设",
    )

    r = 4
    r = _section(ws, r, "关键指标概览")
    kpis = [
        ("基线CAPEX（元）", float(baseline.capex_y0_rmb)),
        ("基线回本周期（月）", str(baseline.payback_months)),
        ("华普可行方案数量", int(provider_df.shape[0])),
        ("双方共赢方案数量", int(everyone_df.shape[0])),
    ]
    for i, (k, v) in enumerate(kpis):
        rr = r + (i // 2)
        cc = 1 + (i % 2) * 5
        ws.merge_cells(start_row=rr, start_column=cc, end_row=rr, end_column=cc + 1)
        ws.cell(rr, cc, k).font = SECTION_FONT
        ws.cell(rr, cc, k).alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=rr, start_column=cc + 2, end_row=rr, end_column=cc + 4)
        val_cell = ws.cell(rr, cc + 2, v)
        val_cell.font = Font(color="111827", bold=True, size=12)
        val_cell.alignment = Alignment(horizontal="left")
        if isinstance(v, (int, float)):
            val_cell.number_format = "#,##0"
    r += 3

    r = _section(ws, r, "推荐方案（客户节省+华普毛利提升+更快回本）", c1=1, c2=10)
    if best is None:
        ws.cell(r, 1, "在当前约束下未找到双方共赢方案。").font = BODY_FONT
        r += 2
    else:
        summary_rows = [
            ("合同年限（年）", int(best["term_years"])),
            ("前6年年服务费（元）", float(best["annual_service_fee_rmb"])),
            ("平均年服务费（元）", float(best["average_client_payment_rmb_per_year"])),
            ("后4年年降幅（元/年）", float(best["last_four_year_fee_reduction_rmb"])),
            ("首期款（元）", float(best["upfront_rmb"])),
            ("降本模式", OPEX_MODE_CN.get(str(best["opex_mode"]), str(best["opex_mode"]))),
            ("AI降本比例", float(best["ai_opex_reduction_pct"])),
            ("回本周期（月）", int(best["payback_months"])),
            ("较基线回本改善（月）", float(best["payback_improvement_months"])),
            ("客户最少年节省（元）", float(best["min_client_savings_rmb_per_year"])),
            ("华普最少年毛利提升（元）", float(best["min_provider_gross_profit_uplift_rmb_per_year"])),
            ("华普项目NPV（元）", float(best["npv_project_rmb"])),
            ("客户价值缺口PV（元）", float(best["client_gap_rmb"])),
        ]
        for i, (k, v) in enumerate(summary_rows):
            ws.cell(r + i, 1, k).font = BODY_FONT
            ws.cell(r + i, 2, v).font = Font(color="111827", bold=True, size=10)
            ws.cell(r + i, 2).number_format = "#,##0" if isinstance(v, (int, float)) else "General"
        r += len(summary_rows) + 1

    # Screenshot-style comparison table (baseline vs best)
    if best is not None:
        best["gp_year1_rmb"] = float(best_result.provider_accounting_gross_profit_rmb_y.get(1)) if best_result is not None else 0.0
        best["net_client_payment_y1_rmb"] = float(best_result.client_payment_rmb_y.get(1)) if best_result is not None else float(best["annual_service_fee_rmb"])
        best["average_client_payment_rmb_per_year"] = float(best_result.average_client_payment_rmb_per_year) if best_result is not None else float(best["annual_service_fee_rmb"])
        best["baseline_npv_rmb"] = float(baseline.npv_project_rmb)
        r = _section(ws, r, "基线方案 vs 推荐方案对比", c1=1, c2=10)
        summ = _summary_block(
            baseline_fee_rmb_y1=float(baseline.revenue_rmb_y.get(1)),
            baseline_gp_rmb_y1=float(baseline.accounting_gross_profit_rmb_y.get(1)),
            baseline_payback_months=baseline.payback_months,
            offer=best,
        )
        end_r, end_c = _write_table(ws, r, 1, summ)
        r = end_r + 1
        if best_result is not None:
            r = _section(ws, r, "能源托管 vs LaaS 年度现金流对比（元/年）", c1=1, c2=10)
            best_yearly = _localize_simple_comparison(simple_cashflow_comparison_table(best_result, baseline))
            end_r, end_c = _write_table(ws, r, 1, best_yearly)
            r = end_r + 1

    # Client-facing: remove provider-NPV scatter chart; show a clean tier list instead.
    r = _section(ws, r, "10个共赢方案分层（多机制）", c1=1, c2=10)
    show = topN.copy()
    if not show.empty and "opex_mode" in show.columns:
        show["opex_mode"] = show["opex_mode"].map(lambda x: OPEX_MODE_CN.get(str(x), str(x)))
    keep = [
        "term_years",
        "annual_service_fee_rmb",
        "last_four_year_fee_reduction_rmb",
        "upfront_rmb",
        "opex_mode",
        "ai_opex_reduction_pct",
        "payback_months",
        "min_client_savings_rmb_per_year",
        "min_provider_gross_profit_uplift_rmb_per_year",
    ]
    keep = [c for c in keep if c in show.columns]
    show = show[keep].head(10)
    show = show.rename(
        columns={
            "term_years": "合同年限",
            "annual_service_fee_rmb": "前6年年服务费",
            "last_four_year_fee_reduction_rmb": "后4年年降幅",
            "upfront_rmb": "首期款",
            "opex_mode": "降本模式",
            "ai_opex_reduction_pct": "AI降本比例",
            "payback_months": "回本周期(月)",
            "min_client_savings_rmb_per_year": "客户最少年节省",
            "min_provider_gross_profit_uplift_rmb_per_year": "华普最少年毛利提升",
        }
    )
    end_r, end_c = _write_table(ws, r, 1, show)
    for rr in range(r + 1, end_r):
        for cc in range(1, 1 + len(show.columns)):
            header = ws.cell(r, cc).value
            if header in ("前6年年服务费", "后4年年降幅", "首期款", "客户最少年节省", "华普最少年毛利提升"):
                ws.cell(rr, cc).number_format = "#,##0"
            if header == "AI降本比例":
                ws.cell(rr, cc).number_format = "0.00%"

    # 2) Executive Summary
    ws2 = wb.create_sheet("执行摘要", 1)
    _set_col_widths(ws2, {1: 22, 2: 90, 3: 16})
    _header_bar(ws2, "执行摘要", "能源托管转LaaS的可行方案区间与关键权衡")
    rr = 4
    rr = _section(ws2, rr, "目标", c1=1, c2=3)
    ws2.cell(rr, 1, "目标").font = SECTION_FONT
    ws2.cell(rr, 2, "在可追溯、可解释的测算框架下，寻找能够让客户逐年少付、让华普逐年毛利更高、且较基线更快回本的LaaS商业条款。").alignment = WRAP_TOP
    rr += 2
    rr = _section(ws2, rr, "硬性约束", c1=1, c2=3)
    bullets = [
        "合同年限不超过10年",
        "华普静态现金回本不超过36个月",
        "华普回本必须快于基线方案",
        "华普每年会计毛利不低于基线方案",
        "客户每年支付不高于基线方案",
        "客户价值缺口需小于等于0，即LaaS支付现值减去基线支付现值再减去保障价值现值后不能为正",
    ]
    for i, btxt in enumerate(bullets):
        ws2.cell(rr + i, 2, f"- {btxt}").alignment = WRAP_TOP
    rr += len(bullets) + 1

    rr = _section(ws2, rr, "推荐逻辑", c1=1, c2=3)
    if best is None:
        ws2.cell(rr, 2, "当前假设下未找到双方共赢方案，建议重新审视客户价值假设、降本模式或商业结构。").alignment = WRAP_TOP
    else:
        ws2.cell(rr, 2, f"推荐逻辑优先最大化客户最少年节省与华普最少年毛利提升，其次优先更快回本与更高NPV。当前推荐方案为：合同{int(best['term_years'])}年，前6年年服务费{best['annual_service_fee_rmb']:,.0f}元，平均年服务费{best['average_client_payment_rmb_per_year']:,.0f}元，后4年年降幅{best['last_four_year_fee_reduction_rmb']:,.0f}元，首期款{best['upfront_rmb']:,.0f}元，降本模式为{OPEX_MODE_CN.get(str(best['opex_mode']), str(best['opex_mode']))}。").alignment = WRAP_TOP

    # 3) Assumptions
    ws3 = wb.create_sheet("假设输入", 2)
    _set_col_widths(ws3, {1: 34, 2: 22, 3: 70})
    _header_bar(ws3, "假设与输入", "所有假设均显式列示，所有输入均可追溯到底层数据")
    rr = 4
    rr = _section(ws3, rr, "搜索区间")
    rows = [
        ("测算期（年）", horizon_years, ""),
        ("回本约束（月）", payback_constraint_months, ""),
        ("折现率（年化）", discount_rate, "用于华普NPV测算，同时作为客户默认折现率"),
        ("服务费下限（基线比例）", fee_low, "服务费搜索区间 = 基线年托管费 × 下限至上限"),
        ("服务费上限（基线比例）", fee_high, ""),
        ("服务费步数", fee_steps, ""),
        ("后4年年降幅区间（元/年）", "、".join(f"{x:,.0f}" for x in last_four_year_fee_reduction_grid), f"仅作用于第7-10年，且尾部服务费底线为前6年服务费的{int(MIN_LAST_FOUR_YEAR_FEE_PCT * 100)}%"),
        ("首期款区间（元）", "、".join(f"{x:,.0f}" for x in upfront_grid), "客户于第0年支付，并计入客户成本现值"),
        ("降本模式", "、".join(OPEX_MODE_CN.get(x, x) for x in opex_modes), "AI+光伏模式假设电费降为0，未额外计入光伏CAPEX"),
        ("AI降本比例区间", "、".join(f"{int(x*100)}%" for x in reduction_grid), f"仅适用于比例型降本模式，比例上限为{int(MAX_AI_OPEX_REDUCTION_PCT * 100)}%"),
    ]
    for i, (k, v, note) in enumerate(rows):
        ws3.cell(rr + i, 1, k).font = BODY_FONT
        ws3.cell(rr + i, 2, v).font = Font(color="111827", bold=True, size=10)
        ws3.cell(rr + i, 3, note).alignment = WRAP_TOP
    rr += len(rows) + 2
    rr = _section(ws3, rr, "客户价值假设（SLA/风险转移）")
    for i, (k, v) in enumerate(asdict(client_value).items()):
        key_cn = {
            "baseline_outage_hours_per_year": "基线年停电小时数",
            "laas_guaranteed_outage_hours_per_year": "LaaS承诺年停电小时数",
            "outage_cost_rmb_per_hour": "单位停电损失（元/小时）",
            "sla_credit_share_to_client": "保障价值归客户比例",
            "client_discount_rate_annual": "客户折现率（年化）",
        }.get(k, k)
        ws3.cell(rr + i, 1, key_cn).font = BODY_FONT
        ws3.cell(rr + i, 2, v).font = Font(color="111827", bold=True, size=10)
    rr += len(asdict(client_value)) + 2
    rr = _section(ws3, rr, "数据来源说明")
    ws3.cell(rr, 1, "基线服务费来源").font = BODY_FONT
    ws3.cell(rr, 2, "income_analysis.csv 中“托管收入”行").alignment = WRAP_TOP
    ws3.cell(rr + 1, 1, "现金OPEX来源").font = BODY_FONT
    ws3.cell(rr + 1, 2, "opex.csv 中“改造后电费、职工薪酬费用、维修材料费、车辆费用、管理费用”").alignment = WRAP_TOP
    ws3.cell(rr + 2, 1, "CAPEX来源").font = BODY_FONT
    ws3.cell(rr + 2, 2, "capex.csv 中“总投资”").alignment = WRAP_TOP

    # 4) Outputs (Pareto frontier + recommended points)
    ws4 = wb.create_sheet("结果输出", 3)
    _set_col_widths(ws4, {1: 18, 2: 18, 3: 16, 4: 16, 5: 18, 6: 16, 7: 14, 8: 14, 9: 14, 10: 14})
    _header_bar(ws4, "结果输出", "方案全集、帕累托前沿与代表性推荐方案")
    rr = 4
    rr = _section(ws4, rr, "帕累托前沿（华普可行方案，最大化NPV、最小化客户价值缺口）")
    frontier_show = frontier[
        [
            "term_years",
            "annual_service_fee_rmb",
            "average_client_payment_rmb_per_year",
            "last_four_year_fee_reduction_rmb",
            "upfront_rmb",
            "opex_mode",
            "ai_opex_reduction_pct",
            "payback_months",
            "npv_project_rmb",
            "client_gap_rmb",
            "client_benefit_pass",
            "min_client_savings_rmb_per_year",
            "min_provider_gross_profit_uplift_rmb_per_year",
        ]
    ].copy()
    frontier_show = _localize_envelope_df(frontier_show)
    end_r, end_c = _write_table(ws4, rr, 1, frontier_show.head(200))
    rr = end_r + 2
    rr = _section(ws4, rr, "代表性前沿方案（用于管理层讨论）")
    rep_show = rep[
        [
            "term_years",
            "annual_service_fee_rmb",
            "average_client_payment_rmb_per_year",
            "last_four_year_fee_reduction_rmb",
            "upfront_rmb",
            "opex_mode",
            "ai_opex_reduction_pct",
            "payback_months",
            "npv_project_rmb",
            "client_gap_rmb",
            "min_client_savings_rmb_per_year",
            "min_provider_gross_profit_uplift_rmb_per_year",
        ]
    ].copy()
    rep_show = _localize_envelope_df(rep_show)
    _write_table(ws4, rr, 1, rep_show)

    # 5) Appendix / Backup
    ws5 = wb.create_sheet("附录测算", 4)
    _set_col_widths(ws5, {1: 14, 2: 18, 3: 14, 4: 14, 5: 16, 6: 16, 7: 16, 8: 16, 9: 16, 10: 18, 11: 18, 12: 18})
    _header_bar(ws5, "附录测算", "基线现金流、推荐方案对比及方案样本明细")
    rr = 4
    rr = _section(ws5, rr, "基线方案年度现金流（元/年）")
    base_df = _localize_baseline_cashflow(baseline_summary_table(baseline))
    _write_table(ws5, rr, 1, base_df)
    rr += len(base_df) + 4
    if best_result is not None:
        rr = _section(ws5, rr, "能源托管 vs LaaS 年度现金流对比（元/年）")
        best_df = _localize_simple_comparison(simple_cashflow_comparison_table(best_result, baseline))
        _write_table(ws5, rr, 1, best_df)
        rr += len(best_df) + 4
    rr = _section(ws5, rr, "方案全集样本（前2000行）")
    sample = _localize_envelope_df(env_df.head(2000))
    _write_table(ws5, rr, 1, sample)

    # Traceability JSON as text
    rr = 4
    ws6 = wb.create_sheet("追溯说明", 5)
    _set_col_widths(ws6, {1: 24, 2: 110})
    _header_bar(ws6, "追溯说明", "关键口径、来源文件与测算定义说明")
    bundle = provenance_bundle(baseline, best_result)
    rr = 4
    rr = _section(ws6, rr, "关键追溯说明", c1=1, c2=2)
    trace_rows = [
        ("基线服务费定义", "来自托管收入；对应客户支付，也是基线方案下华普收入。"),
        ("LaaS服务费定义", "按合同收费规则计算：前6年主费率，后4年按降幅调整，首期款按预付在全周期分摊。"),
        ("客户价值缺口定义", "LaaS支付现值减去基线支付现值，再减去保障价值现值；小于等于0代表客户整体更优。"),
        ("华普毛利定义", "毛利 = 收入 - 现金OPEX - 折旧。"),
        ("回本定义", "第0年含CAPEX与首期款，累计净现金流首次转正所对应的月份即为回本点。"),
        ("主要来源文件", "income_analysis.csv、opex.csv、loan.csv、capex.csv。"),
    ]
    for i, (k, v) in enumerate(trace_rows):
        ws6.cell(rr + i, 1, k).font = BODY_FONT
        ws6.cell(rr + i, 2, v).alignment = WRAP_TOP

    # Finishing touches
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A3"

    OUT_CN.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_CN)
    return OUT_CN


def main() -> None:
    out = build_workbook()
    print(f"Wrote Excel: {out}")


if __name__ == "__main__":
    main()

