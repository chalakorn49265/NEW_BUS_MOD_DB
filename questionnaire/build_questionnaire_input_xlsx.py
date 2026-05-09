"""Build questionnaire_01_input.xlsx: CN 填写 + EN English; subset via questionnaire_input_row_ids.txt."""

from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
ROW_IDS_FILE = ROOT / "questionnaire_input_row_ids.txt"

# Count-like or year-like integers (non-negative whole numbers).
WHOLE_NONNEG_IDS = frozenset({"A3", "A4", "D5", "J2", "H10"})

# kWh / price fields: allow N/A as text or non-negative number (custom formula per row).
NA_OR_NONNEG_DECIMAL_IDS = frozenset({"C4b", "C4c", "C4d"})


def _load_zh_map() -> dict[str, tuple[str, str]]:
    path = ROOT / "questionnaire_01_zh_map.py"
    spec = importlib.util.spec_from_file_location("questionnaire_01_zh_map", path)
    if spec is None or spec.loader is None:
        raise FileNotFoundError(path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.ZH_MAP


def _cell(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    if v is None:
        return ""
    return str(v).strip()


def _load_row_id_order() -> list[str] | None:
    if not ROW_IDS_FILE.is_file():
        return None
    lines = ROW_IDS_FILE.read_text(encoding="utf-8").splitlines()
    out: list[str] = []
    for line in lines:
        s = line.split("#", 1)[0].strip()
        if s:
            out.append(s)
    return out if out else None


def _master_by_id(ws_src) -> dict[str, dict[str, str]]:
    """qid -> dtype (col5), unit_en (col6), question_en (col4)."""
    by_id: dict[str, dict[str, str]] = {}
    for r in range(2, ws_src.max_row + 1):
        qid = _cell(ws_src, r, 1)
        if not qid:
            continue
        by_id[qid] = {
            "dtype": _cell(ws_src, r, 5),
            "unit_en": _cell(ws_src, r, 6) or _cell(ws_src, r, 5),
            "question_en": _cell(ws_src, r, 4),
        }
    return by_id


# Chinese / English list options. Keys must exist for every Categorical/Binary id used in subsets.
DROPDOWNS_CN: dict[str, list[str]] = {
    "A2": ["新建", "改造", "扩建", "纯运维升级"],
    "A6": ["高压钠（HPS）", "LED", "太阳能", "混合（请填%）"],
    "B8": ["是", "否"],
    "B9": ["上升", "下降", "持平"],
    "C4a": ["单一电价（平段）", "分时电价（峰谷）", "多段或其它（请说明）"],
    "C7": ["是", "否"],
    "E1": ["是", "否"],
    "G11": ["是", "否", "不确定"],
    "G12": ["是", "否"],
    "J1": ["是（全覆盖）", "是（仅部分）", "否", "未知"],
    "J3": ["节能分享", "保证节能量", "固定托管或管理费", "租赁或租用", "混合或其它（说明）"],
    "J4": ["业主或政府方", "ESCO或承包商", "分摊或转付", "不清楚"],
    "H7": ["是", "否", "未决"],
    "I2": ["接受", "拒绝", "未决"],
    "I5": ["移交", "回购", "续约", "拆回"],
    "K6": ["是", "否", "仅脱敏"],
    "C11": ["单表", "多表", "估算分摊", "混合（说明）"],
    "D9": ["内部", "外包", "混合（请填%）"],
    "F1": ["好", "一般", "差", "未知（备注）"],
}

DROPDOWNS_EN: dict[str, list[str]] = {
    "A2": ["New installation", "Retrofit", "Expansion", "O&M upgrade only"],
    "A6": ["HPS", "LED", "Solar", "Mix (specify %)"],
    "B8": ["Yes", "No"],
    "B9": ["Rising", "Falling", "Flat"],
    "C4a": ["Flat (single rate)", "Time-of-use (TOU)", "Multi-tier or other (describe)"],
    "C7": ["Yes", "No"],
    "E1": ["Yes", "No"],
    "G11": ["Yes", "No", "Uncertain"],
    "G12": ["Yes", "No"],
    "J1": ["Yes (whole portfolio)", "Partial (subset only)", "No", "Unknown"],
    "J3": [
        "Shared savings",
        "Guaranteed savings",
        "Fixed fee or management fee",
        "Lease or rental",
        "Mixed or other (describe)",
    ],
    "J4": [
        "Owner or public client pays",
        "ESCO or contractor pays",
        "Shared or pass-through",
        "Unclear",
    ],
    "H7": ["Yes", "No", "Undecided"],
    "I2": ["Accept", "Reject", "Undecided"],
    "I5": ["Transfer", "Buyout", "Renewal", "Retrieval"],
    "K6": ["Yes", "No", "De-identified only"],
    "C11": ["Single meter", "Multiple meters", "Estimated allocation", "Mixed (describe)"],
    "D9": ["In-house", "Outsourced", "Mixed (specify %)"],
    "F1": ["Good", "Fair", "Poor", "Unknown (notes)"],
}


def _sheet_id_order(ws_src) -> list[str]:
    order: list[str] = []
    seen: set[str] = set()
    for r in range(2, ws_src.max_row + 1):
        qid = _cell(ws_src, r, 1)
        if qid and qid not in seen:
            seen.add(qid)
            order.append(qid)
    return order


def _build_row_lists(
    ws_src,
    id_order: list[str] | None,
) -> tuple[list[tuple[str, str, str]], list[tuple[str, str, str]], list[str]]:
    zh = _load_zh_map()
    master = _master_by_id(ws_src)
    if id_order is None:
        id_order = _sheet_id_order(ws_src)
    cn_rows: list[tuple[str, str, str]] = []
    en_rows: list[tuple[str, str, str]] = []
    dtypes: list[str] = []
    for qid in id_order:
        if qid not in master:
            raise KeyError(f"Question id {qid!r} not found in questionnaire_01.xlsx Questions sheet")
        if qid not in zh:
            raise KeyError(f"Missing ZH_MAP entry for {qid!r} — update questionnaire_01_zh_map.py")
        qzh, unit_zh = zh[qid]
        m = master[qid]
        cn_rows.append((qid, qzh, unit_zh))
        en_rows.append((qid, m["question_en"], m["unit_en"]))
        dtypes.append(m["dtype"])
    return cn_rows, en_rows, dtypes


def _dropdown_order(data_rows: list[tuple[str, str, str]]) -> list[str]:
    seen: set[str] = set()
    order: list[str] = []
    for qid, _, _ in data_rows:
        if qid in DROPDOWNS_CN and qid not in seen:
            seen.add(qid)
            order.append(qid)
    return order


def _fill_list_sheet(sheet, dropdown_order: list[str], lang: str) -> dict[str, str]:
    src = DROPDOWNS_CN if lang == "cn" else DROPDOWNS_EN
    list_col_by_id: dict[str, str] = {}
    col_idx = 1
    for qid in dropdown_order:
        opts = src[qid]
        letter = get_column_letter(col_idx)
        list_col_by_id[qid] = letter
        sheet.cell(1, col_idx, qid)
        sheet.cell(1, col_idx).font = Font(bold=True)
        for i, opt in enumerate(opts, start=2):
            sheet.cell(i, col_idx, opt)
        sheet.column_dimensions[letter].width = 30 if lang == "en" else 24
        col_idx += 1
    return list_col_by_id


def _apply_list_validations(
    ws,
    data_rows: list[tuple[str, str, str]],
    list_col_by_id: dict[str, str],
    lists_sheet_name: str,
    lang: str,
) -> None:
    src = DROPDOWNS_CN if lang == "cn" else DROPDOWNS_EN
    for i, (qid, _q, _u) in enumerate(data_rows, start=2):
        if qid not in list_col_by_id:
            continue
        letter = list_col_by_id[qid]
        last_row = 1 + len(src[qid])
        formula = f"={lists_sheet_name}!${letter}$2:${letter}${last_row}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(i, 4))


def _apply_numeric_and_date(ws, data_rows: list[tuple[str, str, str]], dtypes: list[str]) -> None:
    for i, (qid, _q, _u) in enumerate(data_rows, start=2):
        dtype = dtypes[i - 2]
        col = f"D{i}"
        if dtype in ("Categorical", "Binary"):
            continue
        if dtype == "Qualitative":
            continue
        if dtype != "Quantitative":
            continue
        if qid in DROPDOWNS_CN:
            continue
        if qid == "INT6":
            dv = DataValidation(
                type="date",
                operator="between",
                formula1="1990-01-01",
                formula2="2050-12-31",
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add(ws.cell(i, 4))
            continue
        if qid in NA_OR_NONNEG_DECIMAL_IDS:
            f1 = f'=OR({col}="N/A",{col}="n/a",AND(ISNUMBER({col}),{col}>=0))'
            dv = DataValidation(type="custom", formula1=f1, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(ws.cell(i, 4))
            continue
        if qid in WHOLE_NONNEG_IDS:
            dv = DataValidation(
                type="whole",
                operator="greaterThanOrEqual",
                formula1="0",
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add(ws.cell(i, 4))
            continue
        dv = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(ws.cell(i, 4))


def _style_header(ws, headers: tuple[str, str, str, str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")


def _fill_data_sheet(ws, data_rows: list[tuple[str, str, str]]) -> None:
    for i, (qid, qtext, unit_text) in enumerate(data_rows, start=2):
        ws.cell(i, 1, qid)
        ws.cell(i, 2, qtext)
        ws.cell(i, 3, unit_text)
        ws.cell(i, 4, None)
        ws.cell(i, 3).fill = PatternFill("solid", fgColor="F3F3F3")


def build() -> Path:
    src_path = ROOT / "questionnaire_01.xlsx"
    if not src_path.is_file():
        raise FileNotFoundError(src_path)

    id_order = _load_row_id_order()
    wb_src = load_workbook(src_path, data_only=True)
    ws_src = wb_src["Questions"]

    cn_rows, en_rows, dtypes = _build_row_lists(ws_src, id_order)
    dropdown_order = _dropdown_order(cn_rows)

    for qid in dropdown_order:
        if qid not in DROPDOWNS_EN or qid not in DROPDOWNS_CN:
            raise ValueError(f"Missing dropdown list for {qid}")
        if len(DROPDOWNS_CN[qid]) != len(DROPDOWNS_EN[qid]):
            raise ValueError(f"CN/EN option count mismatch for {qid}")

    for i, qid in enumerate(r[0] for r in cn_rows):
        dt = dtypes[i]
        if dt in ("Categorical", "Binary") and qid not in DROPDOWNS_CN:
            raise ValueError(f"Categorical/Binary id {qid!r} needs DROPDOWNS_CN/EN entries")

    out_path = ROOT / "questionnaire_01_input.xlsx"
    wb = Workbook()

    ws_cn = wb.active
    ws_cn.title = "填写"
    _style_header(ws_cn, ("#", "问题", "单位", "回答"))
    _fill_data_sheet(ws_cn, cn_rows)

    ws_en = wb.create_sheet("English")
    _style_header(ws_en, ("#", "Question", "Unit", "Answer"))
    _fill_data_sheet(ws_en, en_rows)

    lists_cn = wb.create_sheet("Lists")
    lists_cn.sheet_state = "hidden"
    lists_en = wb.create_sheet("Lists_EN")
    lists_en.sheet_state = "hidden"

    col_cn = _fill_list_sheet(lists_cn, dropdown_order, "cn")
    col_en = _fill_list_sheet(lists_en, dropdown_order, "en")

    for ws in (ws_cn, ws_en):
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 68 if ws.title == "English" else 62
        ws.column_dimensions["C"].width = 28 if ws.title == "English" else 24
        ws.column_dimensions["D"].width = 40 if ws.title == "English" else 36

    _apply_list_validations(ws_cn, cn_rows, col_cn, "Lists", "cn")
    _apply_list_validations(ws_en, en_rows, col_en, "Lists_EN", "en")

    _apply_numeric_and_date(ws_cn, cn_rows, dtypes)
    _apply_numeric_and_date(ws_en, en_rows, dtypes)

    wb.save(out_path)
    return out_path


def main() -> None:
    p = build()
    print("Wrote", p)


if __name__ == "__main__":
    main()
