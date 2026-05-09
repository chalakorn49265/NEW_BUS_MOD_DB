"""Build feishu_form_simple.xlsx / .csv from embedded field spec (no runtime deps beyond openpyxl)."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# (order, type_zh, title_zh, title_en, hint, required, options, json_path)
FIELDS: list[tuple] = [
    (1, "单行文本", "项目编号", "Project ID", "内部机会/项目代码", "Y", "", "identity.project_id"),
    (2, "单行文本", "客户名称", "Client name", "客户法定或常用名称", "Y", "", "identity.client_name"),
    (3, "单行文本", "国家/地区", "Country", "如：中国 / Mozambique", "Y", "", "identity.country"),
    (4, "单行文本", "币种（ISO）", "Currency (ISO 4217)", "全部金额统一币种，如 CNY、USD、MZN", "Y", "", "capex_triplet.currency"),
    (5, "单行文本", "基准年度/期间", "Baseline period", "如：FY2024、自然年2024", "N", "", "baseline_period"),
    (6, "日期", "填写日期", "Submission date", "拜访或提交当日", "Y", "", "identity.submitted_at"),
    (7, "单行文本", "填写人", "Submitted by", "姓名 + 角色", "N", "", "identity.submitted_by"),
    (8, "单选题", "项目类型", "Project type", "", "Y", "新建|改造|扩建|纯运维升级", "project_type"),
    (9, "数字", "灯具数量（盏）", "Number of lights", "计划改造或存量相关盏数", "Y", "", "scale.number_of_lights"),
    (10, "数字", "现状年电费（本币/年）", "Annual electricity cost", "基准期内可归因于道路照明的电费", "N", "", "→ baseline electricity"),
    (11, "数字", "现状年运维费（本币/年）", "Annual O&M cost", "人工+材料+外包运维等", "N", "", "→ baseline opex"),
    (12, "数字", "设备更新/采购预算（本币/年）", "Capex / replacement budget", "设备类资本性更新，若单独列账", "N", "", "→ baseline capex budget"),
    (13, "数字", "年总支出（如仅知一项可只填此）", "Total annual lighting spend", "若只掌握总包数字，可只填此项", "N", "", "→ B1 total"),
    (14, "数字", "年用电量（kWh，如可知）", "Annual kWh", "近一年或基准年", "N", "", "energy_baseline_kwh"),
    (15, "数字", "每晚开灯小时数", "Hours on per night", "可写平均；有季节差异请备注", "N", "", "operating_hours_night"),
    (16, "单选题", "是否已有调光/控制", "Dimming / control", "", "N", "无|是|部分", "existing_control"),
    (17, "多行文本", "现状灯型与占比", "Incumbent fixture mix", "如：高压钠 60% + 普通 LED 40%", "N", "", "fixture_mix"),
    (
        18,
        "单选题",
        "可接受合同年限",
        "Acceptable term (years)",
        "选最接近的一档",
        "N",
        "≤3年|4–7年|8–12年|12年以上|未决",
        "commercial_laas.term_years",
    ),
    (19, "单选题", "资产是否必须归客户", "Asset ownership must be client", "", "N", "是|否|未决", "ownership_preference"),
    (20, "多行文本", "付款/审批主体", "Paying & approval entity", "如：财政/城管/路灯所/其他", "N", "", "payer_entity"),
    (21, "多行文本", "其他必要说明", "Other notes", "与金额口径、拆表、范围边界相关", "N", "", "notes"),
]


def _fill_sheet(ws) -> None:
    headers = [
        "序号",
        "题型",
        "题目标题（中文）",
        "Title (EN)",
        "填写说明",
        "必填",
        "选项（单选/多选，用 | 分隔）",
        "映射（内部JSON路径/备注）",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
    for row_i, row in enumerate(FIELDS, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row_i, c, val if val is not None else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["H"].width = 28
    for r in range(1, ws.max_row + 1):
        for c in (5, 6, 8):
            if ws.cell(r, c).value:
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    root = Path(__file__).resolve().parent
    xlsx_path = root / "feishu_form_simple.xlsx"
    csv_path = root / "feishu_form_simple.csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "Form_fields"
    _fill_sheet(ws)

    inst = wb.create_sheet("使用说明", 1)
    inst["A1"] = "飞书问卷 / 多维表格 — 使用说明"
    inst["A1"].font = Font(bold=True, size=12)
    inst["A3"] = (
        "1. 飞书「问卷」目前通常需逐题创建；本表供复制「题目标题」与「题型」到表单设计器，"
        "或先在本表内确认字段再手工录入。"
    )
    inst["A4"] = (
        "2. 若使用「多维表格」作收集：可新建表，将本目录下 feishu_form_simple.csv 导入为数据表，"
        "再按列类型改为文本/数字/单选等；或仅把本表当字段清单对照创建字段。"
    )
    inst["A5"] = (
        "3. 金额题请与「币种」列一致；与 B1/B2–B4 拆表关系见完整版问卷 Definitions 表或 docs/feishu_field_mapping.md。"
    )
    inst["A6"] = "4. 带「→」的映射为业务备注，用于后续 intake 映射到 project_capex_pack，非飞书内置字段。"
    inst.column_dimensions["A"].width = 100
    for r in range(3, 7):
        inst.cell(r, 1).alignment = Alignment(wrap_text=True)

    wb.save(xlsx_path)

    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "order",
                "type_zh",
                "title_zh",
                "title_en",
                "hint",
                "required",
                "options",
                "json_path",
            ]
        )
        w.writerows(FIELDS)

    print("Wrote", xlsx_path)
    print("Wrote", csv_path)


if __name__ == "__main__":
    main()
