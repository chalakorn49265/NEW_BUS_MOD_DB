"""Build questionnaire_01_3col.{csv,xlsx} (EN) and questionnaire_01_3col_zh.{csv,xlsx} (ZH) from questionnaire/questionnaire_01.xlsx."""

from __future__ import annotations

import csv
import importlib.util
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def _cell(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    if v is None:
        return ""
    return str(v).strip()


def _load_zh_map(script_dir: Path) -> dict[str, tuple[str, str]]:
    path = script_dir / "questionnaire_01_zh_map.py"
    spec = importlib.util.spec_from_file_location("questionnaire_01_zh_map", path)
    if spec is None or spec.loader is None:
        raise FileNotFoundError(path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.ZH_MAP


def extract_raw(workbook_path: Path) -> list[tuple[str, str, str, str]]:
    """Rows: (qid, question_en, unit_en, answer)."""
    wb = load_workbook(workbook_path, data_only=True)
    if "Questions" not in wb.sheetnames:
        raise ValueError(f"No Questions sheet in {workbook_path}")
    ws = wb["Questions"]
    rows_out: list[tuple[str, str, str, str]] = []
    for r in range(2, ws.max_row + 1):
        qid = _cell(ws, r, 1)
        qtext = _cell(ws, r, 4)
        dtype = _cell(ws, r, 5)
        unit_fmt = _cell(ws, r, 6)
        ans = _cell(ws, r, 8)
        if not qid and not qtext:
            continue
        unit = unit_fmt if unit_fmt else dtype
        rows_out.append((qid, qtext, unit, ans))
    return rows_out


def rows_en(raw: list[tuple[str, str, str, str]]) -> list[tuple[str, str, str]]:
    out: list[tuple[str, str, str]] = []
    for qid, qtext, unit, ans in raw:
        if qid and qtext:
            question = f"{qid} — {qtext}"
        elif qtext:
            question = qtext
        else:
            question = qid
        out.append((question, unit, ans))
    return out


def rows_zh(
    raw: list[tuple[str, str, str, str]],
    zh_map: dict[str, tuple[str, str]],
) -> list[tuple[str, str, str]]:
    out: list[tuple[str, str, str]] = []
    for qid, _qtext, _unit, ans in raw:
        if qid not in zh_map:
            raise KeyError(f"Missing Chinese mapping for question id: {qid!r}")
        qz, uz = zh_map[qid]
        question = f"{qid} — {qz}" if qid else qz
        out.append((question, uz, ans))
    return out


def write_outputs(
    rows: list[tuple[str, str, str]],
    csv_path: Path,
    xlsx_path: Path,
    *,
    headers: tuple[str, str, str],
) -> None:
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(list(headers))
        w.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
    for r, (q, u, a) in enumerate(rows, start=2):
        ws.cell(r, 1, q)
        ws.cell(r, 2, u)
        ws.cell(r, 3, a)
    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 36
    wb.save(xlsx_path)


def main() -> None:
    root = Path(__file__).resolve().parent
    src = root / "questionnaire_01.xlsx"
    if not src.is_file():
        raise SystemExit(f"Missing source workbook: {src}")
    raw = extract_raw(src)
    zh_map = _load_zh_map(root)

    en_rows = rows_en(raw)
    zh_rows = rows_zh(raw, zh_map)

    write_outputs(
        en_rows,
        root / "questionnaire_01_3col.csv",
        root / "questionnaire_01_3col.xlsx",
        headers=("question", "unit", "answer"),
    )
    write_outputs(
        zh_rows,
        root / "questionnaire_01_3col_zh.csv",
        root / "questionnaire_01_3col_zh.xlsx",
        headers=("问题", "单位", "回答"),
    )
    print(f"Wrote {len(en_rows)} rows EN: questionnaire_01_3col.csv / .xlsx")
    print(f"Wrote {len(zh_rows)} rows ZH: questionnaire_01_3col_zh.csv / .xlsx")


if __name__ == "__main__":
    main()
