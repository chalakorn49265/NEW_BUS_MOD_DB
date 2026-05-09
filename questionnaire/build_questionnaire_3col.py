"""Emit questionnaire_simple_3col.{csv,xlsx} with columns: question, unit, answer."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# (question, unit) — answer column left blank for fill-in
ROWS: list[tuple[str, str]] = [
    ("项目编号（内部机会/项目代码）", "文本"),
    ("客户名称", "文本"),
    ("国家/地区", "文本"),
    ("币种（ISO 4217，如 CNY / USD / MZN）", "代码"),
    ("基准年度/期间（如 FY2024、自然年2024）", "文本"),
    ("填写日期", "YYYY-MM-DD"),
    ("填写人（姓名 + 角色）", "文本"),
    ("项目类型（新建 / 改造 / 扩建 / 纯运维升级 — 选一）", "类别"),
    ("灯具数量", "盏"),
    ("现状年电费（可归因于道路照明）", "本币/年"),
    ("现状年运维费（人工 + 材料 + 外包等）", "本币/年"),
    ("设备更新或采购预算（资本性，若单独列账）", "本币/年"),
    ("年总照明支出（若只知道总数可只填此项）", "本币/年"),
    ("年用电量（近一年或基准年）", "kWh/年"),
    ("每晚开灯小时数（平均；有季节差异可在备注说明）", "小时/晚"),
    ("是否已有调光或智能控制（无 / 是 / 部分）", "类别"),
    ("现状灯型与占比（如：高压钠 60% + LED 40%）", "文本"),
    ("可接受合同年限（≤3年 / 4–7年 / 8–12年 / 12年以上 / 未决）", "类别"),
    ("资产是否必须归客户（是 / 否 / 未决）", "类别"),
    ("付款或审批主体（如财政 / 城管 / 路灯所）", "文本"),
    ("其他必要说明（口径、拆表、范围边界）", "文本"),
]


def main() -> None:
    root = Path(__file__).resolve().parent
    csv_path = root / "questionnaire_simple_3col.csv"
    xlsx_path = root / "questionnaire_simple_3col.xlsx"

    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["question", "unit", "answer"])
        for q, u in ROWS:
            w.writerow([q, u, ""])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["question", "unit", "answer"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
    for r, (q, u) in enumerate(ROWS, start=2):
        ws.cell(r, 1, q)
        ws.cell(r, 2, u)
        ws.cell(r, 3, "")
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    wb.save(xlsx_path)

    print("Wrote", csv_path)
    print("Wrote", xlsx_path)


if __name__ == "__main__":
    main()
