from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "business_model_comparison" / "roadlight_four_models_owner_fixed_cn_yearly_diff_cn_backup.xlsx"
OUT = ROOT / "business_model_comparison" / "roadlight_four_models_owner_questionB_ownerNPV.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FONT = Font(color="1F3864", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header(ws, r: int, c1: int, c2: int, text: str) -> None:
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _add_mode_b_inputs(wb: openpyxl.Workbook) -> dict[str, str]:
    """
    Appends a Mode_B block to 假设 and returns a map of logical names -> cell address.
    We deliberately place this after row 30 to avoid disturbing existing references.
    """
    ws = wb["假设"]
    start = ws.max_row + 2  # currently 30, so start at 32

    _header(ws, start, 1, 9, "Mode_B（Question B）：业主支付不锁定，比较“甲方NPV(总支出)”谁最低")
    r = start + 1
    ws.cell(r, 1, "开关").font = SUBHEADER_FONT
    ws.cell(r, 2, "OwnerSpendMode").font = SUBHEADER_FONT
    ws.cell(r, 3, "Fixed / ModelBased").font = SUBHEADER_FONT
    ws.cell(r, 4, "说明").font = SUBHEADER_FONT
    ws.cell(r, 5, "选择 ModelBased 才会回答 Question B").alignment = WRAP
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    r += 1

    # Mode switch value cell
    ws.cell(r, 1, "OwnerSpendMode（Fixed / ModelBased）")
    ws.cell(r, 2, "ModelBased")
    ws.cell(r, 4, "Fixed=沿用B16锁定；ModelBased=按各模式计费结构算业主支出").alignment = WRAP
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    mode_cell = ws.cell(r, 2).coordinate
    r += 2

    ws.cell(r, 1, "折现率（甲方，年）").font = SUBHEADER_FONT
    ws.cell(r, 2, 0.12).number_format = "0.00%"
    owner_disc = ws.cell(r, 2).coordinate
    ws.cell(r, 4, "用于 Owner_NPV_Comparison").alignment = WRAP
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    r += 1

    ws.cell(r, 1, "合约期（年）").font = SUBHEADER_FONT
    ws.cell(r, 2, "=B6")
    term_years = ws.cell(r, 2).coordinate
    r += 2

    _header(ws, r, 1, 9, "Mode_B 关键参数（让四模式在甲方侧自然拉开差异）")
    r += 1

    # EMC shared-savings parameters
    ws.cell(r, 1, "EMC：甲方保留节省比例（OwnerRetainShare）").alignment = WRAP
    ws.cell(r, 2, 0.5).number_format = "0.00%"
    emc_owner_retain = ws.cell(r, 2).coordinate
    ws.cell(r, 4, "例：50%表示节省的一半归甲方；另一半用于支付服务方").alignment = WRAP
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    r += 1

    ws.cell(r, 1, "EMC：固定服务费（年，可选）").alignment = WRAP
    ws.cell(r, 2, 0)
    emc_fixed_fee = ws.cell(r, 2).coordinate
    r += 1

    # Custody fee (托管)
    ws.cell(r, 1, "托管：托管费（年，覆盖电费+运维）").alignment = WRAP
    ws.cell(r, 2, "=B16")  # default parity with old lock, but now editable
    custody_fee = ws.cell(r, 2).coordinate
    ws.cell(r, 4, "默认等于B16；你可改为报价/指数化费用").alignment = WRAP
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    r += 1

    # LaaS subscription
    ws.cell(r, 1, "LaaS：订阅费（年）").alignment = WRAP
    ws.cell(r, 2, "=B16")
    laas_fee = ws.cell(r, 2).coordinate
    r += 1

    ws.cell(r, 1, "LaaS：SLA罚则（年，默认0）").alignment = WRAP
    ws.cell(r, 2, 0)
    laas_penalty = ws.cell(r, 2).coordinate
    r += 1

    # Lease rate to compute rent from CAPEX
    ws.cell(r, 1, "租赁：隐含利率（年，用于算租金PMT）").alignment = WRAP
    ws.cell(r, 2, "=B18")
    lease_rate_cell = ws.cell(r, 2).coordinate
    r += 1

    ws.cell(r, 1, "租赁：租金（年，=PMT(隐含利率, 年, CAPEX)）").alignment = WRAP
    ws.cell(r, 2, f"=-PMT(${lease_rate_cell[0]}${lease_rate_cell[1:]},$B$6,$B$8)")
    lease_rent_cell = ws.cell(r, 2).coordinate
    ws.cell(r, 4, "默认按CAPEX与利率/期限生成；租赁下电费与运维通常仍由甲方承担").alignment = WRAP
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    r += 2

    for rr in range(start, r):
        for c in range(1, 10):
            ws.cell(rr, c).alignment = WRAP

    return {
        "mode": f"假设!{mode_cell}",
        "owner_disc": f"假设!{owner_disc}",
        "term_years": f"假设!{term_years}",
        "emc_owner_retain": f"假设!{emc_owner_retain}",
        "emc_fixed_fee": f"假设!{emc_fixed_fee}",
        "custody_fee": f"假设!{custody_fee}",
        "laas_fee": f"假设!{laas_fee}",
        "laas_penalty": f"假设!{laas_penalty}",
        "lease_rent": f"假设!{lease_rent_cell}",
    }


def _apply_mode_b_owner_outflow(wb: openpyxl.Workbook, refs: dict[str, str]) -> None:
    """
    Update each 模式 sheet's 甲方实际流出 (column F rows 6-15) to be:
      IF(Mode="Fixed", 假设!B16, ModelBasedOwnerPayment_t)
    where the model-based payment is defined per model.
    """
    mode = refs["mode"]

    # Common baseline/delivered components already exist in 假设:
    # B9 baseline electricity, B10 baseline O&M, B11 delivered electricity, B12 delivered O&M
    baseline_total = "('假设'!$B$9+'假设'!$B$10)"
    delivered_total = "('假设'!$B$11+'假设'!$B$12)"
    savings = f"({baseline_total}-{delivered_total})"

    # ModelBased owner payments
    emc_payment = f"({delivered_total}+{refs['emc_fixed_fee']}+{refs['emc_owner_retain']}*{savings})"
    custody_payment = f"({refs['custody_fee']})"
    laas_payment = f"({refs['laas_fee']}+{refs['laas_penalty']})"
    lease_payment = f"({refs['lease_rent']}+'假设'!$B$11+'假设'!$B$12)"  # lease + owner still pays elec+O&M

    mapping = {
        "EMC模式": emc_payment,
        "能源托管模式": custody_payment,
        "LaaS模式": laas_payment,
        "租赁模式": lease_payment,
    }

    for sheet, payment in mapping.items():
        ws = wb[sheet]
        for r in range(6, 16):  # owner block year 1..10 lives here
            # Column F is 甲方实际流出 in these templates
            ws.cell(r, 6).value = f"=IF({mode}=\"Fixed\",'假设'!$B$16,{payment})"


def _add_owner_npv_sheet(wb: openpyxl.Workbook, refs: dict[str, str]) -> None:
    name = "Owner_NPV_Comparison"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    _set_col_widths(ws, {1: 20, 2: 18, 3: 18, 4: 18, 5: 18, 6: 22, 7: 34})

    _header(ws, 1, 1, 7, "Question B：哪种模式让甲方付得更少？（甲方NPV(总支出) 最低）")
    ws.cell(3, 1, "折现率（甲方，年）").font = SUBHEADER_FONT
    ws.cell(3, 2, f"={refs['owner_disc']}")
    ws.cell(3, 4, "Mode").font = SUBHEADER_FONT
    ws.cell(3, 5, f"={refs['mode']}")
    ws.cell(4, 1, "Baseline（市电含电缆）年成本").font = SUBHEADER_FONT
    ws.cell(4, 2, "='假设'!$B$29")
    ws.cell(4, 4, "Baseline NPV(cost)").font = SUBHEADER_FONT
    ws.cell(4, 5, "=NPV($B$3,$B$11:$K$11)")
    ws.cell(5, 1, "说明").font = SUBHEADER_FONT
    ws.cell(
        5,
        2,
        "本页按各模式表中的“甲方实际流出（Mode_B公式）”计算NPV(cost)。ΔNPV>0 表示相对基线省钱。",
    ).alignment = WRAP
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)

    # Table header
    headers = ["模式", "Y1", "Y2", "Y3", "Y4", "…", "NPV(cost) and Delta vs baseline"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(7, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Pull owner outflows from each model sheet F6:F15
    models = [("EMC", "EMC模式"), ("能源托管", "能源托管模式"), ("LaaS", "LaaS模式"), ("租赁", "租赁模式")]
    # Baseline annual series row (for NPV range)
    ws.cell(10, 1, "Baseline annual cost series").font = SUBHEADER_FONT
    ws.cell(11, 1, "Baseline").font = SUBHEADER_FONT
    for i, col in enumerate(range(2, 12), start=1):  # B..K represent Y1..Y10
        ws.cell(11, col, f"='假设'!$B$29")

    row = 8
    for label, sheet in models:
        ws.cell(row, 1, label)
        ws.cell(row, 2, f"='{sheet}'!F6")
        ws.cell(row, 3, f"='{sheet}'!F7")
        ws.cell(row, 4, f"='{sheet}'!F8")
        ws.cell(row, 5, f"='{sheet}'!F9")
        ws.cell(row, 6, "…")
        ws.cell(row, 7, f"=NPV($B$3,'{sheet}'!F6:F15)&\" | Delta=\"&($E$4-NPV($B$3,'{sheet}'!F6:F15))")
        row += 1


def main() -> None:
    if not SRC.is_file():
        raise SystemExit(f"Missing source workbook: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=False)
    refs = _add_mode_b_inputs(wb)
    _apply_mode_b_owner_outflow(wb, refs)
    _add_owner_npv_sheet(wb, refs)
    wb.save(OUT)
    print(f"Wrote: {OUT}")


if __name__ == "__main__":
    main()

